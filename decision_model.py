#!/usr/bin/env python3
"""
Karar Modeli — Tek Model
=========================
Telegram verileri + piyasa verileri + teknik indikatörler (3 TF)
hep birlikte tek modelde eğitilir.

Kullanım:
    python decision_model.py
    python decision_model.py --target 4h
    python decision_model.py --output rapor.xlsx

Gereksinimler:
    pip install pandas numpy scikit-learn openpyxl
"""

import sqlite3
import logging
import argparse
import warnings
import numpy as np
import pandas as pd
from datetime import datetime, timezone
from pathlib import Path

from sklearn.ensemble import RandomForestClassifier, GradientBoostingClassifier
from sklearn.tree import DecisionTreeClassifier, export_text
from sklearn.model_selection import StratifiedKFold, cross_val_score, cross_val_predict
from sklearn.metrics import classification_report
from sklearn.preprocessing import LabelEncoder

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from feature_builder import load_full_dataset, FEATURE_COLUMNS as IND_FEATURE_COLUMNS

log = logging.getLogger(__name__)
warnings.filterwarnings("ignore", category=FutureWarning)

BT_DB_PATH = "backtest_results.db"
OUT_PATH   = "decision_analysis.xlsx"

STRONG_WIN_PCT = 1.5
WEAK_WIN_PCT   = 0.0

# ─────────────────────────────────────────────────────────────────
# ÖZELLİK GRUPLARI
# ─────────────────────────────────────────────────────────────────

# Telegram + piyasa + pre-signal
TELEGRAM_FEATURES = [
    "hour", "weekday",
    "fear_greed", "fng_momentum_7d", "fng_momentum_14d",
    "btc_pct_1h", "btc_pct_24h",
    "trend_score", "trend_momentum",
    "pre_pct_1h", "pre_pct_4h",
    "pre_volatility_1h", "pre_volatility_4h",
    "pre_volume_rel_1h", "pre_volume_rel_4h",
    "pre_efficiency_1h",
    "pre_taker_buy_ratio_1h",
    "pump_before_4h", "pump_before_24h",
]

# İndikatör feature'ları (3 TF × 10 indikatör = 30 feature)
# feature_builder.py'den import edilen IND_FEATURE_COLUMNS listesi

ALL_FEATURES = TELEGRAM_FEATURES + IND_FEATURE_COLUMNS

# Ablation study grupları
FEATURE_GROUPS = {
    "1_telegram_context": ["hour", "weekday"],
    "2_market": [
        "fear_greed", "fng_momentum_7d", "fng_momentum_14d",
        "btc_pct_1h", "btc_pct_24h",
        "trend_score", "trend_momentum",
    ],
    "3_pre_signal": [
        "pre_pct_1h", "pre_pct_4h",
        "pre_volatility_1h", "pre_volatility_4h",
        "pre_volume_rel_1h", "pre_volume_rel_4h",
        "pre_efficiency_1h",
        "pre_taker_buy_ratio_1h",
        "pump_before_4h", "pump_before_24h",
    ],
    "4_indicators_5m": [f"ind_{c}_5m" for c in [
        "rsi_14", "macd_histogram", "macd_cross",
        "bb_pctb", "bb_bandwidth", "bb_squeeze",
        "ema_alignment", "price_vs_ema200",
        "obv_slope", "volume_ratio",
    ]],
    "5_indicators_15m": [f"ind_{c}_15m" for c in [
        "rsi_14", "macd_histogram", "macd_cross",
        "bb_pctb", "bb_bandwidth", "bb_squeeze",
        "ema_alignment", "price_vs_ema200",
        "obv_slope", "volume_ratio",
    ]],
    "6_indicators_1h": [f"ind_{c}_1h" for c in [
        "rsi_14", "macd_histogram", "macd_cross",
        "bb_pctb", "bb_bandwidth", "bb_squeeze",
        "ema_alignment", "price_vs_ema200",
        "obv_slope", "volume_ratio",
    ]],
}


# ─────────────────────────────────────────────────────────────────
# HEDEF DEĞİŞKEN
# ─────────────────────────────────────────────────────────────────

def define_target(df: pd.DataFrame, horizon: str = "1h") -> pd.Series:
    pct_col = f"pct_{horizon}"
    if pct_col not in df.columns:
        raise ValueError(f"{pct_col} kolonu bulunamadı!")

    targets = []
    for _, row in df.iterrows():
        pct = row[pct_col]
        direction = row.get("direction", "NEUTRAL")

        if pd.isna(pct):
            targets.append(None)
            continue

        is_long = direction in ("LONG", "LONG_IMPL")
        is_short = direction == "SHORT"

        if is_long:
            effective_pct = pct
        elif is_short:
            effective_pct = -pct
        else:
            effective_pct = abs(pct)

        if effective_pct >= STRONG_WIN_PCT:
            targets.append("STRONG_WIN")
        elif effective_pct >= WEAK_WIN_PCT:
            targets.append("WEAK_WIN")
        else:
            targets.append("LOSS")

    return pd.Series(targets, index=df.index, name=f"target_{horizon}")


# ─────────────────────────────────────────────────────────────────
# VERİ HAZIRLAMA
# ─────────────────────────────────────────────────────────────────

def prepare_data(df: pd.DataFrame, target_col: str) -> tuple:
    df_model = df[df["direction"].isin(["LONG", "SHORT", "LONG_IMPL"])].copy()
    df_model = df_model.dropna(subset=[target_col])

    available = [f for f in ALL_FEATURES if f in df_model.columns]
    missing = [f for f in ALL_FEATURES if f not in df_model.columns]
    if missing:
        log.warning(f"  Eksik özellikler ({len(missing)}): {missing[:10]}...")

    X = df_model[available].copy()
    y = df_model[target_col].copy()

    for col in X.columns:
        if X[col].isna().any():
            med = X[col].median()
            X[col] = X[col].fillna(med if not pd.isna(med) else 0)

    log.info(f"[model] Veri: {len(X)} satır × {len(X.columns)} özellik")
    log.info(f"[model] Sınıf dağılımı: {y.value_counts().to_dict()}")

    return X, y, available, df_model


# ─────────────────────────────────────────────────────────────────
# MODEL EĞİTİMİ
# ─────────────────────────────────────────────────────────────────

def train_models(X, y):
    models = {
        "random_forest": RandomForestClassifier(
            n_estimators=200, max_depth=8, min_samples_leaf=10,
            class_weight="balanced", random_state=42, n_jobs=-1,
        ),
        "gradient_boost": GradientBoostingClassifier(
            n_estimators=150, max_depth=5, learning_rate=0.1,
            min_samples_leaf=10, random_state=42,
        ),
        "decision_tree": DecisionTreeClassifier(
            max_depth=5, min_samples_leaf=15,
            class_weight="balanced", random_state=42,
        ),
    }

    cv = StratifiedKFold(n_splits=5, shuffle=True, random_state=42)
    results = {}

    for name, model in models.items():
        f1_scores = cross_val_score(model, X, y, cv=cv, scoring="f1_weighted")
        acc_scores = cross_val_score(model, X, y, cv=cv, scoring="accuracy")
        model.fit(X, y)

        results[name] = {
            "model": model,
            "f1_mean": f1_scores.mean(),
            "f1_std": f1_scores.std(),
            "acc_mean": acc_scores.mean(),
            "acc_std": acc_scores.std(),
        }
        log.info(f"  {name}: F1={f1_scores.mean():.3f}±{f1_scores.std():.3f}  Acc={acc_scores.mean():.3f}")

    return results


# ─────────────────────────────────────────────────────────────────
# ABLATION STUDY
# ─────────────────────────────────────────────────────────────────

def ablation_study(X, y, available_features):
    cv = StratifiedKFold(n_splits=5, shuffle=True, random_state=42)
    results = {}
    cumulative = []
    prev_f1 = 0

    for group_name, features in FEATURE_GROUPS.items():
        group_available = [f for f in features if f in available_features]
        if not group_available:
            continue

        cumulative.extend(group_available)
        X_sub = X[cumulative].copy()

        model = RandomForestClassifier(
            n_estimators=100, max_depth=6, min_samples_leaf=10,
            class_weight="balanced", random_state=42, n_jobs=-1,
        )
        f1_scores = cross_val_score(model, X_sub, y, cv=cv, scoring="f1_weighted")
        current_f1 = f1_scores.mean()
        improvement = current_f1 - prev_f1

        results[group_name] = {
            "features_count": len(cumulative),
            "new_features": len(group_available),
            "f1_mean": current_f1,
            "f1_std": f1_scores.std(),
            "improvement": improvement,
        }
        log.info(f"  [ablation] {group_name}: F1={current_f1:.3f} (+{improvement:+.3f}, {len(cumulative)} feat)")
        prev_f1 = current_f1

    return results


# ─────────────────────────────────────────────────────────────────
# KARAR FONKSİYONU
# ─────────────────────────────────────────────────────────────────

def make_decision(predicted, proba_dict, max_proba):
    loss_prob = proba_dict.get("LOSS", 0)

    if predicted == "STRONG_WIN" and max_proba > 0.5:
        return "EXECUTE"
    elif predicted != "LOSS" and loss_prob < 0.4:
        return "CAUTION"
    else:
        return "SKIP"


# ─────────────────────────────────────────────────────────────────
# EXCEL RAPOR
# ─────────────────────────────────────────────────────────────────

def _style_header(ws, row, ncols):
    fill = PatternFill(start_color="2D3436", end_color="2D3436", fill_type="solid")
    font = Font(color="FFFFFF", bold=True, size=10)
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")


def _style_title(ws, row, text, ncols):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = Font(bold=True, size=14, color="2D3436")
    cell.alignment = Alignment(horizontal="left")


def generate_report(results, ablation, importances, df, target_col,
                    available_features, out_path):
    wb = Workbook()

    # ── Sheet 1: Model Karşılaştırma ────────────────────────
    ws1 = wb.active
    ws1.title = "Model Karşılaştırma"
    _style_title(ws1, 1, "Model Performans Karşılaştırması (Tek Model — Tüm Veriler)", 6)

    headers = ["Model", "F1 (Ort)", "F1 (Std)", "Accuracy (Ort)", "Accuracy (Std)", "Durum"]
    for c, h in enumerate(headers, 1):
        ws1.cell(row=3, column=c, value=h)
    _style_header(ws1, 3, len(headers))

    best_f1 = max(r["f1_mean"] for r in results.values())
    row = 4
    for name, r in results.items():
        ws1.cell(row=row, column=1, value=name)
        ws1.cell(row=row, column=2, value=round(r["f1_mean"], 4))
        ws1.cell(row=row, column=3, value=round(r["f1_std"], 4))
        ws1.cell(row=row, column=4, value=round(r["acc_mean"], 4))
        ws1.cell(row=row, column=5, value=round(r["acc_std"], 4))
        ws1.cell(row=row, column=6, value="EN İYİ" if r["f1_mean"] == best_f1 else "")
        if r["f1_mean"] == best_f1:
            ws1.cell(row=row, column=6).font = Font(bold=True, color="27AE60")
        row += 1

    for col in ["A","B","C","D","E","F"]:
        ws1.column_dimensions[col].width = 20

    # ── Sheet 2: Ablation Study ──────────────────────────────
    ws2 = wb.create_sheet("Ablation Study")
    _style_title(ws2, 1, "Özellik Grubu Katkı Analizi (Ablation Study)", 6)
    headers2 = ["Grup", "Yeni Özellik", "Toplam Özellik", "F1", "F1 Std", "İyileşme"]
    for c, h in enumerate(headers2, 1):
        ws2.cell(row=3, column=c, value=h)
    _style_header(ws2, 3, len(headers2))

    row = 4
    for group_name, a in ablation.items():
        ws2.cell(row=row, column=1, value=group_name)
        ws2.cell(row=row, column=2, value=a["new_features"])
        ws2.cell(row=row, column=3, value=a["features_count"])
        ws2.cell(row=row, column=4, value=round(a["f1_mean"], 4))
        ws2.cell(row=row, column=5, value=round(a["f1_std"], 4))
        cell = ws2.cell(row=row, column=6, value=round(a["improvement"], 4))
        if a["improvement"] > 0.01:
            cell.font = Font(color="27AE60", bold=True)
        elif a["improvement"] < -0.01:
            cell.font = Font(color="E74C3C", bold=True)
        row += 1

    for col in ["A","B","C","D","E","F"]:
        ws2.column_dimensions[col].width = 22

    # ── Sheet 3: Feature Importance ──────────────────────────
    ws3 = wb.create_sheet("Feature Importance")
    _style_title(ws3, 1, "Özellik Önem Sıralaması (En İyi Model)", 3)
    headers3 = ["Sıra", "Özellik", "Önem"]
    for c, h in enumerate(headers3, 1):
        ws3.cell(row=3, column=c, value=h)
    _style_header(ws3, 3, len(headers3))

    row = 4
    for i, (feat, imp) in enumerate(importances[:40], 1):
        ws3.cell(row=row, column=1, value=i)
        ws3.cell(row=row, column=2, value=feat)
        ws3.cell(row=row, column=3, value=round(imp, 5))
        if i <= 5:
            ws3.cell(row=row, column=2).font = Font(bold=True)
        row += 1

    ws3.column_dimensions["A"].width = 8
    ws3.column_dimensions["B"].width = 30
    ws3.column_dimensions["C"].width = 12

    # ── Sheet 4: Kanal Performansı ───────────────────────────
    ws4 = wb.create_sheet("Kanal Performansı")
    _style_title(ws4, 1, "Kanal Bazlı Sinyal Doğruluğu", 7)

    if target_col in df.columns:
        channel_stats = df.groupby("channel").agg(
            total=(target_col, "count"),
            strong_win=(target_col, lambda x: (x == "STRONG_WIN").sum()),
            weak_win=(target_col, lambda x: (x == "WEAK_WIN").sum()),
            loss=(target_col, lambda x: (x == "LOSS").sum()),
        ).reset_index()
        channel_stats["win_rate"] = (channel_stats["strong_win"] + channel_stats["weak_win"]) / channel_stats["total"] * 100
        channel_stats["strong_rate"] = channel_stats["strong_win"] / channel_stats["total"] * 100
        channel_stats = channel_stats.sort_values("win_rate", ascending=False)

        headers4 = ["Kanal", "Toplam", "Güçlü Kazanç", "Zayıf Kazanç", "Kayıp", "Win Rate %", "Güçlü %"]
        for c, h in enumerate(headers4, 1):
            ws4.cell(row=3, column=c, value=h)
        _style_header(ws4, 3, len(headers4))

        row = 4
        for _, ch in channel_stats.iterrows():
            ws4.cell(row=row, column=1, value=ch["channel"])
            ws4.cell(row=row, column=2, value=int(ch["total"]))
            ws4.cell(row=row, column=3, value=int(ch["strong_win"]))
            ws4.cell(row=row, column=4, value=int(ch["weak_win"]))
            ws4.cell(row=row, column=5, value=int(ch["loss"]))
            ws4.cell(row=row, column=6, value=round(ch["win_rate"], 1))
            ws4.cell(row=row, column=7, value=round(ch["strong_rate"], 1))
            row += 1

        for col in ["A","B","C","D","E","F","G"]:
            ws4.column_dimensions[col].width = 16

    # ── Sheet 5: Sinyal Kararları (Out-of-Fold) ─────────────
    ws5 = wb.create_sheet("Sinyal Kararları")
    _style_title(ws5, 1, "Her Sinyal İçin Model Kararı (Out-of-Fold)", 10)

    df_directed = df[df["direction"].isin(["LONG", "SHORT", "LONG_IMPL"])].copy()
    if not df_directed.empty and target_col in df_directed.columns:
        df_valid = df_directed.dropna(subset=[target_col])
        avail = [f for f in ALL_FEATURES if f in df_valid.columns]
        X_all = df_valid[avail].copy()
        y_all = df_valid[target_col].copy()

        for col in X_all.columns:
            if X_all[col].isna().any():
                X_all[col] = X_all[col].fillna(X_all[col].median() if not X_all[col].isna().all() else 0)

        # Out-of-fold predictions
        from sklearn.base import clone
        best_name = max(results, key=lambda k: results[k]["f1_mean"])
        best_model = results[best_name]["model"]
        cv = StratifiedKFold(n_splits=5, shuffle=True, random_state=42)

        oof_pred = cross_val_predict(clone(best_model), X_all, y_all, cv=cv)
        oof_proba = cross_val_predict(clone(best_model), X_all, y_all, cv=cv, method="predict_proba")
        classes = sorted(y_all.unique())

        headers5 = [
            "Sinyal TS", "Kanal", "Ticker", "Yön", "Gerçek",
            "Tahmin", "Güven %", "Karar", "pct_1h", "pct_4h"
        ]
        for c, h in enumerate(headers5, 1):
            ws5.cell(row=3, column=c, value=h)
        _style_header(ws5, 3, len(headers5))

        row = 4
        decisions_list = []
        for idx_pos, (df_idx, sig_row) in enumerate(df_valid.iterrows()):
            pred = oof_pred[idx_pos]
            proba = oof_proba[idx_pos]
            max_proba = max(proba)
            proba_dict = dict(zip(classes, proba))
            decision = make_decision(pred, proba_dict, max_proba)
            decisions_list.append(decision)

            ws5.cell(row=row, column=1, value=str(sig_row.get("signal_ts", ""))[:19])
            ws5.cell(row=row, column=2, value=sig_row.get("channel", ""))
            ws5.cell(row=row, column=3, value=sig_row.get("ticker", ""))
            ws5.cell(row=row, column=4, value=sig_row.get("direction", ""))
            ws5.cell(row=row, column=5, value=sig_row.get(target_col, ""))
            ws5.cell(row=row, column=6, value=pred)
            ws5.cell(row=row, column=7, value=round(max_proba * 100, 1))

            cell = ws5.cell(row=row, column=8, value=decision)
            if decision == "EXECUTE":
                cell.font = Font(color="27AE60", bold=True)
            elif decision == "SKIP":
                cell.font = Font(color="E74C3C", bold=True)
            else:
                cell.font = Font(color="F39C12", bold=True)

            pct_1h = sig_row.get("pct_1h")
            if pct_1h is not None and not pd.isna(pct_1h):
                ws5.cell(row=row, column=9, value=round(float(pct_1h), 2))
            pct_4h = sig_row.get("pct_4h")
            if pct_4h is not None and not pd.isna(pct_4h):
                ws5.cell(row=row, column=10, value=round(float(pct_4h), 2))

            row += 1
            if row > 2003:
                break

        for col in ["A","B","C","D","E","F","G","H","I","J"]:
            ws5.column_dimensions[col].width = 16

        # ── Sheet 6: Karar Başarı Oranları ───────────────────
        ws6 = wb.create_sheet("Karar Başarı Oranları")
        _style_title(ws6, 1, "EXECUTE / CAUTION / SKIP Başarı Analizi (Out-of-Fold)", 8)

        headers6 = ["Karar", "Sinyal Sayısı", "STRONG_WIN", "WEAK_WIN", "LOSS", "Win Rate %", "Güçlü %", "Ort. pct_1h"]
        for c, h in enumerate(headers6, 1):
            ws6.cell(row=3, column=c, value=h)
        _style_header(ws6, 3, len(headers6))

        df_eval = df_valid.copy()
        df_eval["decision_oof"] = decisions_list

        row = 4
        for decision_type in ["EXECUTE", "CAUTION", "SKIP"]:
            sub = df_eval[df_eval["decision_oof"] == decision_type]
            total = len(sub)
            if total == 0:
                continue
            sw = (sub[target_col] == "STRONG_WIN").sum()
            ww = (sub[target_col] == "WEAK_WIN").sum()
            loss = (sub[target_col] == "LOSS").sum()
            wr = (sw + ww) / total * 100
            sr = sw / total * 100
            avg_pct = sub["pct_1h"].mean() if "pct_1h" in sub.columns else 0

            ws6.cell(row=row, column=1, value=decision_type)
            ws6.cell(row=row, column=2, value=total)
            ws6.cell(row=row, column=3, value=sw)
            ws6.cell(row=row, column=4, value=ww)
            ws6.cell(row=row, column=5, value=loss)
            ws6.cell(row=row, column=6, value=round(wr, 1))
            ws6.cell(row=row, column=7, value=round(sr, 1))
            ws6.cell(row=row, column=8, value=round(float(avg_pct), 3) if not pd.isna(avg_pct) else 0)

            if decision_type == "EXECUTE":
                ws6.cell(row=row, column=1).font = Font(color="27AE60", bold=True)
            elif decision_type == "SKIP":
                ws6.cell(row=row, column=1).font = Font(color="E74C3C", bold=True)
            row += 1

        row += 1
        ws6.cell(row=row, column=1, value="GENEL BASELINE").font = Font(bold=True)
        total_all = len(df_eval)
        win_all = ((df_eval[target_col] == "STRONG_WIN") | (df_eval[target_col] == "WEAK_WIN")).sum()
        ws6.cell(row=row, column=2, value=total_all)
        ws6.cell(row=row, column=6, value=round(win_all / total_all * 100, 1))
        avg_all = df_eval["pct_1h"].mean() if "pct_1h" in df_eval.columns else 0
        ws6.cell(row=row, column=8, value=round(float(avg_all), 3) if not pd.isna(avg_all) else 0)

        for col in ["A","B","C","D","E","F","G","H"]:
            ws6.column_dimensions[col].width = 16

    # ── Sheet 7: Karar Kuralları ─────────────────────────────
    ws7 = wb.create_sheet("Karar Kuralları")
    _style_title(ws7, 1, "Decision Tree Karar Kuralları (Okunabilir)", 1)

    if "decision_tree" in results:
        dt_model = results["decision_tree"]["model"]
        rules = export_text(dt_model, feature_names=available_features, max_depth=5)
        row = 3
        for line in rules.split("\n"):
            ws7.cell(row=row, column=1, value=line)
            ws7.cell(row=row, column=1).font = Font(name="Consolas", size=9)
            row += 1
        ws7.column_dimensions["A"].width = 80

    wb.save(out_path)
    log.info(f"[rapor] Kaydedildi: {out_path}")


# ─────────────────────────────────────────────────────────────────
# ANA İŞLEM
# ─────────────────────────────────────────────────────────────────

def run_decision_analysis(bt_db: str = BT_DB_PATH,
                          target_horizon: str = "1h",
                          out_path: str = OUT_PATH):
    sep = "=" * 60
    log.info(sep)
    log.info(f"KARAR MODELİ — Hedef: win_{target_horizon}")
    log.info(f"Telegram + Piyasa + İndikatörler (5m/15m/1h) — TEK MODEL")
    log.info(sep)

    # 1. Veri yükle
    log.info("[1/5] Veri yükleniyor...")
    df = load_full_dataset(bt_db)
    if df.empty:
        log.error("Veri seti boş!")
        return

    # 2. Hedef değişken
    log.info("[2/5] Hedef değişken tanımlanıyor...")
    target_col = f"target_{target_horizon}"
    df[target_col] = define_target(df, target_horizon)

    # 3. Veri hazırlık
    log.info("[3/5] Veri hazırlanıyor...")
    X, y, available_features, df_model = prepare_data(df, target_col)

    if len(X) < 50:
        log.error(f"Yeterli veri yok: {len(X)} satır (min 50 gerekli)")
        return

    # 4. Model eğitimi
    log.info("[4/5] Modeller eğitiliyor...")
    results = train_models(X, y)

    best_name = max(results, key=lambda k: results[k]["f1_mean"])
    best_model = results[best_name]["model"]
    log.info(f"\nEn iyi model: {best_name} (F1={results[best_name]['f1_mean']:.3f})")

    # Feature importance
    importances = sorted(
        zip(available_features, best_model.feature_importances_),
        key=lambda x: x[1], reverse=True
    )
    log.info("\nTop 15 özellik:")
    for feat, imp in importances[:15]:
        log.info(f"  {feat:35s} {imp:.4f}")

    # Ablation study
    log.info("\n[5/5] Ablation study...")
    ablation = ablation_study(X, y, available_features)

    # Excel rapor
    generate_report(results, ablation, importances, df, target_col, available_features, out_path)

    log.info(f"\n{sep}")
    log.info("ANALİZ TAMAMLANDI")
    log.info(f"  En iyi model: {best_name}")
    log.info(f"  F1 Score: {results[best_name]['f1_mean']:.3f}")
    log.info(f"  Accuracy: {results[best_name]['acc_mean']:.3f}")
    log.info(f"  Toplam feature: {len(available_features)}")
    log.info(f"  Rapor: {out_path}")
    log.info(sep)

    return {
        "best_model": best_name,
        "f1": results[best_name]["f1_mean"],
        "accuracy": results[best_name]["acc_mean"],
    }


def main():
    parser = argparse.ArgumentParser(description="Karar Modeli — Tek Model")
    parser.add_argument("--bt-db", default=BT_DB_PATH, help="Backtest DB yolu")
    parser.add_argument("--target", default="1h", choices=["1h", "4h", "1d"],
                        help="Hedef zaman dilimi (varsayılan: 1h)")
    parser.add_argument("--output", default=OUT_PATH, help="Excel çıktı dosyası")
    parser.add_argument("-v", "--verbose", action="store_true")
    args = parser.parse_args()

    level = logging.DEBUG if args.verbose else logging.INFO
    logging.basicConfig(
        level=level,
        format="%(asctime)s  %(levelname)-7s  %(message)s",
        datefmt="%H:%M:%S",
    )

    run_decision_analysis(
        bt_db=args.bt_db,
        target_horizon=args.target,
        out_path=args.output,
    )


if __name__ == "__main__":
    main()
