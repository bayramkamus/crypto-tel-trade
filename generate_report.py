#!/usr/bin/env python3
"""
Telegram Sinyal Raporu Uretici
================================
pump_research.db -> pump_research_daily.xlsx

Kullanim:
    python generate_report.py
    python generate_report.py --db pump_research.db --out rapor.xlsx
"""

import re
import sqlite3
import argparse
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from excel_styles import (
    C, fill, bold_font, reg_font, center, thin_border,
    title_row, header_row, stripe_row, set_col_widths,
)
from scraping.ticker_parser import extract_ticker, extract_direction

# ─────────────────────────────────────────────────────────────────
# SABITLER
# ─────────────────────────────────────────────────────────────────

DB_PATH  = "pump_research.db"
OUT_PATH = "pump_research_daily.xlsx"

BINANCE_TICKERS = {
    "BTC","ETH","BNB","SOL","XRP","DOGE","ADA","AVAX","LINK","DOT",
    "UNI","AAVE","LTC","BCH","TRX","NEAR","TAO","OP","ARB","IMX",
    "PEPE","SHIB","DOGS","ENA","GRT","SNX","WLD","MORPHO","IOTA",
    "FLOW","ARK","BAT","LPT","PYTH","TIA","DYDX","CRO","ALICE",
    "DUSK","BIO","KAIA","TON","STORJ","XVG","FRAX","PNUT","MOODENG",
    "SKY","HYPE","ZRO","GRASS","WLFI","ETHFI","VVV","INIT","OM",
    "MANTRA","KERNEL","OPN","PARTI","GPS","SOON","RESOLV","ZORA",
    "MUBARAK","INJ","APT","SUI","FIL","ATOM","ALGO","VET","SAND",
    "MANA","AXS","GALA","CHZ","ENJ","THETA","FTM","HBAR","EGLD",
    "MATIC","KNC","BAL","CRV","SUSHI","YFI","COMP","MKR","LDO",
    "GMX","PERP","ACT","ARC","LIT","LYN","MIRA","NIGHT",
    "TOWNS","UB","BARD","AVNT","PUMP","COAI","EDEN","ETHW",
    "GIGGLE","GWEI","HUMA","JELLYJELLY","PIPPIN","TOWN","XPL",
    "ASTER","PENDLE","STG","RDNT","ACE",
}

# ─────────────────────────────────────────────────────────────────
# PARSER
# ─────────────────────────────────────────────────────────────────

def parse_messages(db_path: str) -> list:
    conn = sqlite3.connect(db_path)
    c = conn.cursor()
    c.execute("""
        SELECT m.id, m.channel_name, m.timestamp, m.message_text, ch.title
        FROM messages m
        LEFT JOIN channels ch ON m.channel_id = ch.channel_id
        WHERE m.message_text IS NOT NULL
        ORDER BY m.timestamp ASC
    """)
    raw = c.fetchall()
    conn.close()

    results = []
    for (rid, channel, ts_str, text, ch_title) in raw:
        ticker, direction, entry, tp, sl, leverage = _parse_signal(text)
        msg_type = _classify(text, ticker, direction, entry)

        news_tickers = []
        if msg_type == "news":
            news_tickers = _extract_news_tickers(text)
            if not ticker and news_tickers:
                ticker = news_tickers[0]

        results.append({
            "id":            rid,
            "channel":       channel,
            "channel_title": ch_title or channel,
            "date":          ts_str[:10],
            "time":          ts_str[11:16],
            "ticker":        ticker,
            "news_tickers":  ", ".join(news_tickers) if news_tickers else "",
            "direction":     direction,
            "leverage":      leverage,
            "entry":         entry,
            "tp":            tp,
            "sl":            sl,
            "msg_type":      msg_type,
            "text":          text,
        })
    return results


def _parse_signal(text: str):
    ticker    = extract_ticker(text)
    direction = extract_direction(text)

    entry_m = re.search(r"Entry[:\s]+([0-9][0-9. \-]+)", text, re.IGNORECASE)
    entry = entry_m.group(1).strip()[:40] if entry_m else ""

    sl_m = re.search(r"(?:Stop.Loss|StopLoss|SL)[:\s]+\$?([0-9.]+)", text, re.IGNORECASE)
    sl = sl_m.group(1) if sl_m else ""

    tp_m = re.search(r"(?:Target|TP)[s]?[:\s]+([0-9$., \-]+)", text, re.IGNORECASE)
    tp = tp_m.group(1).strip()[:70] if tp_m else ""

    lev_m = re.search(r"Leverage[:\s]+(\S+)", text, re.IGNORECASE)
    lev = lev_m.group(1).rstrip(",;") if lev_m else ""

    return ticker, direction, entry, tp, sl, lev


def _classify(text, ticker, direction, entry):
    if ticker and (entry or direction):
        return "signal"
    for t in BINANCE_TICKERS:
        if re.search(rf"\b{re.escape(t)}\b", text):
            return "news"
    return "other"


def _extract_news_tickers(text):
    found = []
    for t in BINANCE_TICKERS:
        if re.search(rf"\b{re.escape(t)}\b", text):
            found.append(t)
    priority = ["BTC","ETH","SOL","XRP","BNB","DOGE","ADA"]
    found.sort(key=lambda x: (priority.index(x) if x in priority else 99, x))
    return found


# ─────────────────────────────────────────────────────────────────
# GUNLUK OZET
# ─────────────────────────────────────────────────────────────────

def build_daily_summary(messages: list) -> list:
    by_date = defaultdict(list)
    for m in messages:
        by_date[m["date"]].append(m)

    rows = []
    for date in sorted(by_date.keys()):
        day_msgs = by_date[date]
        signals = [m for m in day_msgs if m["msg_type"] == "signal"]
        news    = [m for m in day_msgs if m["msg_type"] == "news"]
        longs   = [s for s in signals if s["direction"] == "LONG"]
        shorts  = [s for s in signals if s["direction"] == "SHORT"]

        all_tickers = [m["ticker"] for m in day_msgs if m["ticker"]]
        top_ticker = Counter(all_tickers).most_common(1)
        top_ticker = top_ticker[0][0] if top_ticker else ""

        channels = sorted(set(m["channel"] for m in day_msgs))

        rows.append({
            "date":           date,
            "weekday":        datetime.strptime(date, "%Y-%m-%d").strftime("%A"),
            "total_msgs":     len(day_msgs),
            "signal_count":   len(signals),
            "news_count":     len(news),
            "other_count":    len(day_msgs) - len(signals) - len(news),
            "long_count":     len(longs),
            "short_count":    len(shorts),
            "top_ticker":     top_ticker,
            "unique_tickers": len(set(m["ticker"] for m in day_msgs if m["ticker"])),
            "active_channels": len(channels),
            "channels":       ", ".join(channels),
        })
    return rows


# ─────────────────────────────────────────────────────────────────
# EXCEL YAZICI
# ─────────────────────────────────────────────────────────────────

def write_excel(messages: list, daily: list, out_path: str):
    wb = Workbook()
    _write_daily_summary(wb, daily, messages)
    _write_flat_table(wb, messages)
    _write_signals_only(wb, messages)
    wb.save(out_path)


def _write_daily_summary(wb, daily, messages):
    ws = wb.active
    ws.title = "📅 Gunluk Ozet"
    ws.sheet_view.showGridLines = False

    date_range = f"{daily[0]['date']} – {daily[-1]['date']}" if daily else ""
    title_row(ws,
              f"📅  GUNLUK SİNYAL OZETİ  ·  {len(daily)} gun  ·  {date_range}",
              C["dark"], 12, font_size=14, height=36)

    total_msgs  = sum(d["total_msgs"] for d in daily)
    total_sigs  = sum(d["signal_count"] for d in daily)
    total_news  = sum(d["news_count"] for d in daily)
    total_long  = sum(d["long_count"] for d in daily)
    total_short = sum(d["short_count"] for d in daily)

    kpis = [
        ("A", "Toplam Mesaj",   total_msgs,  C["blue"]),
        ("C", "Sinyal",         total_sigs,  C["green"]),
        ("E", "Haber (news)",   total_news,  C["orange"]),
        ("G", "Long / Short",   f"{total_long} / {total_short}", C["purple"]),
    ]
    for col, label, val, color in kpis:
        c2 = chr(ord(col) + 1)
        for r in [3, 4, 5]:
            ws.merge_cells(f"{col}{r}:{c2}{r}")
        lc = ws.cell(3, ord(col)-64, label)
        lc.font = bold_font(); lc.fill = fill(color); lc.alignment = center()
        vc = ws.cell(4, ord(col)-64, val)
        vc.font = Font(name="Arial", bold=True, size=18, color=color)
        vc.fill = fill(C["white"]); vc.alignment = center()
        ws.cell(5, ord(col)-64).fill = fill(color)
        ws.cell(5, ord(col)-64+1).fill = fill(color)
    for r in [3,4,5]:
        ws.row_dimensions[r].height = 26

    HDR = ["Tarih","Gun","Toplam\nMesaj","Sinyal","Haber\n(News)","Diger",
           "LONG","SHORT","En Cok\nGecen","Unique\nTicker","Aktif\nKanal","Kanallar"]
    header_row(ws, 7, HDR, C["dark"], h=30)
    set_col_widths(ws, [12,11,10,9,9,8,8,8,12,10,10,55])

    for i, d in enumerate(daily, start=8):
        stripe_row(ws, i, 12, i % 2 == 0)

        ws.cell(i, 1, d["date"]).font = reg_font(C["blue"])
        ws.cell(i, 2, d["weekday"][:3])
        ws.cell(i, 3, d["total_msgs"]).alignment = center()

        sc = ws.cell(i, 4, d["signal_count"]); sc.alignment = center()
        if d["signal_count"] > 0:
            sc.fill = fill(C["l_green"]); sc.font = bold_font(C["green"], 9)

        nc = ws.cell(i, 5, d["news_count"]); nc.alignment = center()
        if d["news_count"] > 0:
            nc.fill = fill(C["l_amber"]); nc.font = reg_font(C["orange"])

        ws.cell(i, 6, d["other_count"]).alignment = center()

        lc = ws.cell(i, 7, d["long_count"]); lc.alignment = center()
        if d["long_count"] > 0:
            lc.fill = fill(C["l_green"]); lc.font = bold_font(C["green"], 9)

        sc2 = ws.cell(i, 8, d["short_count"]); sc2.alignment = center()
        if d["short_count"] > 0:
            sc2.fill = fill(C["l_red"]); sc2.font = bold_font(C["red"], 9)

        tt = ws.cell(i, 9, d["top_ticker"]); tt.alignment = center()
        if d["top_ticker"]:
            tt.font = bold_font(C["purple"], 9); tt.fill = fill(C["l_purp"])

        ws.cell(i, 10, d["unique_tickers"]).alignment = center()
        ws.cell(i, 11, d["active_channels"]).alignment = center()
        ws.cell(i, 12, d["channels"]).font = reg_font(C["dim"], 8)

    # Toplam satiri
    n = len(daily) + 8
    ws.row_dimensions[n].height = 18
    stripe_row(ws, n, 12, False)
    ws.cell(n, 1, "TOPLAM").font = bold_font(C["dark"])
    for c_idx in [3,4,5,6,7,8]:
        ws.cell(n, c_idx, f"=SUM({get_column_letter(c_idx)}8:{get_column_letter(c_idx)}{n-1})")
        ws.cell(n, c_idx).alignment = center()
    for col in range(1, 13):
        ws.cell(n, col).fill = fill(C["mid"])
        ws.cell(n, col).border = thin_border()
        ws.cell(n, col).font = bold_font(C["dark"], 9)


def _write_flat_table(wb, messages):
    ws = wb.create_sheet("📋 Tum Mesajlar")
    ws.sheet_view.showGridLines = False

    sig_cnt  = sum(1 for m in messages if m["msg_type"] == "signal")
    news_cnt = sum(1 for m in messages if m["msg_type"] == "news")
    title_row(ws,
              f"📋  TUM MESAJLAR  ·  {len(messages)} kayit  ·  "
              f"{sig_cnt} sinyal  ·  {news_cnt} haber",
              C["blue"], 13)

    HDR = ["#","Tarih","Saat","Kanal","Tur","Ticker","Yon","Kaldirac",
           "Giris","TP","SL","Haber Ticker'lari","Mesaj Ozeti"]
    header_row(ws, 2, HDR, C["blue"])
    set_col_widths(ws, [5,11,7,24,9,10,9,9,22,35,12,22,60])

    TYPE_STYLE = {
        "signal": (C["l_green"],  C["green"],  "📡 Sinyal"),
        "news":   (C["l_amber"],  C["orange"], "📰 Haber"),
        "other":  (C["gray"],     C["dim"],    "—"),
    }

    for i, m in enumerate(messages, start=3):
        stripe_row(ws, i, 13, i % 2 == 0)

        ws.cell(i, 1, i-2).alignment = center()
        ws.cell(i, 2, m["date"]).font = reg_font(C["blue"])
        ws.cell(i, 3, m["time"])
        ws.cell(i, 4, f"@{m['channel']}").font = reg_font(C["purple"], 8)

        bg, fc, label = TYPE_STYLE.get(m["msg_type"], (C["white"], C["dark"], m["msg_type"]))
        tc = ws.cell(i, 5, label)
        tc.fill = fill(bg); tc.font = bold_font(fc, 8); tc.alignment = center()

        if m["ticker"]:
            tkc = ws.cell(i, 6, m["ticker"])
            tkc.font = bold_font(C["blue"], 9); tkc.alignment = center()

        if m["direction"]:
            dc = ws.cell(i, 7, m["direction"]); dc.alignment = center()
            if m["direction"] == "LONG":
                dc.fill = fill(C["l_green"]); dc.font = bold_font(C["green"], 9)
            else:
                dc.fill = fill(C["l_red"]); dc.font = bold_font(C["red"], 9)

        ws.cell(i, 8, m["leverage"]).alignment = center()
        ws.cell(i, 9, m["entry"])
        ws.cell(i, 10, m["tp"])
        ws.cell(i, 11, m["sl"])
        ws.cell(i, 12, m["news_tickers"]).font = reg_font(C["orange"], 8)
        ws.cell(i, 13, m["text"][:130].replace("\n", " ")).font = reg_font(C["dim"], 8)
        ws.row_dimensions[i].height = 13


def _write_signals_only(wb, messages):
    signals = [m for m in messages if m["msg_type"] == "signal"]
    ws = wb.create_sheet("📡 Sinyaller")
    ws.sheet_view.showGridLines = False

    title_row(ws, f"📡  ALIM/SATIM SİNYALLERİ  ·  {len(signals)} adet", C["green"], 11)

    HDR = ["#","Tarih","Saat","Kanal","Ticker","Yon","Kaldirac",
           "Giris Fiyati","Hedefler (TP)","Stop-Loss","Mesaj Ozeti"]
    header_row(ws, 2, HDR, C["green"])
    set_col_widths(ws, [5,11,7,24,12,9,10,22,38,12,60])

    for i, s in enumerate(signals, start=3):
        stripe_row(ws, i, 11, i % 2 == 0)

        ws.cell(i, 1, i-2).alignment = center()
        ws.cell(i, 2, s["date"]).font = reg_font(C["blue"])
        ws.cell(i, 3, s["time"])
        ws.cell(i, 4, f"@{s['channel']}").font = reg_font(C["purple"], 8)

        tc = ws.cell(i, 5, s["ticker"])
        tc.font = bold_font(C["blue"], 10); tc.alignment = center()

        if s["direction"]:
            dc = ws.cell(i, 6, s["direction"]); dc.alignment = center()
            if s["direction"] == "LONG":
                dc.fill = fill(C["l_green"]); dc.font = bold_font(C["green"], 9)
            else:
                dc.fill = fill(C["l_red"]); dc.font = bold_font(C["red"], 9)

        ws.cell(i, 7, s["leverage"]).alignment = center()
        ws.cell(i, 8, s["entry"])
        ws.cell(i, 9, s["tp"])
        ws.cell(i, 10, s["sl"])
        ws.cell(i, 11, s["text"][:130].replace("\n", " ")).font = reg_font(C["dim"], 8)
        ws.row_dimensions[i].height = 14


# ─────────────────────────────────────────────────────────────────
# RUN (programatik cagri icin)
# ─────────────────────────────────────────────────────────────────

def run(db: str = DB_PATH, out: str = OUT_PATH):
    db  = str(Path(db).resolve())
    out = str(Path(out).resolve())

    print(f"[1/3] Mesajlar okunuyor: {db}")
    messages = parse_messages(db)
    sig_cnt  = sum(1 for m in messages if m["msg_type"] == "signal")
    news_cnt = sum(1 for m in messages if m["msg_type"] == "news")
    print(f"      → {len(messages)} mesaj  |  {sig_cnt} sinyal  |  {news_cnt} haber")

    print("[2/3] Gunluk ozet hesaplaniyor...")
    daily = build_daily_summary(messages)
    print(f"      → {len(daily)} gun")

    print(f"[3/3] Excel yaziliyor: {out}")
    write_excel(messages, daily, out)
    print(f"      ✅ Tamamlandi: {out}")


# ─────────────────────────────────────────────────────────────────
# CLI
# ─────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Telegram sinyal raporu uretici")
    parser.add_argument("--db",  default=DB_PATH)
    parser.add_argument("--out", default=OUT_PATH)
    args = parser.parse_args()
    run(db=args.db, out=args.out)


if __name__ == "__main__":
    main()
