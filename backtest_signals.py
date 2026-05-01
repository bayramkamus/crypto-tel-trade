#!/usr/bin/env python3
"""
Sinyal Volatilite Backtester
================================
Her sinyal icin Binance'ten fiyat verisi ceker:
  +5dk / +30dk / +1sa / +4sa / +1gun hareketleri

Kullanim:
    python backtest_signals.py
    python backtest_signals.py --db pump_research.db --out backtest.xlsx --force

Gereksinimler:
    pip install requests openpyxl
"""

import re
import time
import sqlite3
import argparse
from datetime import datetime, timezone
from pathlib import Path
from collections import Counter, defaultdict

import requests
from openpyxl import Workbook
from openpyxl.styles import Font

from excel_styles import (
    C, fill, bold_font, reg_font, center,
    title_row, header_row, stripe_row, set_col_widths, pct_cell,
)
from scraping.ticker_parser import extract_ticker, extract_direction

# ─────────────────────────────────────────────────────────────────
# SABITLER
# ─────────────────────────────────────────────────────────────────

DB_PATH     = "pump_research.db"
BT_DB_PATH  = "backtest_results.db"
OUT_PATH    = "backtest_signals.xlsx"

BINANCE_SPOT    = "https://api.binance.com/api/v3/klines"
BINANCE_FUTURES = "https://fapi.binance.com/fapi/v1/klines"

PRE_SIGNAL_WINDOW_SPECS = [
    ("5m", 5),
    ("30m", 30),
    ("1h", 60),
    ("4h", 240),
]
PRE_SIGNAL_METRIC_SPECS = [
    ("pre_pct",              "REAL"),
    ("pre_range",            "REAL"),
    ("pre_volatility",       "REAL"),
    ("pre_volume_rel",       "REAL"),
    ("pre_trade_rel",        "REAL"),
    ("pre_efficiency",       "REAL"),
    ("pre_taker_buy_ratio",  "REAL"),   # pencere toplam taker-buy / toplam hacim
]
PRE_SIGNAL_COLUMNS = [
    f"{metric}_{window_label}"
    for metric, _ in PRE_SIGNAL_METRIC_SPECS
    for window_label, _ in PRE_SIGNAL_WINDOW_SPECS
]
PRE_SIGNAL_SQL_COLUMNS = ",\n".join(
    f"            {col:<20} REAL"
    for col in PRE_SIGNAL_COLUMNS
)

BT_VALUE_COLUMNS = [
    "message_db_id", "channel", "signal_ts", "ticker", "symbol", "market",
    "direction", "entry_raw", "price_signal",
    "price_5m", "price_30m", "price_1h", "price_4h", "price_1d",
    "pct_5m", "pct_30m", "pct_1h", "pct_4h", "pct_1d",
    "candle_open", "candle_high", "candle_low", "candle_close",
    "candle_volume", "trade_count",
    "candle_taker_buy_vol", "candle_taker_buy_ratio",   # yönlü hacim baskısı
    "prev_volume", "volume_ma_24", "volume_ma_48",
    # Sinyal öncesi pump uyarısı (Katman 1 — ML'e girmiyor, sadece tanımlayıcı)
    "pump_before_4h", "pump_before_24h",
    # BTC piyasa bağlamı (Katman 2 — ML feature)
    "btc_pct_1h", "btc_pct_24h",
    *PRE_SIGNAL_COLUMNS,
]


# ─────────────────────────────────────────────────────────────────
# VERİTABANI
# ─────────────────────────────────────────────────────────────────

def init_backtest_table(db_path: str = BT_DB_PATH):
    conn = sqlite3.connect(db_path)
    conn.execute(f"""
        CREATE TABLE IF NOT EXISTS signal_backtest (
            id              INTEGER PRIMARY KEY AUTOINCREMENT,
            message_db_id   INTEGER UNIQUE,
            channel         TEXT,
            signal_ts       TEXT,
            ticker          TEXT,
            symbol          TEXT,
            market          TEXT,
            direction       TEXT,
            entry_raw       TEXT,
            price_signal    REAL,
            price_5m        REAL,
            price_30m       REAL,
            price_1h        REAL,
            price_4h        REAL,
            price_1d        REAL,
            pct_5m          REAL,
            pct_30m         REAL,
            pct_1h          REAL,
            pct_4h          REAL,
            pct_1d          REAL,
            candle_open     REAL,
            candle_high     REAL,
            candle_low      REAL,
            candle_close    REAL,
            candle_volume   REAL,
            trade_count     INTEGER,
            prev_volume     REAL,
            volume_ma_24    REAL,
            volume_ma_48    REAL,
{PRE_SIGNAL_SQL_COLUMNS},
            fetched_at      TEXT
        )
    """)
    # Mevcut DB'lere yeni kolonlari ekle (migration)
    existing = {row[1] for row in conn.execute("PRAGMA table_info(signal_backtest)")}
    migration_cols = [
        ("candle_open",             "REAL"),
        ("candle_high",             "REAL"),
        ("candle_low",              "REAL"),
        ("candle_close",            "REAL"),
        ("candle_volume",           "REAL"),
        ("trade_count",             "INTEGER"),
        ("candle_taker_buy_vol",    "REAL"),
        ("candle_taker_buy_ratio",  "REAL"),
        ("prev_volume",             "REAL"),
        ("volume_ma_24",            "REAL"),
        ("volume_ma_48",            "REAL"),
    ]
    migration_cols.extend((col, "REAL") for col in PRE_SIGNAL_COLUMNS)
    # Sinyal öncesi pump + BTC bağlamı
    migration_cols.extend([
        ("pump_before_4h",  "REAL"),
        ("pump_before_24h", "REAL"),
        ("btc_pct_1h",      "REAL"),
        ("btc_pct_24h",     "REAL"),
    ])
    # backtest_trends.py tarafından doldurulan trend kolonları
    migration_cols.extend([("trend_score", "REAL"), ("trend_momentum", "REAL")])
    # backtest_context.py tarafından doldurulan FnG kolonları
    migration_cols.extend([("fear_greed", "INTEGER"), ("fear_greed_label", "TEXT")])
    for col, typedef in migration_cols:
        if col not in existing:
            conn.execute(f"ALTER TABLE signal_backtest ADD COLUMN {col} {typedef}")
    conn.commit()
    conn.close()


def load_cached(db_path: str = BT_DB_PATH) -> set:
    """
    Daha önce tamamlanmış (fetched_at dolu) satırların message_db_id'lerini döner.
    Bu set'te olan mesajlar tekrar işlenmez.
    """
    conn = sqlite3.connect(db_path)
    tables = {r[0] for r in conn.execute(
        "SELECT name FROM sqlite_master WHERE type='table'"
    )}
    if "signal_backtest" not in tables:
        conn.close()
        return set()
    result = {
        r[0] for r in conn.execute(
            "SELECT message_db_id FROM signal_backtest WHERE fetched_at IS NOT NULL"
        )
    }
    conn.close()
    return result


def save_backtest_row(db_path: str, row: dict):
    conn = sqlite3.connect(db_path)
    columns = BT_VALUE_COLUMNS + ["fetched_at"]
    placeholders = ",".join("?" for _ in columns)
    values = [row.get(col) for col in BT_VALUE_COLUMNS]
    values.append(datetime.now(timezone.utc).isoformat())
    conn.execute(
        f"""
        INSERT OR REPLACE INTO signal_backtest
        ({", ".join(columns)})
        VALUES ({placeholders})
        """,
        values,
    )
    conn.commit()
    conn.close()


# ─────────────────────────────────────────────────────────────────
# PARSER
# ─────────────────────────────────────────────────────────────────

def _last_processed_per_channel(bt_db: str = BT_DB_PATH) -> dict[str, str]:
    """
    backtest_results.db'deki HER KANAL icin en son islenmis sinyal zamanini doner.
    Donus: {channel_name: max_signal_ts_iso}

    Tablo yoksa ya da bossa bos dict doner. Bos dict, parse_signals icin
    "tum mesajlari oku" anlamina gelir.
    """
    try:
        conn = sqlite3.connect(bt_db)
        tables = {r[0] for r in conn.execute(
            "SELECT name FROM sqlite_master WHERE type='table'"
        )}
        if "signal_backtest" not in tables:
            conn.close()
            return {}
        rows = conn.execute("""
            SELECT channel, MAX(signal_ts)
            FROM signal_backtest
            WHERE fetched_at IS NOT NULL
              AND channel IS NOT NULL
            GROUP BY channel
        """).fetchall()
        conn.close()
        return {ch: ts for ch, ts in rows if ch and ts}
    except Exception:
        return {}


def _channels_in_messages(db_path: str) -> set[str]:
    """pump_research.db/messages tablosundaki tum DISTINCT kanal adlari."""
    try:
        conn = sqlite3.connect(db_path)
        rows = conn.execute("""
            SELECT DISTINCT channel_name
            FROM messages
            WHERE channel_name IS NOT NULL AND message_text IS NOT NULL
        """).fetchall()
        conn.close()
        return {r[0] for r in rows if r[0]}
    except Exception:
        return set()


def _detect_new_channels(db_path: str, bt_db: str) -> set[str]:
    """
    pump_research.db'de mesaji olan ama signal_backtest'te HENUZ kaydi olmayan
    kanallari doner. Yeni eklenen kanallar bu sete duser; pipeline onlar icin
    tum gecmis mesajlari islemeli.
    """
    msg_channels = _channels_in_messages(db_path)
    if not msg_channels:
        return set()
    seen = set(_last_processed_per_channel(bt_db).keys())
    return msg_channels - seen


# Geriye uyumluluk: eski global cutoff alan kod kalmasin diye yardimci shim.
# Artik dahili kullanim _last_processed_per_channel'a tasindi.
def _last_processed_ts(bt_db: str = BT_DB_PATH) -> str | None:
    """DEPRECATED — yeni kod _last_processed_per_channel kullansin."""
    per_channel = _last_processed_per_channel(bt_db)
    if not per_channel:
        return None
    return min(per_channel.values())  # en eski kanalin son islenme zamani


def parse_signals(db_path: str, bt_db: str = BT_DB_PATH,
                  force: bool = False) -> list[dict]:
    """
    Ticker içeren mesajları toplar.

    Iki fazli incremental mod (varsayilan):

      FAZ 1 — YENI KANAL TESPITI:
        signal_backtest'te kaydi olmayan kanallari tespit eder ve
        bu kanallarin TUM gecmis mesajlarini ceker. Boylece run_collector
        ile yeni eklenip 30 gunluk backfill yapilan kanalin verisi
        pipeline'a tam olarak girer.

      FAZ 2 — MEVCUT KANALLAR:
        Daha once islenmis kanallar icin kendi kanal-basina son islenmis
        signal_ts'inden sonraki mesajlari ceker.

      Iki fazin sonucu timestamp'e gore birlestirilip sirali doner.

    force=True ise tum kanal cutoff'lari atlanir, tum mesajlar islenir.

    Yön etiketleri:
      LONG       — açık LONG/BUY
      SHORT      — açık SHORT/SELL
      LONG_IMPL  — emoji/anahtar kelime ile örtülü boğa sinyali
      NEUTRAL    — ticker var ama yön sinyali yok
    """
    # ────────────────────────────────────────────────────────────
    # YOL A: --force veya hic islenmis sinyal yok → tam tarama
    # ────────────────────────────────────────────────────────────
    per_channel_cutoff = {} if force else _last_processed_per_channel(bt_db)

    if force or not per_channel_cutoff:
        conn = sqlite3.connect(db_path)
        c = conn.cursor()
        c.execute("""
            SELECT id, channel_name, timestamp, message_text
            FROM messages WHERE message_text IS NOT NULL
            ORDER BY timestamp ASC
        """)
        rows = c.fetchall()
        conn.close()
        if force:
            print("  📅 Tam tarama (--force): pump_research.db'deki tum mesajlar okunuyor")
        else:
            print("  📅 Tam tarama: pump_research.db'deki tum mesajlar okunuyor (henuz islenmis sinyal yok)")
        return _rows_to_signals(rows)

    # ────────────────────────────────────────────────────────────
    # YOL B: iki fazli incremental
    # ────────────────────────────────────────────────────────────
    new_channels = _detect_new_channels(db_path, bt_db)
    existing_channels = set(per_channel_cutoff.keys())

    rows: list = []

    conn = sqlite3.connect(db_path)
    c = conn.cursor()
    try:
        # ─── FAZ 1: YENI KANALLAR — tum mesajlar ───────────────
        if new_channels:
            preview = ", ".join(f"@{ch}" for ch in sorted(new_channels)[:5])
            more = f" (+{len(new_channels) - 5} daha)" if len(new_channels) > 5 else ""
            print(
                f"  📅 FAZ 1 — YENI KANAL: {len(new_channels)} kanal signal_backtest'te yok "
                f"→ tum gecmis mesajlari cekiliyor"
            )
            print(f"     {preview}{more}")

            placeholders = ",".join("?" * len(new_channels))
            c.execute(
                f"""
                SELECT id, channel_name, timestamp, message_text
                FROM messages
                WHERE message_text IS NOT NULL
                  AND channel_name IN ({placeholders})
                """,
                tuple(sorted(new_channels)),
            )
            new_rows = c.fetchall()
            rows.extend(new_rows)
            print(f"     → {len(new_rows)} mesaj cekildi (yeni kanal toplam)")
        else:
            print("  📅 FAZ 1 — YENI KANAL: yok, atlandi")

        # ─── FAZ 2: MEVCUT KANALLAR — kanal-basina cutoff ──────
        if existing_channels:
            print(
                f"  📅 FAZ 2 — MEVCUT KANALLAR: {len(existing_channels)} kanal "
                f"icin kanal-basina cutoff uygulaniyor"
            )

            placeholders = ",".join("?" * len(existing_channels))
            channels_ordered = sorted(existing_channels)
            params: list = list(channels_ordered)

            # Her kanal icin (channel, cutoff) ciftini VALUES tablosu yerine
            # CTE+UNION ALL ile inject etmek yerine, daha basit yol:
            # Python'da kanal bazli per-cutoff icin kucuk bir lookup yapip
            # tek SQL'de IN listesi ile cekip Python'da filtreleyelim.
            c.execute(
                f"""
                SELECT id, channel_name, timestamp, message_text
                FROM messages
                WHERE message_text IS NOT NULL
                  AND channel_name IN ({placeholders})
                """,
                tuple(params),
            )
            candidate_rows = c.fetchall()

            existing_rows = [
                r for r in candidate_rows
                if r[2] is not None and r[2] >= per_channel_cutoff.get(r[1], "")
            ]
            rows.extend(existing_rows)
            print(
                f"     → {len(candidate_rows)} aday, {len(existing_rows)} mesaj "
                f"cutoff sonrasi (mevcut kanal toplam)"
            )
        else:
            print("  📅 FAZ 2 — MEVCUT KANALLAR: yok, atlandi")
    finally:
        conn.close()

    # ─── Birlestir ve timestamp'e gore sirala ──────────────────
    rows.sort(key=lambda r: (r[2] or "", r[0]))
    return _rows_to_signals(rows)


def _rows_to_signals(rows: list) -> list[dict]:
    """messages tablosu satirlarini ticker'li sinyal dict'lerine donusturur."""
    signals = []
    for (rid, ch, ts, msg) in rows:
        ticker = extract_ticker(msg)
        if not ticker:
            continue                          # ticker yoksa hiç dahil etme

        direction = extract_direction(msg)    # LONG | SHORT | LONG_IMPL | None

        entry_m = re.search(r"Entry[:\s]+([0-9][0-9. \-]+)", msg, re.IGNORECASE)
        entry = entry_m.group(1).strip()[:40] if entry_m else ""

        signals.append({
            "message_db_id": rid,
            "channel":       ch,
            "signal_ts":     ts,
            "ticker":        ticker,
            "direction":     direction if direction else "NEUTRAL",
            "entry_raw":     entry,
        })
    return signals


# ─────────────────────────────────────────────────────────────────
# BİNANCE API
# ─────────────────────────────────────────────────────────────────

_symbol_cache: dict = {}


def _ts_ms(ts_str: str) -> int:
    ts_str = ts_str.replace("+00:00", "").replace("Z", "")
    if "+" in ts_str:
        ts_str = ts_str[:ts_str.index("+")]
    dt = datetime.fromisoformat(ts_str).replace(tzinfo=timezone.utc)
    return int(dt.timestamp() * 1000)


def _klines(url: str, symbol: str, interval: str,
            start_ms: int, limit: int):
    params = {
        "symbol": symbol, "interval": interval,
        "startTime": start_ms, "limit": limit,
    }
    try:
        r = requests.get(url, params=params, timeout=10)
        if r.status_code == 429:
            retry = int(r.headers.get("Retry-After", 30))
            print(f"    ⏳ Rate limit — {retry}s bekleniyor...")
            time.sleep(retry)
            r = requests.get(url, params=params, timeout=10)
        if r.status_code == 200:
            return r.json()
    except requests.RequestException as e:
        print(f"    ⚠ Istek hatasi: {e}")
    return None


def _close(candle: list) -> float:
    return float(candle[4])


def _open(candle: list) -> float:
    return float(candle[1])


def _high(candle: list) -> float:
    return float(candle[2])


def _low(candle: list) -> float:
    return float(candle[3])


def _volume(candle: list) -> float:
    return float(candle[5])


def _trade_count(candle: list) -> int:
    return int(candle[8])


def _mean_or_none(values: list[float]) -> float | None:
    valid = [v for v in values if v is not None]
    if not valid:
        return None
    return round(sum(valid) / len(valid), 8)


def _std_or_none(values: list[float]) -> float | None:
    valid = [float(v) for v in values if v is not None]
    if len(valid) < 2:
        return None
    mean_val = sum(valid) / len(valid)
    variance = sum((v - mean_val) ** 2 for v in valid) / (len(valid) - 1)
    return round(variance ** 0.5, 6)


def _ratio_or_none(num: float | None, den: float | None,
                   digits: int = 4) -> float | None:
    if num is None or den in (None, 0):
        return None
    return round(num / den, digits)


def _taker_buy_vol(candle: list) -> float:
    """Binance klines[9]: taker buy base asset volume (alış tarafı hacim)."""
    return float(candle[9])


def _candle_snapshot(candle: list) -> dict:
    vol        = _volume(candle)
    taker_buy  = _taker_buy_vol(candle)
    return {
        "candle_open":             _open(candle),
        "candle_high":             _high(candle),
        "candle_low":              _low(candle),
        "candle_close":            _close(candle),
        "candle_volume":           vol,
        "trade_count":             _trade_count(candle),
        "candle_taker_buy_vol":    taker_buy,
        # >0.5 = alış baskısı, <0.5 = satış baskısı
        "candle_taker_buy_ratio":  round(taker_buy / vol, 4) if vol > 0 else None,
    }


def _compute_pre_signal_metrics(prev_candles: list,
                                signal_price: float | None) -> dict:
    metrics = {col: None for col in PRE_SIGNAL_COLUMNS}
    if signal_price is None:
        return metrics

    for window_label, window_size in PRE_SIGNAL_WINDOW_SPECS:
        if len(prev_candles) < window_size:
            continue

        window = prev_candles[-window_size:]
        baseline = (
            prev_candles[-2 * window_size:-window_size]
            if len(prev_candles) >= 2 * window_size else []
        )

        start_open = _open(window[0])
        start_close = _close(window[0])
        end_close = _close(window[-1])
        highs = [_high(candle) for candle in window]
        lows = [_low(candle) for candle in window]
        volumes      = [_volume(candle)          for candle in window]
        trades       = [_trade_count(candle)     for candle in window]
        taker_buys   = [_taker_buy_vol(candle)   for candle in window]

        metrics[f"pre_pct_{window_label}"] = calc_pct(start_close, signal_price)

        if start_open:
            price_range = max(highs) - min(lows)
            metrics[f"pre_range_{window_label}"] = round(price_range / start_open * 100, 3)
            if price_range > 0:
                metrics[f"pre_efficiency_{window_label}"] = round(
                    abs(end_close - start_open) / price_range,
                    4,
                )

        returns = []
        prev_close = None
        for candle in window:
            close_val = _close(candle)
            if prev_close not in (None, 0):
                returns.append((close_val - prev_close) / prev_close * 100)
            prev_close = close_val
        metrics[f"pre_volatility_{window_label}"] = _std_or_none(returns)

        # Pencere genelinde taker buy oranı (yönlü hacim baskısı)
        total_vol   = sum(volumes)
        total_taker = sum(taker_buys)
        metrics[f"pre_taker_buy_ratio_{window_label}"] = (
            round(total_taker / total_vol, 4) if total_vol > 0 else None
        )

        if baseline:
            baseline_volumes = [_volume(candle)      for candle in baseline]
            baseline_trades  = [_trade_count(candle) for candle in baseline]
            metrics[f"pre_volume_rel_{window_label}"] = _ratio_or_none(
                _mean_or_none(volumes),
                _mean_or_none(baseline_volumes),
            )
            metrics[f"pre_trade_rel_{window_label}"] = _ratio_or_none(
                _mean_or_none(trades),
                _mean_or_none(baseline_trades),
            )

    return metrics


def fetch_prices(ticker: str, signal_ts: str) -> dict:
    symbol_usdt = ticker + "USDT"
    result = {
        "symbol": symbol_usdt, "market": "not_found",
        "price_signal": None,
        "price_5m": None, "price_30m": None,
        "price_1h": None, "price_4h": None, "price_1d": None,
        "candle_open": None, "candle_high": None,
        "candle_low": None, "candle_close": None,
        "candle_volume": None, "trade_count": None,
        "candle_taker_buy_vol": None, "candle_taker_buy_ratio": None,
        "prev_volume": None, "volume_ma_24": None, "volume_ma_48": None,
        "pump_before_4h": None, "pump_before_24h": None,
        "btc_pct_1h": None, "btc_pct_24h": None,
        **{col: None for col in PRE_SIGNAL_COLUMNS},
    }
    start_ms = _ts_ms(signal_ts)

    # Hangi market?
    market_url = None
    signal_1m = None
    if symbol_usdt in _symbol_cache:
        cached = _symbol_cache[symbol_usdt]
        if cached == "spot":
            market_url = BINANCE_SPOT; result["market"] = "spot"
        elif cached == "futures":
            market_url = BINANCE_FUTURES; result["market"] = "futures"
        else:
            return result
        signal_test = _klines(market_url, symbol_usdt, "1m", start_ms, 1)
        if signal_test:
            signal_1m = signal_test[0]
    else:
        test = _klines(BINANCE_SPOT, symbol_usdt, "1m", start_ms, 1)
        if test:
            _symbol_cache[symbol_usdt] = "spot"
            market_url = BINANCE_SPOT; result["market"] = "spot"
            signal_1m = test[0]
        else:
            test = _klines(BINANCE_FUTURES, symbol_usdt, "1m", start_ms, 1)
            if test:
                _symbol_cache[symbol_usdt] = "futures"
                market_url = BINANCE_FUTURES; result["market"] = "futures"
                signal_1m = test[0]
            else:
                _symbol_cache[symbol_usdt] = None
                return result

    if not signal_1m:
        return result

    signal_open_ms = int(signal_1m[0])
    result.update(_candle_snapshot(signal_1m))
    result["price_signal"] = result["candle_close"]

    # Gecmis 1m klines -> hacim ortalamalari ve sinyal oncesi momentum/range
    hist_1m = _klines(
        market_url,
        symbol_usdt,
        "1m",
        max(0, signal_open_ms - 480 * 60 * 1000),
        481,
    )
    time.sleep(0.10)

    if hist_1m:
        current_idx = next(
            (idx for idx, candle in enumerate(hist_1m) if int(candle[0]) == signal_open_ms),
            None,
        )
        if current_idx is None:
            current_idx = len(hist_1m) - 1

        prev_candles = hist_1m[:current_idx]
        prev_volumes = [_volume(c) for c in prev_candles]

        if prev_volumes:
            result["prev_volume"] = prev_volumes[-1]
        if len(prev_volumes) >= 24:
            result["volume_ma_24"] = _mean_or_none(prev_volumes[-24:])
        if len(prev_volumes) >= 48:
            result["volume_ma_48"] = _mean_or_none(prev_volumes[-48:])
        result.update(_compute_pre_signal_metrics(prev_candles, result["price_signal"]))

    # 1m klines -> sinyal, +5dk, +30dk, +1sa
    kl_1m = _klines(market_url, symbol_usdt, "1m", signal_open_ms, 62)
    time.sleep(0.15)

    if kl_1m:
        result["price_signal"] = _close(kl_1m[0])
        if len(kl_1m) > 5:  result["price_5m"]  = _close(kl_1m[5])
        if len(kl_1m) > 30: result["price_30m"] = _close(kl_1m[30])
        if len(kl_1m) > 60: result["price_1h"]  = _close(kl_1m[60])
        elif len(kl_1m) > 1: result["price_1h"] = _close(kl_1m[-1])

    # 1h klines -> +4sa, +1gun + pump_before hesabı
    # Sinyal anından 24 saat öncesini de kapsayacak şekilde çek
    kl_1h_start = max(0, start_ms - 24 * 60 * 60 * 1000)
    kl_1h_full = _klines(market_url, symbol_usdt, "1h", kl_1h_start, 50)
    time.sleep(0.15)

    if kl_1h_full:
        # Sinyal anındaki candle'ı bul
        signal_idx = None
        for idx, c in enumerate(kl_1h_full):
            if int(c[0]) >= start_ms:
                signal_idx = idx
                break
        if signal_idx is None:
            signal_idx = len(kl_1h_full) - 1

        # Geleceğe doğru: +4sa, +1gun
        if signal_idx + 4 < len(kl_1h_full):
            result["price_4h"] = _close(kl_1h_full[signal_idx + 4])
        if signal_idx + 24 < len(kl_1h_full):
            result["price_1d"] = _close(kl_1h_full[signal_idx + 24])
        elif signal_idx + 1 < len(kl_1h_full):
            result["price_1d"] = _close(kl_1h_full[-1])

        # ── pump_before: sinyal öncesi ne kadar hareket olmuş ──
        p_signal = result["price_signal"]
        if p_signal and signal_idx >= 4:
            price_4h_ago = _close(kl_1h_full[signal_idx - 4])
            if price_4h_ago and price_4h_ago > 0:
                result["pump_before_4h"] = round(
                    (p_signal - price_4h_ago) / price_4h_ago * 100, 3
                )
        if p_signal and signal_idx >= 24:
            price_24h_ago = _close(kl_1h_full[signal_idx - 24])
            if price_24h_ago and price_24h_ago > 0:
                result["pump_before_24h"] = round(
                    (p_signal - price_24h_ago) / price_24h_ago * 100, 3
                )

    # ── BTC piyasa bağlamı ──
    btc_kl = _klines(BINANCE_SPOT, "BTCUSDT", "1h",
                     max(0, start_ms - 24 * 60 * 60 * 1000), 26)
    time.sleep(0.10)

    if btc_kl:
        # BTC sinyal anındaki candle'ı bul
        btc_signal_idx = None
        for idx, c in enumerate(btc_kl):
            if int(c[0]) >= start_ms:
                btc_signal_idx = idx
                break
        if btc_signal_idx is None:
            btc_signal_idx = len(btc_kl) - 1

        btc_now = _close(btc_kl[btc_signal_idx])
        if btc_now:
            if btc_signal_idx >= 1:
                btc_1h_ago = _close(btc_kl[btc_signal_idx - 1])
                if btc_1h_ago and btc_1h_ago > 0:
                    result["btc_pct_1h"] = round(
                        (btc_now - btc_1h_ago) / btc_1h_ago * 100, 3
                    )
            if btc_signal_idx >= 24:
                btc_24h_ago = _close(btc_kl[btc_signal_idx - 24])
                if btc_24h_ago and btc_24h_ago > 0:
                    result["btc_pct_24h"] = round(
                        (btc_now - btc_24h_ago) / btc_24h_ago * 100, 3
                    )

    return result


def calc_pct(price_now, price_then):
    if price_now and price_then and price_now != 0:
        return round((price_then - price_now) / price_now * 100, 3)
    return None


# ─────────────────────────────────────────────────────────────────
# BACKTEST CALIŞTIRICISI
# ─────────────────────────────────────────────────────────────────

def run_backtest(db_path: str, bt_path: str = BT_DB_PATH,
                 force: bool = False):
    init_backtest_table(bt_path)

    # İnkremental: force=False ise sadece son tarihten sonraki mesajları oku
    cached_ids = load_cached(bt_path) if not force else set()
    signals    = parse_signals(db_path, bt_db=bt_path, force=force)

    # cached_ids ile ikinci güvenlik filtresi (since_ts eşit tarihli satırlar için)
    todo = [s for s in signals if s["message_db_id"] not in cached_ids]

    total_in_bt = len(cached_ids)
    print(f"\n📊 pump_research.db'den okunan: {len(signals)} | "
          f"Yeni (işlenecek): {len(todo)} | "
          f"Önceden tamamlanmış: {total_in_bt}")

    if not todo:
        print("✅ Yeni sinyal yok. pump_research.db'e yeni mesaj geldiğinde tekrar çalıştır.")
        return total_in_bt

    print(f"⏱  Tahmini süre: ~{len(todo)*0.6:.0f} saniye\n")

    for i, sig in enumerate(todo, 1):
        ticker = sig["ticker"]
        print(f"[{i:3}/{len(todo)}] {sig['signal_ts'][:16]}  "
              f"{ticker:<12} {sig['direction'] or '?':5}  ", end="", flush=True)

        prices = fetch_prices(ticker, sig["signal_ts"])
        p0 = prices["price_signal"]

        row = {
            **sig, **prices,
            "pct_5m":  calc_pct(p0, prices["price_5m"]),
            "pct_30m": calc_pct(p0, prices["price_30m"]),
            "pct_1h":  calc_pct(p0, prices["price_1h"]),
            "pct_4h":  calc_pct(p0, prices["price_4h"]),
            "pct_1d":  calc_pct(p0, prices["price_1d"]),
        }
        save_backtest_row(bt_path, row)

        if prices["market"] == "not_found":
            print("⚫ bulunamadi")
        else:
            p5  = f"{prices['price_5m']:.4g}" if prices["price_5m"] else "—"
            pct = f"{row['pct_1h']:+.2f}%" if row["pct_1h"] else "—"
            p0_str = f"{p0:.4g}" if p0 else "?"
            print(f"✅ {prices['market']:7}  "
                  f"fiyat={p0_str}  +5m={p5}  +1sa={pct}")

    print(f"\n✅ Backtest tamamlandi. {len(todo)} sinyal islendi.")
    return len(signals)


# ─────────────────────────────────────────────────────────────────
# EXCEL YAZICI
# ─────────────────────────────────────────────────────────────────

def write_excel(bt_path: str = BT_DB_PATH, out_path: str = OUT_PATH):
    conn = sqlite3.connect(bt_path)
    c = conn.cursor()
    c.execute("""
        SELECT message_db_id, channel, signal_ts, ticker, symbol, market,
               direction, entry_raw, price_signal,
               price_5m, price_30m, price_1h, price_4h, price_1d,
               pct_5m, pct_30m, pct_1h, pct_4h, pct_1d
        FROM signal_backtest ORDER BY signal_ts ASC
    """)
    rows = c.fetchall()
    conn.close()

    if not rows:
        print("⚠ Backtest verisi bulunamadi.")
        return

    found = [r for r in rows if r[5] != "not_found"]
    print(f"\n📊 Excel yaziliyor: {len(rows)} satir ({len(found)} fiyat verisi)")

    wb = Workbook()
    _write_main_sheet(wb, rows)
    _write_summary_sheet(wb, rows)
    _write_channel_sheet(wb, rows)
    wb.save(out_path)
    print(f"✅ Kaydedildi: {out_path}")


def _write_main_sheet(wb, rows):
    ws = wb.active
    ws.title = "📊 Sinyal Backtesti"
    ws.sheet_view.showGridLines = False

    found_cnt = sum(1 for r in rows if r[5] != "not_found")
    title_row(ws,
              f"📊  SİNYAL VOLATİLİTE ANALİZİ  ·  "
              f"{found_cnt}/{len(rows)} sinyal fiyat verisi mevcut",
              C["dark"], 19)

    HDR = [
        "#", "Tarih", "Saat", "Kanal", "Ticker", "Borsa", "Yon", "Giris",
        "Sinyal\nFiyati",
        "+5dk\nFiyat", "+5dk\n%",
        "+30dk\nFiyat", "+30dk\n%",
        "+1sa\nFiyat", "+1sa\n%",
        "+4sa\nFiyat", "+4sa\n%",
        "+1gun\nFiyat", "+1gun\n%",
    ]
    header_row(ws, 2, HDR, C["dark"], h=32)
    set_col_widths(ws, [4,11,7,22,10,9,8,20,12,11,9,11,9,11,9,11,9,11,9])

    PRICE_FMT = "#,##0.0####"

    for i, r in enumerate(rows, start=3):
        (mid, ch, ts, ticker, symbol, market, direction,
         entry_raw, p0, p5m, p30m, p1h, p4h, p1d,
         pct5m, pct30m, pct1h, pct4h, pct1d) = r

        stripe_row(ws, i, 19, i % 2 == 0)

        ws.cell(i, 1, i-2).alignment = center()
        ws.cell(i, 2, ts[:10]).font = reg_font(C["blue"])
        ws.cell(i, 3, ts[11:16])
        ws.cell(i, 4, f"@{ch}").font = Font(name="Arial", size=8, color=C["purple"])

        tc = ws.cell(i, 5, ticker)
        tc.font = bold_font(C["blue"], 9); tc.alignment = center()

        mc = ws.cell(i, 6, market); mc.alignment = center()
        if market == "spot":
            mc.fill = fill(C["l_green"]); mc.font = bold_font(C["green"], 8)
        elif market == "futures":
            mc.fill = fill(C["l_blue"]); mc.font = bold_font(C["blue"], 8)
        else:
            mc.fill = fill(C["gray"]); mc.font = reg_font(C["dim"], 8)

        if direction:
            dc = ws.cell(i, 7, direction); dc.alignment = center()
            if direction == "LONG":
                dc.fill = fill(C["l_green"]); dc.font = bold_font(C["green"], 9)
            else:
                dc.fill = fill(C["l_red"]); dc.font = bold_font(C["red"], 9)

        ws.cell(i, 8, entry_raw).font = reg_font(C["dim"], 8)

        for col, price in [(9,p0),(10,p5m),(12,p30m),(14,p1h),(16,p4h),(18,p1d)]:
            if price is not None:
                pc = ws.cell(i, col, price)
                pc.number_format = PRICE_FMT
                pc.alignment = center(); pc.font = reg_font(size=9)

        for col, pct_val in [(11,pct5m),(13,pct30m),(15,pct1h),(17,pct4h),(19,pct1d)]:
            pct_cell(ws, i, col, pct_val, direction)


def _write_summary_sheet(wb, rows):
    ws = wb.create_sheet("🎯 Ticker Ozeti")
    ws.sheet_view.showGridLines = False

    found = [r for r in rows if r[5] != "not_found" and r[8] is not None]
    title_row(ws,
              f"🎯  TİCKER BAZLI ORTALAMA HAREKETLER  ·  {len(found)} sinyal",
              C["green"], 12)

    HDR = ["Ticker","Borsa","Yon","Sinyal\nSayisi",
           "Ort +5dk%","Ort +30dk%","Ort +1sa%","Ort +4sa%","Ort +1gun%",
           "Max +1sa%","Min +1sa%","Kanal(lar)"]
    header_row(ws, 2, HDR, C["green"], h=30)
    set_col_widths(ws, [10,9,8,10,11,11,11,11,11,11,11,45])

    ticker_data = defaultdict(list)
    for r in found:
        (mid, ch, ts, ticker, symbol, market, direction,
         entry_raw, p0, p5m, p30m, p1h, p4h, p1d,
         pct5m, pct30m, pct1h, pct4h, pct1d) = r
        ticker_data[ticker].append({
            "market": market, "direction": direction or "", "ch": ch,
            "pct5m": pct5m, "pct30m": pct30m, "pct1h": pct1h,
            "pct4h": pct4h, "pct1d": pct1d,
        })

    def avg(lst):
        valid = [x for x in lst if x is not None]
        return round(sum(valid) / len(valid), 3) if valid else None

    sorted_tickers = sorted(ticker_data.items(), key=lambda x: -len(x[1]))

    for i, (ticker, records) in enumerate(sorted_tickers, start=3):
        stripe_row(ws, i, 12, i % 2 == 0, height=15)

        tc = ws.cell(i, 1, ticker)
        tc.font = bold_font(C["blue"], 10); tc.alignment = center()

        markets = Counter(r["market"] for r in records)
        best_mkt = markets.most_common(1)[0][0]
        mc = ws.cell(i, 2, best_mkt); mc.alignment = center()
        mc.fill = fill(C["l_green"] if best_mkt == "spot" else C["l_blue"])
        mc.font = bold_font(C["green"] if best_mkt == "spot" else C["blue"], 8)

        dirs = Counter(r["direction"] for r in records if r["direction"])
        ws.cell(i, 3, " / ".join(f"{d}:{n}" for d,n in dirs.most_common())).alignment = center()
        ws.cell(i, 4, len(records)).alignment = center()
        ws.cell(i, 4).font = bold_font(C["dark"], 10)

        for col, key in [(5,"pct5m"),(6,"pct30m"),(7,"pct1h"),(8,"pct4h"),(9,"pct1d")]:
            a = avg([r[key] for r in records])
            if a is not None:
                cc = ws.cell(i, col, f"{a:+.2f}%")
                cc.alignment = center(); cc.font = reg_font(size=9)
                if a > 0:   cc.fill = fill(C["l_green"]); cc.font = bold_font(C["green"], 9)
                elif a < 0: cc.fill = fill(C["l_red"]);   cc.font = bold_font(C["red"], 9)

        pcts_1h = [r["pct1h"] for r in records if r["pct1h"] is not None]
        if pcts_1h:
            mx = ws.cell(i, 10, f"{max(pcts_1h):+.2f}%")
            mx.alignment = center(); mx.fill = fill(C["l_green"]); mx.font = bold_font(C["green"], 9)
            mn = ws.cell(i, 11, f"{min(pcts_1h):+.2f}%")
            mn.alignment = center(); mn.fill = fill(C["l_red"]); mn.font = bold_font(C["red"], 9)

        chs = sorted(set(r["ch"] for r in records))
        ws.cell(i, 12, ", ".join(chs)).font = reg_font(C["dim"], 8)


def _write_channel_sheet(wb, rows):
    ws = wb.create_sheet("📡 Kanal Analizi")
    ws.sheet_view.showGridLines = False

    title_row(ws,
              "📡  KANAL BAZLI PERFORMANS  ·  Sinyal sonrasi ortalama hareketler",
              C["purple"], 10)

    HDR = ["Kanal","Sinyal\nSayisi","Fiyat Verisi",
           "Ort +5dk%","Ort +30dk%","Ort +1sa%","Ort +4sa%","Ort +1gun%",
           "Yon Dagilimi","En Cok Ticker"]
    header_row(ws, 2, HDR, C["purple"], h=28)
    set_col_widths(ws, [26,10,12,11,11,11,11,11,20,30])

    ch_data = defaultdict(list)
    for r in rows:
        ch_data[r[1]].append(r)

    def avg(lst):
        valid = [x for x in lst if x is not None]
        return round(sum(valid) / len(valid), 3) if valid else None

    for i, (ch, records) in enumerate(
            sorted(ch_data.items(), key=lambda x: -len(x[1])), start=3):
        stripe_row(ws, i, 10, i % 2 == 0, height=16)

        ws.cell(i, 1, f"@{ch}").font = Font(
            name="Arial", size=9, color=C["purple"], bold=True)
        ws.cell(i, 2, len(records)).alignment = center()

        with_price = [r for r in records if r[8] is not None]
        wp_c = ws.cell(i, 3, len(with_price)); wp_c.alignment = center()
        if len(with_price) == len(records):
            wp_c.fill = fill(C["l_green"])
        elif len(with_price) == 0:
            wp_c.fill = fill(C["l_red"])

        for col, idx in [(4,14),(5,15),(6,16),(7,17),(8,18)]:
            a = avg([r[idx] for r in records if r[idx] is not None])
            if a is not None:
                cc = ws.cell(i, col, f"{a:+.2f}%")
                cc.alignment = center(); cc.font = reg_font(size=9)
                if a > 0:   cc.fill = fill(C["l_green"]); cc.font = bold_font(C["green"], 9)
                elif a < 0: cc.fill = fill(C["l_red"]);   cc.font = bold_font(C["red"], 9)

        dirs = Counter(r[6] for r in records if r[6])
        ws.cell(i, 9, " / ".join(f"{d}:{n}" for d,n in dirs.most_common())).alignment = center()

        top_t = Counter(r[3] for r in records if r[3]).most_common(3)
        ws.cell(i, 10, ", ".join(f"{t}({n})" for t,n in top_t)).font = reg_font(C["dim"], 8)


# ─────────────────────────────────────────────────────────────────
# RUN (programatik cagri icin)
# ─────────────────────────────────────────────────────────────────

def run(db: str = DB_PATH, btdb: str = BT_DB_PATH,
        out: str = OUT_PATH, force: bool = False,
        excel_only: bool = False):
    """
    Backtest'i calistir ve Excel uret.
    main.py'den veya komut satirindan cagirilabilir.
    """
    db   = str(Path(db).resolve())
    btdb = str(Path(btdb).resolve())
    out  = str(Path(out).resolve())

    init_backtest_table(btdb)

    if not excel_only:
        run_backtest(db, bt_path=btdb, force=force)

    write_excel(bt_path=btdb, out_path=out)


# ─────────────────────────────────────────────────────────────────
# CLI
# ─────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Sinyal volatilite backtester")
    parser.add_argument("--db",    default=DB_PATH)
    parser.add_argument("--btdb",  default=BT_DB_PATH)
    parser.add_argument("--out",   default=OUT_PATH)
    parser.add_argument("--force", action="store_true")
    parser.add_argument("--excel-only", action="store_true")
    args = parser.parse_args()
    run(db=args.db, btdb=args.btdb, out=args.out,
        force=args.force, excel_only=args.excel_only)


if __name__ == "__main__":
    main()
