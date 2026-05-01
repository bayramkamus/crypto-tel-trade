#!/usr/bin/env python3
"""
Sinyal ML Analiz Motoru
========================
backtest_results.db uzerinden:
  1. Win rate analizi (kanal / saat / gun / yon)
  2. Korelasyon matrisi (market cap dahil)
  3. Karar agaci (yorumlanabilir kurallar)
  4. Market cap etkisi (LARGE / MID / SMALL cap)
  5. Kanal performans siralamasi

Kullanim:
    python ml_analysis.py
    python ml_analysis.py --target win_4h --no-volume
"""

import time
import sqlite3
import argparse
import warnings
from pathlib import Path

import requests
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from excel_styles import (
    C, fill, bold_font, reg_font, center, left, thin_border,
    title_row, header_row, stripe_row, set_col_widths,
    wr_cell, pct_cell,
)

try:
    from sklearn.tree import DecisionTreeClassifier, export_text
    from sklearn.ensemble import RandomForestClassifier
    from sklearn.preprocessing import LabelEncoder
    from sklearn.model_selection import cross_val_score, StratifiedKFold
    from sklearn.metrics import balanced_accuracy_score
    SKLEARN_OK = True
except ImportError:
    SKLEARN_OK = False
    print("⚠  scikit-learn bulunamadi. Karar agaci analizi devre disi.")

warnings.filterwarnings("ignore")

# ─────────────────────────────────────────────────────────────────
# SABITLER
# ─────────────────────────────────────────────────────────────────

BT_DB_PATH = "backtest_results.db"
OUT_PATH   = "ml_analysis.xlsx"

COINGECKO_MARKETS = "https://api.coingecko.com/api/v3/coins/markets"

PRE_SIGNAL_WINDOW_SPECS = [
    ("5m", "5m", 5),
    ("30m", "30m", 30),
    ("1h", "1h", 60),
    ("4h", "4h", 240),
]
PRE_SIGNAL_METRIC_SPECS = [
    ("pre_pct",             "Momentum",        "Pencere baslangic close -> signal close % degisim"),
    ("pre_range",           "Range",           "(max high - min low) / pencere ilk open"),
    ("pre_volatility",      "Volatilite",      "Pencere icindeki 1m getiri std"),
    ("pre_volume_rel",      "Hacim Orani",     "Pencere ort hacim / onceki es pencere"),
    ("pre_trade_rel",       "Trade Orani",     "Pencere ort trade / onceki es pencere"),
    ("pre_efficiency",      "Efficiency",      "|close - open| / range"),
    ("pre_taker_buy_ratio", "Taker Buy Oran",  "Pencere toplam taker-buy / toplam hacim  (>0.5=alis baskisi)"),
]
PRE_SIGNAL_COLUMNS = [
    f"{metric}_{window_key}"
    for metric, _, _ in PRE_SIGNAL_METRIC_SPECS
    for window_key, _, _ in PRE_SIGNAL_WINDOW_SPECS
]
FEATURE_SPECS = [
    (
        f"{metric}_{window_key}",
        window_label,
        indicator_label,
        formula,
    )
    for window_key, window_label, _ in PRE_SIGNAL_WINDOW_SPECS
    for metric, indicator_label, formula in PRE_SIGNAL_METRIC_SPECS
]
FEATURE_WINDOW_ORDER = {
    window_label: idx for idx, (_, window_label, _) in enumerate(PRE_SIGNAL_WINDOW_SPECS)
}

LOAD_BACKTEST_COLUMNS = [
    "message_db_id", "channel", "signal_ts", "ticker", "symbol", "market",
    "direction", "entry_raw", "price_signal",
    "price_5m", "price_30m", "price_1h", "price_4h", "price_1d",
    "pct_5m", "pct_30m", "pct_1h", "pct_4h", "pct_1d",
    "candle_open", "candle_high", "candle_low", "candle_close",
    "candle_volume", "trade_count",
    "candle_taker_buy_vol", "candle_taker_buy_ratio",
    "prev_volume", "volume_ma_24", "volume_ma_48",
    *PRE_SIGNAL_COLUMNS,
    "trend_score", "trend_momentum",   # Google Trends zenginleştirmesi
    # Katman 1 — sadece uyarı/gösterge, ML'e girmiyor
    "pump_before_4h", "pump_before_24h",
    # Katman 2 — ML feature'ları
    "btc_pct_1h", "btc_pct_24h",
    "fear_greed", "fear_greed_label",
    "fng_momentum_7d", "fng_momentum_14d",
]


# ─────────────────────────────────────────────────────────────────
# VERİ YUKLEME
# ─────────────────────────────────────────────────────────────────

def load_backtest(db_path: str) -> pd.DataFrame:
    conn = sqlite3.connect(db_path)
    existing = {row[1] for row in conn.execute("PRAGMA table_info(signal_backtest)")}
    selected = [c for c in LOAD_BACKTEST_COLUMNS if c in existing]
    df = pd.read_sql_query(f"""
        SELECT {", ".join(selected)}
        FROM signal_backtest
        ORDER BY signal_ts ASC
    """, conn)
    conn.close()
    for col in LOAD_BACKTEST_COLUMNS:
        if col not in df.columns:
            df[col] = np.nan
    return df


def _safe_div(num: pd.Series, den: pd.Series) -> pd.Series:
    num = pd.to_numeric(num, errors="coerce")
    den = pd.to_numeric(den, errors="coerce")
    return num / den.where(den != 0)


def enrich_features(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df["entry_raw"] = df["entry_raw"].fillna("")
    df["signal_ts"] = pd.to_datetime(df["signal_ts"], utc=True)
    df["date"]      = df["signal_ts"].dt.date.astype(str)
    df["hour"]      = df["signal_ts"].dt.hour
    df["weekday"]   = df["signal_ts"].dt.day_name()
    df["weekday_n"] = df["signal_ts"].dt.dayofweek

    def time_block(h):
        if 0  <= h < 6:  return "Gece (0-6)"
        if 6  <= h < 12: return "Sabah (6-12)"
        if 12 <= h < 18: return "Ogleden Sonra (12-18)"
        return "Aksam (18-24)"
    df["time_block"] = df["hour"].apply(time_block)

    # Yön normalize: LONG_IMPL'i win/fav hesabında LONG gibi davran;
    # NEUTRAL mesajlarda win/fav = NaN (yön belirsiz).
    _LONG_DIRS  = {"LONG", "LONG_IMPL"}
    _SHORT_DIRS = {"SHORT"}
    _UNDEF_DIRS = {"NEUTRAL", "", None}

    for period in ["5m", "30m", "1h", "4h", "1d"]:
        pct_col = f"pct_{period}"
        win_col = f"win_{period}"
        fav_col = f"fav_{period}"
        if pct_col in df.columns:
            df[win_col] = df.apply(
                lambda r, p=pct_col: (
                    np.nan if r["direction"] in _UNDEF_DIRS or pd.isna(r[p]) else
                    1 if (r["direction"] in _LONG_DIRS  and r[p] > 0) else
                    1 if (r["direction"] in _SHORT_DIRS and r[p] < 0) else
                    0
                ), axis=1
            )
            df[fav_col] = df.apply(
                lambda r, p=pct_col: (
                    r[p]  if r["direction"] in _LONG_DIRS  else
                    -r[p] if r["direction"] in _SHORT_DIRS else
                    np.nan
                ), axis=1
            )

    df["has_entry"] = (df["entry_raw"].str.len() > 0).astype(int)
    # dir_signal: LONG/LONG_IMPL → +1, SHORT → -1, NEUTRAL → NaN
    df["dir_signal"] = df["direction"].map(
        {"LONG": 1, "LONG_IMPL": 1, "SHORT": -1}
    )

    numeric_cols = [
        "candle_open", "candle_high", "candle_low", "candle_close",
        "candle_volume", "trade_count",
        "candle_taker_buy_vol", "candle_taker_buy_ratio",
        "prev_volume", "volume_ma_24", "volume_ma_48",
        *PRE_SIGNAL_COLUMNS,
        "trend_score", "trend_momentum", "trend_dir_signal",
    ]
    for col in numeric_cols:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    df["range"] = df["candle_high"] - df["candle_low"]
    df["range_norm"] = _safe_div(df["range"], df["candle_open"])
    df["range_close"] = _safe_div(df["range"], df["candle_close"])
    df["volume_delta"] = _safe_div(df["candle_volume"] - df["prev_volume"], df["prev_volume"])
    df["volume_rel_24"] = _safe_div(df["candle_volume"], df["volume_ma_24"])
    df["volume_rel_48"] = _safe_div(df["candle_volume"], df["volume_ma_48"])
    df["trade_density"] = _safe_div(df["trade_count"], df["range"])
    df["liquidity_pressure"] = _safe_div(df["candle_volume"], df["range"])
    df["efficiency"] = _safe_div((df["candle_close"] - df["candle_open"]).abs(), df["range"])
    df["volvol"] = _safe_div(df["candle_volume"], df["range_norm"])

    # ── Lag Score ────────────────────────────────────────────────────
    # Sinyalin ne kadar geç geldiğini ölçer:
    #   +1.0 → hareket tamamıyla signal öncesinde oldu   (geç giriş)
    #    0.0 → hareket eşit paylaşılmış veya yok        (dengeli)
    #   -1.0 → sinyal öncesi fiyat düştü, sonrası çıktı (kontra-momentum)
    # Yüksek pozitif lag_score + LONG  →  tepeye biniş riski yüksek
    if "pre_pct_1h" in df.columns and "pct_1h" in df.columns:
        pre  = pd.to_numeric(df["pre_pct_1h"], errors="coerce")
        post = pd.to_numeric(df["pct_1h"],     errors="coerce")
        df["lag_score_1h"] = (pre / (pre.abs() + post.abs() + 1e-6)).replace(
            [np.inf, -np.inf], np.nan
        )

    # trend_dir_signal: trend_momentum × dir_signal
    if "trend_momentum" in df.columns and "dir_signal" in df.columns:
        tm = pd.to_numeric(df["trend_momentum"], errors="coerce")
        ds = pd.to_numeric(df["dir_signal"],     errors="coerce")
        df["trend_dir_signal"] = (tm * ds).replace([np.inf, -np.inf], np.nan)

    # ── Yön-ayarlı momentum terimleri ─────────────────────────────
    # Raw pre_pct_* yön bilgisi taşımaz: LONG'da +%5 iyi, SHORT'da +%5 kötü.
    # Karar ağacına raw momentum sokarsak LONG/SHORT etkileri birbirini iptal eder.
    # Çözüm: pre_pct_*_dir = pre_pct_* × dir_signal
    #   +değer → momentum işlem yönüyle AYNI   (avantajlı giriş)
    #   -değer → momentum işlem yönüne TERS    (kontra-momentum, riskli)
    #   NaN    → NEUTRAL sinyaller (dir_signal=NaN)
    # Karar ağacına yalnızca _dir versiyonları girer, raw versiyonlar GIRMEZ.
    if "dir_signal" in df.columns:
        ds = pd.to_numeric(df["dir_signal"], errors="coerce")
        for wk, _, _ in PRE_SIGNAL_WINDOW_SPECS:
            raw_col = f"pre_pct_{wk}"
            dir_col = f"pre_pct_{wk}_dir"
            if raw_col in df.columns:
                raw = pd.to_numeric(df[raw_col], errors="coerce")
                df[dir_col] = (raw * ds).replace([np.inf, -np.inf], np.nan)

    # ── BTC & FnG numeric cast ─────────────────────────────────────
    for col in ["btc_pct_1h", "btc_pct_24h", "fear_greed",
                "fng_momentum_7d", "fng_momentum_14d",
                "pump_before_4h", "pump_before_24h"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    # ── BTC yön-ayarlı terimler ────────────────────────────────────
    # Raw btc_pct_* direction-agnostic: BTC +%2 LONG için iyi, SHORT için kötü.
    # Karar ağacında raw versiyon LONG/SHORT etkilerini iptal eder.
    # Çözüm: btc_pct_*_dir = btc_pct_* × dir_signal
    #   +değer → BTC hareketi işlem yönüyle AYNI  (piyasa bağlamı uygun)
    #   -değer → BTC hareketi işlem yönüne TERS   (kontra-piyasa, riskli)
    # Karar ağacına yalnızca _dir versiyonları girer.
    if "dir_signal" in df.columns:
        ds = pd.to_numeric(df["dir_signal"], errors="coerce")
        for btc_col in ["btc_pct_1h", "btc_pct_24h"]:
            if btc_col in df.columns:
                raw = pd.to_numeric(df[btc_col], errors="coerce")
                df[f"{btc_col}_dir"] = (raw * ds).replace([np.inf, -np.inf], np.nan)

    # ── Channel Reliability Score (kümülatif 30 günlük win rate) ───
    # Her sinyal için, o kanalın son 30 gündeki geçmiş win rate'ini hesaplar.
    # Bu "o ana kadarki bilgi" kullandığı için look-ahead bias yoktur.
    if "win_1h" in df.columns and "channel" in df.columns:
        df = df.sort_values("signal_ts").reset_index(drop=True)
        ch_wr_30d = []
        for idx, row in df.iterrows():
            ch = row["channel"]
            sig_ts = row["signal_ts"]
            window_start = sig_ts - pd.Timedelta(days=30)
            # Bu sinyalden ÖNCE, aynı kanaldan, son 30 gündeki sinyaller
            mask = (
                (df["channel"] == ch) &
                (df["signal_ts"] < sig_ts) &
                (df["signal_ts"] >= window_start) &
                (df["win_1h"].notna())
            )
            past = df.loc[mask, "win_1h"]
            if len(past) >= 3:
                ch_wr_30d.append(round(past.mean() * 100, 1))
            else:
                ch_wr_30d.append(np.nan)
        df["channel_win_rate_30d"] = ch_wr_30d

    derived_cols = [col for col, _, _, _ in FEATURE_SPECS]
    available = [col for col in derived_cols if col in df.columns]
    if available:
        df[available] = df[available].replace([np.inf, -np.inf], np.nan)
    return df


# ─────────────────────────────────────────────────────────────────
# MARKET CAP VERİSİ
# ─────────────────────────────────────────────────────────────────

def fetch_marketcap_data(tickers: list) -> dict:
    """CoinGecko API uzerinden market cap ceker. Doner: {ticker: market_cap_usd}"""
    print(f"\n💰 {len(tickers)} ticker icin piyasa degeri (market cap) cekiliyor...")
    targets = {t.upper() for t in tickers}
    cap_map = {}

    for page in (1, 2, 3):
        try:
            r = requests.get(COINGECKO_MARKETS, params={
                "vs_currency": "usd", "order": "market_cap_desc",
                "per_page": 250, "page": page, "sparkline": "false",
            }, timeout=12)

            if r.status_code == 429:
                print("  ⏳ CoinGecko rate limit — 60s bekleniyor...")
                time.sleep(60)
                r = requests.get(COINGECKO_MARKETS, params={
                    "vs_currency": "usd", "order": "market_cap_desc",
                    "per_page": 250, "page": page, "sparkline": "false",
                }, timeout=12)

            if r.status_code != 200:
                print(f"  ⚠ CoinGecko API hatasi: HTTP {r.status_code}")
                break

            for coin in r.json():
                sym = (coin.get("symbol") or "").upper()
                mc  = coin.get("market_cap")
                if sym and mc:
                    cap_map[sym] = float(mc)

            time.sleep(1.2)
        except Exception as e:
            print(f"  ⚠ CoinGecko istegi basarisiz (sayfa {page}): {e}")
            break

        if targets.issubset(cap_map.keys()):
            break

    result = {}
    for t in tickers:
        mc = cap_map.get(t.upper())
        result[t] = mc
        if mc:
            print(f"  {t:<12} ${mc:>20,.0f}")
        else:
            print(f"  {t:<12} ⚫ bulunamadi")
    return result


def assign_cap_tier(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    if "market_cap_usd" not in df.columns:
        df["cap_tier"] = "UNKNOWN"
        return df

    has_cap = df[df["market_cap_usd"].notna()]["market_cap_usd"]
    if len(has_cap) < 3:
        df["cap_tier"] = "UNKNOWN"
        return df

    low_cut  = has_cap.quantile(0.33)
    high_cut = has_cap.quantile(0.67)

    def tier(v):
        if pd.isna(v):    return "UNKNOWN"
        if v <= low_cut:  return "LOW"
        if v <= high_cut: return "MID"
        return "HIGH"

    df["cap_tier"] = df["market_cap_usd"].apply(tier)
    return df


def _cohen_d(a: pd.Series, b: pd.Series) -> float | None:
    a = pd.to_numeric(a, errors="coerce").dropna()
    b = pd.to_numeric(b, errors="coerce").dropna()
    if len(a) < 2 or len(b) < 2:
        return None

    var_a = a.var(ddof=1)
    var_b = b.var(ddof=1)
    pooled_den = (len(a) + len(b) - 2)
    if pooled_den <= 0:
        return None

    pooled_var = (((len(a) - 1) * var_a) + ((len(b) - 1) * var_b)) / pooled_den
    if pd.isna(pooled_var) or pooled_var <= 0:
        return None
    return float((a.mean() - b.mean()) / np.sqrt(pooled_var))


def _permutation_pvalue(long_vals: pd.Series, short_vals: pd.Series,
                        seed: int = 42, iterations: int = 1500) -> float | None:
    long_arr = pd.to_numeric(long_vals, errors="coerce").dropna().to_numpy(dtype=float)
    short_arr = pd.to_numeric(short_vals, errors="coerce").dropna().to_numpy(dtype=float)
    if len(long_arr) < 3 or len(short_arr) < 3:
        return None

    observed = abs(long_arr.mean() - short_arr.mean())
    pooled = np.concatenate([long_arr, short_arr])
    long_count = len(long_arr)
    rng = np.random.default_rng(seed)

    exceed = 0
    for _ in range(iterations):
        perm = rng.permutation(pooled)
        diff = abs(perm[:long_count].mean() - perm[long_count:].mean())
        if diff >= observed - 1e-12:
            exceed += 1

    return float((exceed + 1) / (iterations + 1))


def _feature_status(p_value: float | None,
                    corr_direction: float | None,
                    cohen_d: float | None) -> str:
    strength = max(
        abs(corr_direction) if corr_direction is not None and not pd.isna(corr_direction) else 0,
        abs(cohen_d) if cohen_d is not None and not pd.isna(cohen_d) else 0,
    )
    if p_value is not None and p_value < 0.05 and strength >= 0.35:
        return "Anlamli"
    if p_value is not None and p_value < 0.10 and strength >= 0.20:
        return "Sinirda"
    return "Zayif"


def build_feature_diagnostics(df: pd.DataFrame) -> pd.DataFrame:
    # LONG_IMPL de "boğa" grubuna dahil; NEUTRAL dışlanır
    sub = df[df["direction"].isin(["LONG", "LONG_IMPL", "SHORT"])].copy()
    if sub.empty or "dir_signal" not in sub.columns:
        return pd.DataFrame()

    rows = []
    dir_target = pd.to_numeric(sub["dir_signal"], errors="coerce")

    for idx, (col, window_label, indicator_label, formula) in enumerate(FEATURE_SPECS):
        if col not in sub.columns:
            continue

        series = pd.to_numeric(sub[col], errors="coerce").replace([np.inf, -np.inf], np.nan)
        valid_mask = series.notna() & dir_target.notna()
        if valid_mask.sum() < 8:
            continue

        valid = series.loc[valid_mask]
        dirs = dir_target.loc[valid_mask]
        long_vals = valid.loc[dirs == 1]
        short_vals = valid.loc[dirs == -1]
        if len(long_vals) < 4 or len(short_vals) < 4:
            continue

        corr_direction = valid.corr(dirs)
        cohen_d = _cohen_d(long_vals, short_vals)
        perm_p = _permutation_pvalue(long_vals, short_vals, seed=42 + idx)

        row = {
            "window": window_label,
            "indicator": indicator_label,
            "formula": formula,
            "column": col,
            "n": int(valid.shape[0]),
            "long_mean": round(long_vals.mean(), 4),
            "short_mean": round(short_vals.mean(), 4),
            "delta_long_short": round(long_vals.mean() - short_vals.mean(), 4),
            "corr_direction": round(corr_direction, 3) if pd.notna(corr_direction) else None,
            "cohen_d": round(cohen_d, 3) if cohen_d is not None and not pd.isna(cohen_d) else None,
            "perm_p": round(perm_p, 4) if perm_p is not None else None,
        }
        row["status"] = _feature_status(
            row["perm_p"],
            row["corr_direction"],
            row["cohen_d"],
        )

        rows.append(row)

    diag = pd.DataFrame(rows)
    if len(diag):
        diag["_p_sort"] = diag["perm_p"].fillna(1.0)
        diag["_effect"] = diag[["corr_direction", "cohen_d"]].abs().fillna(0).max(axis=1)
        diag["_window_sort"] = diag["window"].map(FEATURE_WINDOW_ORDER).fillna(99)
        diag = diag.sort_values(
            ["_p_sort", "_effect", "n", "_window_sort"],
            ascending=[True, False, False, True],
        ).drop(columns=["_p_sort", "_effect", "_window_sort"])
    return diag


# ─────────────────────────────────────────────────────────────────
# ANALİZ FONKSİYONLARI
# ─────────────────────────────────────────────────────────────────

def win_rate_table(df: pd.DataFrame, group_col: str,
                   periods=("1h", "4h", "1d")) -> pd.DataFrame:
    rows = []
    for val, grp in df.groupby(group_col):
        row = {group_col: val, "n": len(grp)}
        for p in periods:
            win_col = f"win_{p}"
            fav_col = f"fav_{p}"
            valid = grp[win_col].dropna()
            row[f"win_{p}_%"] = round(valid.mean() * 100, 1) if len(valid) else None
            row[f"n_{p}"]     = len(valid)
            fav_valid = grp[fav_col].dropna()
            row[f"avg_fav_{p}_%"] = round(fav_valid.mean(), 2) if len(fav_valid) else None
        rows.append(row)
    return pd.DataFrame(rows)


def correlation_matrix(df: pd.DataFrame) -> pd.DataFrame:
    """Eski full matris — geriye uyumluluk için tutuluyor."""
    cols = [
        "dir_signal", "hour", "weekday_n", "market_cap_usd", "has_entry",
        *PRE_SIGNAL_COLUMNS,
        "candle_volume", "candle_taker_buy_ratio",
        "prev_volume", "volume_ma_24", "volume_ma_48", "trade_count",
        "range", "range_norm", "range_close",
        "volume_delta", "volume_rel_24", "volume_rel_48",
        "trade_density", "liquidity_pressure", "efficiency", "volvol",
        "lag_score_1h",
        "trend_score", "trend_momentum", "trend_dir_signal",
        "btc_pct_1h", "btc_pct_24h", "fear_greed",
        "fng_momentum_7d", "fng_momentum_14d",
        "channel_win_rate_30d",
        "pct_5m", "pct_30m", "pct_1h", "pct_4h", "pct_1d",
        "fav_1h", "fav_4h", "fav_1d",
        "win_1h", "win_4h", "win_1d",
    ]
    available = [
        c for c in cols
        if c in df.columns and pd.to_numeric(df[c], errors="coerce").notna().sum() >= 2
    ]
    if not available:
        return pd.DataFrame()
    return df[available].apply(pd.to_numeric, errors="coerce").corr().round(3)


def target_correlation(df: pd.DataFrame,
                       targets: list[str] = None) -> pd.DataFrame:
    """
    Hedef odaklı korelasyon: Her feature'ın win/fav hedefleriyle korelasyonu.
    Dönen DataFrame: feature × target, |corr| büyükten küçüğe sıralı.
    """
    if targets is None:
        targets = ["win_1h", "win_4h", "fav_1h", "fav_4h"]

    feature_cols = [
        ("dir_signal",             "Yön Sinyali (+1 LONG, -1 SHORT)"),
        ("hour",                   "Saat"),
        ("weekday_n",              "Haftanın Günü (0=Pzt)"),
        ("has_entry",              "Giriş Fiyatı Var mı"),
        ("candle_volume",          "Candle Hacmi"),
        ("candle_taker_buy_ratio", "Taker Buy Oranı"),
        ("trade_count",            "Trade Sayısı"),
        ("range_norm",             "Range / Open"),
        ("volume_delta",           "Hacim Değişimi"),
        ("volume_rel_24",          "Göreceli Hacim (24h)"),
        ("efficiency",             "Price Efficiency"),
        ("lag_score_1h",           "Lag Skoru (1h)"),
        ("trend_score",            "Google Trend Skoru"),
        ("trend_momentum",         "Trend Momentumu"),
        ("trend_dir_signal",       "Yönlü Trend (momentum×yön)"),
        ("btc_pct_1h",             "BTC 1sa Değişim (raw)"),
        ("btc_pct_24h",            "BTC 24sa Değişim (raw)"),
        ("btc_pct_1h_dir",         "BTC 1sa Yönlü (×yön)"),
        ("btc_pct_24h_dir",        "BTC 24sa Yönlü (×yön)"),
        ("fear_greed",             "Korku/Açgözlülük (FnG)"),
        ("fng_momentum_7d",        "FnG İvme 7 Gün"),
        ("fng_momentum_14d",       "FnG İvme 14 Gün"),
        ("channel_win_rate_30d",   "Kanal WR 30 Gün"),
        ("market_cap_usd",         "Piyasa Değeri (USD)"),
        ("pump_before_4h",         "Pump Before 4h (uyarı)"),
        ("pump_before_24h",        "Pump Before 24h (uyarı)"),
    ]
    # pre_signal kolonlarından sadece momentum ve taker_buy (en anlamlılar)
    for wk, wl, _ in PRE_SIGNAL_WINDOW_SPECS:
        feature_cols.append((f"pre_pct_{wk}",             f"Pre Momentum {wl}"))
        feature_cols.append((f"pre_taker_buy_ratio_{wk}", f"Pre Taker Buy {wl}"))

    rows = []
    for col, label in feature_cols:
        if col not in df.columns:
            continue
        series = pd.to_numeric(df[col], errors="coerce")
        if series.notna().sum() < 10:
            continue

        row = {"feature": col, "label": label, "n": int(series.notna().sum())}
        max_abs = 0
        for t in targets:
            if t not in df.columns:
                row[t] = np.nan
                continue
            t_series = pd.to_numeric(df[t], errors="coerce")
            mask = series.notna() & t_series.notna()
            if mask.sum() < 10:
                row[t] = np.nan
                continue
            corr_val = series[mask].corr(t_series[mask])
            row[t] = round(corr_val, 4) if pd.notna(corr_val) else np.nan
            if pd.notna(corr_val) and abs(corr_val) > max_abs:
                max_abs = abs(corr_val)
        row["max_abs_corr"] = max_abs
        rows.append(row)

    result = pd.DataFrame(rows)
    if len(result):
        result = result.sort_values("max_abs_corr", ascending=False).reset_index(drop=True)
    return result


def _prepare_ml_data(df: pd.DataFrame, target: str = "win_1h") -> dict:
    """
    Ortak ML veri hazırlığı — decision_tree_analysis ve random_forest_analysis tarafından paylaşılır.

    Özellik seçim mantığı:
      1) Kategorik değişkenler LabelEncoder ile kodlanır (kanal, yön, cap_tier, time_block).
      2) Sayısal adaylar arasından HEM win HEM fav ile tutarlı korele olanlar seçilir.
         - |win_corr| >= MIN_WIN_CORR : win rate'i tahmin edebilmeli
         - |fav_corr| >= MIN_FAV_CORR : fav hareketi de açıklamalı
         - sign aynı              : her iki hedefi aynı yönde etkiliyor olmalı
         - |win_corr| >= 0.30×|fav_corr| : sadece büyüklük tahmincisi olan özellikler dışlanır
      3) Eksik değerler medyan ile doldurulur.
    """
    sub = df[(df["market"] != "not_found") & df[target].notna()].copy()
    if len(sub) < 20:
        return {"error": f"Yeterli veri yok ({len(sub)} satir)"}

    le_ch  = LabelEncoder()
    le_dir = LabelEncoder()
    le_cap = LabelEncoder()
    le_tb  = LabelEncoder()

    sub["ch_enc"]  = le_ch.fit_transform(sub["channel"])
    sub["dir_enc"] = le_dir.fit_transform(sub["direction"].fillna("?"))
    sub["cap_enc"] = le_cap.fit_transform(
        sub["cap_tier"] if "cap_tier" in sub.columns
        else pd.Series(["UNKNOWN"] * len(sub))
    )
    sub["tb_enc"]  = le_tb.fit_transform(sub["time_block"])

    base_feature_cols = ["ch_enc", "dir_enc", "cap_enc", "tb_enc", "hour", "weekday_n"]
    feature_cols  = list(base_feature_cols)
    feature_names = [
        "Kanal", "Yon (Long=0/Short=1)",
        "Market Cap Kategorisi", "Zaman Dilimi",
        "Saat", "Haftanin Gunu",
    ]

    numeric_candidates = [
        ("has_entry",               "Giris Var"),
        ("candle_volume",           "Candle Hacmi"),
        ("candle_taker_buy_ratio",  "Taker Buy Oran"),     # yönlü hacim baskısı
        ("trade_count",             "Trade Sayisi"),
        ("range",                   "Range"),
        ("range_norm",              "Range/Open"),
        ("range_close",             "Range/Close"),
        ("volume_delta",            "Hacim Degisimi"),
        ("volume_rel_24",           "Goreceli Hacim (24)"),
        ("volume_rel_48",           "Goreceli Hacim (48)"),
        ("trade_density",           "Trade Yogunlugu"),
        ("liquidity_pressure",      "Likidite Baskisi"),
        ("efficiency",              "Price Efficiency"),
        ("volvol",                  "Vol/Hacim"),
        ("lag_score_1h",            "Lag Skoru (1h)"),     # hareket ne kadar önceden geldi
        ("trend_score",             "Google Trend Skoru"), # sinyal anındaki arama ilgisi 0-100
        ("trend_momentum",          "Trend Momentumu"),    # trend_score / 30g ort → ivme
        ("trend_dir_signal",        "Yönlü Trend"),        # trend_momentum × dir_signal
        ("btc_pct_1h_dir",          "BTC 1sa Yönlü"),      # btc_pct_1h × dir_signal
        ("btc_pct_24h_dir",         "BTC 24sa Yönlü"),     # btc_pct_24h × dir_signal
        ("fear_greed",              "Korku/Acgozluluk"),   # Fear & Greed Index 0-100
        ("fng_momentum_7d",         "FnG Ivme 7g"),        # fear_greed / ort(7g) — kısa vadeli ivme
        ("fng_momentum_14d",        "FnG Ivme 14g"),       # fear_greed / ort(14g) — orta vadeli ivme
        ("channel_win_rate_30d",    "Kanal WR 30g"),       # Kanalın son 30 gün win rate'i
    ]
    # FEATURE_SPECS'ten tüm metrikler — pre_pct_* hariç (raw momentum yön bilgisi taşımaz)
    # pre_pct_* : LONG'da +%5 iyi, SHORT'da +%5 kötü → LONG/SHORT etkileri birbirini iptal eder
    # Modele yalnızca pre_pct_*_dir (yönlü versiyon) girer, aşağıda ekleniyor
    numeric_candidates.extend(
        (col, f"{window_label} {indicator_label}")
        for col, window_label, indicator_label, _ in FEATURE_SPECS
        if not col.startswith("pre_pct_")
    )
    # Yön-ayarlı momentum terimleri: pre_pct_*_dir = pre_pct_* × dir_signal
    #   +değer → momentum işlem yönüyle AYNI   (avantajlı giriş)
    #   -değer → momentum işlem yönüne TERS    (kontra-momentum, riskli)
    for wk, wl, _ in PRE_SIGNAL_WINDOW_SPECS:
        numeric_candidates.append((f"pre_pct_{wk}_dir", f"Yönlü Pre Momentum {wl}"))

    # ── Korelasyon tabanlı özellik filtrelemesi ───────────────────────────────────────────
    fav_target        = target.replace("win_", "fav_")
    # channel_win_rate_30d: win_corr=0.0593 → tam değer 0.06 altında, 0.055 ile dahil edildi
    MIN_WIN_CORR      = 0.055
    MIN_FAV_CORR      = 0.05
    MIN_WIN_FAV_RATIO = 0.30   # win_corr en az fav_corr'un %30'u kadar güçlü olmalı

    def _passes_corr_filter(col: str) -> bool:
        if col not in sub.columns:
            return False
        series = pd.to_numeric(sub[col], errors="coerce")
        if series.notna().sum() < 10:
            return False
        win_series = pd.to_numeric(sub[target], errors="coerce")
        mask_w = series.notna() & win_series.notna()
        if mask_w.sum() < 10:
            return False
        win_corr = series[mask_w].corr(win_series[mask_w])
        if pd.isna(win_corr) or abs(win_corr) < MIN_WIN_CORR:
            return False
        if fav_target not in sub.columns:
            return True
        fav_series = pd.to_numeric(sub[fav_target], errors="coerce")
        mask_f = series.notna() & fav_series.notna()
        if mask_f.sum() < 10:
            return True
        fav_corr = series[mask_f].corr(fav_series[mask_f])
        if pd.isna(fav_corr) or abs(fav_corr) < MIN_FAV_CORR:
            return False
        if win_corr * fav_corr < 0:
            return False
        if abs(win_corr) < MIN_WIN_FAV_RATIO * abs(fav_corr):
            return False
        return True

    numeric_feature_cols = []
    excluded_by_filter   = []
    for col, label in numeric_candidates:
        if col not in sub.columns:
            continue
        if pd.to_numeric(sub[col], errors="coerce").notna().sum() < 10:
            continue
        sub[col] = pd.to_numeric(sub[col], errors="coerce")
        if _passes_corr_filter(col):
            numeric_feature_cols.append(col)
            feature_cols.append(col)
            feature_names.append(label)
        else:
            excluded_by_filter.append(col)

    model_df = sub[feature_cols + [target]].copy()
    for col in numeric_feature_cols:
        med = model_df[col].median()
        model_df[col] = model_df[col].fillna(0 if pd.isna(med) else med)

    X = model_df[feature_cols].values
    y = model_df[target].values.astype(int)

    baseline = round(y.mean() * 100, 1)
    n_splits = min(5, max(2, len(sub) // 10))
    cv = StratifiedKFold(n_splits=n_splits, shuffle=True, random_state=42)

    return {
        "sub":                   sub,
        "X":                     X,
        "y":                     y,
        "feature_cols":          feature_cols,
        "feature_names":         feature_names,
        "numeric_feature_cols":  numeric_feature_cols,
        "excluded_by_filter":    excluded_by_filter,
        "baseline":              baseline,
        "cv":                    cv,
        "le_ch":                 le_ch,
        "le_dir":                le_dir,
        "le_cap":                le_cap,
        "le_tb":                 le_tb,
        "n_samples":             len(sub),
    }


def decision_tree_analysis(df: pd.DataFrame, target: str = "win_1h") -> dict:
    if not SKLEARN_OK:
        return {}

    ml = _prepare_ml_data(df, target)
    if "error" in ml:
        return ml

    sub                  = ml["sub"]
    X                    = ml["X"]
    y                    = ml["y"]
    feature_cols         = ml["feature_cols"]
    feature_names        = ml["feature_names"]
    numeric_feature_cols = ml["numeric_feature_cols"]
    excluded_by_filter   = ml["excluded_by_filter"]
    baseline             = ml["baseline"]
    cv                   = ml["cv"]
    le_ch                = ml["le_ch"]
    le_dir               = ml["le_dir"]
    le_cap               = ml["le_cap"]
    le_tb                = ml["le_tb"]

    # ── Çoklu derinlik denemesi ──────────────────────────────────
    # Metrik: balanced_accuracy — sınıf dengesizliğinde accuracy yanıltıcıdır.
    # Win rate %42 ise "hep ZARARI de" modeli %58 accuracy verir ama %50 balanced_accuracy.
    # class_weight='balanced': ağaç eğitiminde azınlık sınıfa (win=1) daha fazla ağırlık verir.
    # min_leaf: derinlik başına veri büyüklüğüne göre dinamik, minimum 20 örnek zorunlu.
    CV_SCORING     = "balanced_accuracy"
    MIN_LEAF_FLOOR = 20

    depth_results = []
    for depth in [3, 4, 5]:
        min_leaf = max(MIN_LEAF_FLOOR, len(sub) // (2 ** (depth + 2)))
        trial_clf = DecisionTreeClassifier(
            max_depth=depth, min_samples_leaf=min_leaf,
            class_weight="balanced", random_state=42
        )
        try:
            trial_cv = cross_val_score(trial_clf, X, y, cv=cv, scoring=CV_SCORING)
            cv_m = round(trial_cv.mean() * 100, 1)
            cv_s = round(trial_cv.std() * 100, 1)
        except Exception:
            cv_m, cv_s = None, None
        trial_clf.fit(X, y)
        train_a = round(balanced_accuracy_score(y, trial_clf.predict(X)) * 100, 1)
        depth_results.append({
            "depth": depth, "min_leaf": min_leaf,
            "train_acc": train_a, "cv_acc": cv_m, "cv_std": cv_s,
        })

    # En iyi derinliği seç: CV balanced_accuracy - overfitting gap penaltısı
    def _depth_score(d):
        if d["cv_acc"] is None:
            return -999
        gap = d["train_acc"] - d["cv_acc"]
        return d["cv_acc"] - 0.8 * gap

    best_depth_info = max(depth_results, key=_depth_score)
    best_depth      = best_depth_info["depth"]
    best_min_leaf   = best_depth_info["min_leaf"]

    clf = DecisionTreeClassifier(
        max_depth=best_depth, min_samples_leaf=best_min_leaf,
        class_weight="balanced", random_state=42
    )
    cv_scores = []
    try:
        cv_scores = cross_val_score(clf, X, y, cv=cv, scoring=CV_SCORING)
    except Exception:
        pass

    clf.fit(X, y)
    tree_text   = export_text(clf, feature_names=feature_names, max_depth=best_depth)
    importances = list(zip(feature_names, clf.feature_importances_.round(4)))
    importances.sort(key=lambda x: -x[1])

    train_acc = round(balanced_accuracy_score(y, clf.predict(X)) * 100, 1)
    cv_mean   = round(cv_scores.mean() * 100, 1) if len(cv_scores) else None

    rules = _extract_leaf_rules(clf, feature_names, le_ch, le_dir, le_cap, le_tb, X, y)

    return {
        "n_samples":          ml["n_samples"],
        "n_positive":         int(y.sum()),
        "baseline_wr":        baseline,
        "train_acc":          train_acc,
        "cv_acc":             cv_mean,
        "cv_std":             round(cv_scores.std() * 100, 1) if len(cv_scores) else None,
        "best_depth":         best_depth,
        "best_min_leaf":      best_min_leaf,
        "depth_results":      depth_results,
        "tree_text":          tree_text,
        "importances":        importances,
        "rules":              rules,
        "excluded_by_filter": excluded_by_filter,
        "n_features_used":    len(numeric_feature_cols),
    }


def random_forest_analysis(df: pd.DataFrame, target: str = "win_1h") -> dict:
    """
    Random Forest analizi — Karar Ağacının bagging tabanlı topluluk versiyonu.

    Karar ağacına göre avantajları:
      - Overfitting daha az: her ağaç farklı bootstrap örneği + rastgele özellik alt kümesi
      - OOB skoru: yerleşik doğrulama tahmini (cross-val'a gerek yok)
      - Feature importance: yüzlerce ağaçtaki Gini azalmalarının ortalaması → daha stabil

    Config grid: n_estimators × min_samples_leaf → gap-penalized seçim.
    """
    if not SKLEARN_OK:
        return {}

    ml = _prepare_ml_data(df, target)
    if "error" in ml:
        return ml

    sub                  = ml["sub"]
    X                    = ml["X"]
    y                    = ml["y"]
    feature_names        = ml["feature_names"]
    numeric_feature_cols = ml["numeric_feature_cols"]
    excluded_by_filter   = ml["excluded_by_filter"]
    baseline             = ml["baseline"]
    cv                   = ml["cv"]

    CV_SCORING = "balanced_accuracy"

    # ── Config grid ───────────────────────────────────────────────
    # n_estimators: daha fazla ağaç → daha stabil ama daha yavaş
    # min_samples_leaf: yaprak boyutu büyük → daha az overfitting
    configs = [
        {"n_estimators": 100, "min_samples_leaf":  5},
        {"n_estimators": 100, "min_samples_leaf": 10},
        {"n_estimators": 100, "min_samples_leaf": 20},
        {"n_estimators": 200, "min_samples_leaf": 10},
        {"n_estimators": 200, "min_samples_leaf": 20},
    ]

    config_results = []
    for cfg in configs:
        rf_trial = RandomForestClassifier(
            n_estimators=cfg["n_estimators"],
            min_samples_leaf=cfg["min_samples_leaf"],
            class_weight="balanced",
            oob_score=True,
            random_state=42,
            n_jobs=-1,
        )
        try:
            trial_cv = cross_val_score(rf_trial, X, y, cv=cv, scoring=CV_SCORING)
            cv_m = round(trial_cv.mean() * 100, 1)
            cv_s = round(trial_cv.std()  * 100, 1)
        except Exception:
            cv_m, cv_s = None, None
        rf_trial.fit(X, y)
        train_a = round(balanced_accuracy_score(y, rf_trial.predict(X)) * 100, 1)
        oob_a   = round(rf_trial.oob_score_ * 100, 1)
        config_results.append({
            "n_estimators": cfg["n_estimators"],
            "min_leaf":     cfg["min_samples_leaf"],
            "train_acc":    train_a,
            "cv_acc":       cv_m,
            "cv_std":       cv_s,
            "oob_acc":      oob_a,
        })

    # Gap-penalized config seçimi (karar ağacıyla aynı kriter)
    def _cfg_score(d):
        if d["cv_acc"] is None:
            return -999
        gap = d["train_acc"] - d["cv_acc"]
        return d["cv_acc"] - 0.8 * gap

    best_cfg = max(config_results, key=_cfg_score)

    # ── Final model ───────────────────────────────────────────────
    rf = RandomForestClassifier(
        n_estimators=best_cfg["n_estimators"],
        min_samples_leaf=best_cfg["min_leaf"],
        class_weight="balanced",
        oob_score=True,
        random_state=42,
        n_jobs=-1,
    )
    cv_scores = []
    try:
        cv_scores = cross_val_score(rf, X, y, cv=cv, scoring=CV_SCORING)
    except Exception:
        pass

    rf.fit(X, y)
    train_acc = round(balanced_accuracy_score(y, rf.predict(X)) * 100, 1)
    cv_mean   = round(cv_scores.mean() * 100, 1) if len(cv_scores) else None
    oob_score = round(rf.oob_score_ * 100, 1)

    importances = list(zip(feature_names, rf.feature_importances_.round(4)))
    importances.sort(key=lambda x: -x[1])

    # ── Surrogate Tree — RF davranışını kural olarak ifade et ─────
    # RF yorumlanamaz (yüzlerce ağaç). Çözüm: RF'in predict_proba çıktısını
    # hedef alarak tek bir sığ DT (surrogate) eğit. Bu DT:
    #   - RF'in öğrendiği karar sınırlarını basit if/else kuralına dönüştürür
    #   - Yapraklardaki win rate gerçek y etiketlerinden hesaplanır (RF tahmini değil)
    #   - class_weight='balanced' kullanmaz (RF olasılıkları zaten dengeli)
    rf_proba   = rf.predict_proba(X)[:, 1]           # RF olasılık skoru
    rf_binary  = (rf_proba >= 0.5).astype(int)        # RF sınıf tahmini

    surr_min_leaf = max(15, len(sub) // 25)
    surrogate = DecisionTreeClassifier(
        max_depth=4,
        min_samples_leaf=surr_min_leaf,
        random_state=42,
    )
    surrogate.fit(X, rf_binary)

    # Surrogate doğruluk: RF tahminlerine ne kadar sadık?
    surr_fidelity = round(
        (surrogate.predict(X) == rf_binary).mean() * 100, 1
    )
    surrogate_rules = _extract_leaf_rules(
        surrogate, feature_names,
        ml["le_ch"], ml["le_dir"], ml["le_cap"], ml["le_tb"],
        X, y   # gerçek y → win rate'ler gerçek sonuçlardan
    )

    return {
        "n_samples":          ml["n_samples"],
        "n_positive":         int(y.sum()),
        "baseline_wr":        baseline,
        "train_acc":          train_acc,
        "cv_acc":             cv_mean,
        "cv_std":             round(cv_scores.std() * 100, 1) if len(cv_scores) else None,
        "oob_score":          oob_score,
        "best_n_estimators":  best_cfg["n_estimators"],
        "best_min_leaf":      best_cfg["min_leaf"],
        "config_results":     config_results,
        "importances":        importances,
        "surrogate_rules":    surrogate_rules,
        "surr_fidelity":      surr_fidelity,
        "surr_min_leaf":      surr_min_leaf,
        "excluded_by_filter": excluded_by_filter,
        "n_features_used":    len(numeric_feature_cols),
    }


def _extract_leaf_rules(clf, feature_names, le_ch, le_dir, le_cap, le_tb, X, y):
    """
    Karar ağacının yaprak kurallarını çıkarır.

    tree.value tabanlı hesaplama yerine clf.apply(X) ile her örneğin
    hangi yaprağa düştüğünü bulur, oradan gerçek y etiketleriyle win rate
    hesaplar. Bu yaklaşım class_weight='balanced' durumunda da doğru çalışır
    (ağırlıklı sayım sorunu tamamen ortadan kalkar).
    """
    from sklearn.tree import _tree
    tree     = clf.tree_
    leaf_ids = clf.apply(X)          # her örnek → yaprak node_id

    label_maps = {
        0: {i: v for i, v in enumerate(le_ch.classes_)},
        1: {i: v for i, v in enumerate(le_dir.classes_)},
        2: {i: v for i, v in enumerate(le_cap.classes_)},
        3: {i: v for i, v in enumerate(le_tb.classes_)},
    }
    fname_map = {i: n for i, n in enumerate(feature_names)}
    rules = []

    def recurse(node, depth, path):
        if tree.children_left[node] == _tree.TREE_LEAF:
            mask   = leaf_ids == node
            n_node = int(mask.sum())
            wins   = int(y[mask].sum()) if n_node else 0
            wr     = round(wins / n_node * 100, 1) if n_node else 0
            rules.append({
                "kosullar":     " VE ".join(path) if path else "(tumu)",
                "ornek_sayisi": n_node,
                "kazanan":      wins,
                "win_rate_%":   wr,
                "tahmin":       "✅ KARLI" if wr >= 50 else "❌ ZARARI",
            })
            return

        feat_idx = tree.feature[node]
        thresh   = tree.threshold[node]
        fname    = fname_map.get(feat_idx, f"F{feat_idx}")

        if feat_idx in label_maps and thresh < len(label_maps[feat_idx]):
            label      = label_maps[feat_idx].get(int(thresh), str(int(thresh)))
            left_cond  = f"{fname} = {label}"
            right_cond = f"{fname} ≠ {label}"
        else:
            left_cond  = f"{fname} ≤ {thresh:.1f}"
            right_cond = f"{fname} > {thresh:.1f}"

        recurse(tree.children_left[node],  depth + 1, path + [left_cond])
        recurse(tree.children_right[node], depth + 1, path + [right_cond])

    recurse(0, 0, [])
    rules.sort(key=lambda r: -r["win_rate_%"])
    return rules


# ─────────────────────────────────────────────────────────────────
# EXCEL SAYFALARI
# ─────────────────────────────────────────────────────────────────

def write_overview_sheet(wb, df, mcap_data):
    ws = wb.active
    ws.title = "📊 Genel Bakis"
    ws.sheet_view.showGridLines = False

    found  = df[df["market"] != "not_found"]
    w_data = found[found["pct_1h"].notna()]

    title_row(ws, "📊  SİNYAL ML ANALİZİ — GENEL BAKIŞ", C["dark"], 8)

    total_sig  = len(df)
    has_price  = len(found)
    overall_wr = w_data["win_1h"].mean() * 100 if len(w_data) else 0
    avg_fav_1h = w_data["fav_1h"].mean() if len(w_data) else 0
    long_cnt   = len(df[df["direction"].isin(["LONG", "LONG_IMPL"])])
    short_cnt  = len(df[df["direction"] == "SHORT"])
    neutral_cnt = len(df[df["direction"].isin(["NEUTRAL", ""])])

    kpis = [
        ("A", "Toplam Sinyal",    total_sig,              C["blue"]),
        ("C", "Fiyat Verisi Var", has_price,              C["green"]),
        ("E", "+1sa Win Rate",    f"{overall_wr:.1f}%",   C["purple"]),
        ("G", "Ort +1sa Hareket", f"{avg_fav_1h:+.2f}%", C["orange"]),
    ]
    for col, lbl, val, clr in kpis:
        c2 = chr(ord(col)+1)
        for r in [3,4,5]:
            ws.merge_cells(f"{col}{r}:{c2}{r}")
        ws.cell(3, ord(col)-64, lbl).font = bold_font()
        ws.cell(3, ord(col)-64).fill = fill(clr)
        ws.cell(3, ord(col)-64).alignment = center()
        vc = ws.cell(4, ord(col)-64, val)
        vc.font = Font(name="Arial", bold=True, size=20, color=clr)
        vc.fill = fill(C["white"]); vc.alignment = center()
        ws.cell(5, ord(col)-64).fill = fill(clr)
        ws.cell(5, ord(col)-64+1).fill = fill(clr)
    for r in [3,4,5]:
        ws.row_dimensions[r].height = 28

    ws.merge_cells("A6:H6")
    note = ws.cell(6, 1,
        f"  📌  Long(+impl)/Short/Neutral: {long_cnt}/{short_cnt}/{neutral_cnt}  ·  "
        f"Market cap verisi olan ticker: {sum(1 for v in mcap_data.values() if v)}  ·  "
        f"Tarih araligi: {str(df['date'].min())} – {str(df['date'].max())}")
    note.font = reg_font(C["dim"], 9)
    note.fill = fill("F9FAFB"); note.alignment = left()
    ws.row_dimensions[6].height = 18

    header_row(ws, 8, ["Periyot","Veri Sayisi","Genel WR%",
                       "Long WR%","Short WR%","Ort Fav %"], C["dark"], h=22)
    for col_l, w in zip("ABCDEFGH", [14,12,14,12,12,14,14,20]):
        ws.column_dimensions[col_l].width = w

    for i, period in enumerate(["5m","30m","1h","4h","1d"], start=9):
        stripe_row(ws, i, 6, i % 2 == 0, height=16)
        win_col = f"win_{period}"; fav_col = f"fav_{period}"
        sub = df[df[win_col].notna()]
        longs  = sub[sub["direction"].isin(["LONG", "LONG_IMPL"])]
        shorts = sub[sub["direction"] == "SHORT"]
        ws.cell(i, 1, f"+{period}").font = bold_font(C["dark"], 10)
        ws.cell(i, 2, len(sub)).alignment = center()
        wr_cell(ws, i, 3, sub[win_col].mean()*100 if len(sub) else None)
        wr_cell(ws, i, 4, longs[win_col].mean()*100 if len(longs) else None)
        wr_cell(ws, i, 5, shorts[win_col].mean()*100 if len(shorts) else None)
        pct_cell(ws, i, 6, sub[fav_col].mean() if len(sub) else None)


def write_channel_sheet(wb, df):
    ws = wb.create_sheet("🏆 Kanal Performansi")
    ws.sheet_view.showGridLines = False
    title_row(ws, "🏆  KANAL BAZLI WIN RATE VE PERFORMANS SIRALAMASI", C["green"], 11)

    HDR = ["Kanal","Sinyal","Fiyat\nVerisi",
           "+1sa\nWR%","+4sa\nWR%","+1gun\nWR%",
           "Ort +1sa\nFav%","Ort +4sa\nFav%","Max +1sa%","Min +1sa%",
           "Skor\n(WR×Fav)"]
    header_row(ws, 2, HDR, C["green"])
    set_col_widths(ws, [26,9,10,10,10,10,13,13,11,11,13])

    ch_wr = win_rate_table(df[df["market"] != "not_found"], "channel")
    ch_wr = ch_wr.sort_values("win_1h_%", ascending=False, na_position="last")

    for i, (_, row) in enumerate(ch_wr.iterrows(), start=3):
        stripe_row(ws, i, 11, i % 2 == 0, height=16)
        ws.cell(i, 1, f"@{row['channel']}").font = bold_font(C["blue"], 9)
        ws.cell(i, 2, int(row["n"])).alignment = center()
        ws.cell(i, 3, int(row["n_1h"])).alignment = center()
        wr_cell(ws, i, 4, row.get("win_1h_%"))
        wr_cell(ws, i, 5, row.get("win_4h_%"))
        wr_cell(ws, i, 6, row.get("win_1d_%"))
        pct_cell(ws, i, 7, row.get("avg_fav_1h_%"))
        pct_cell(ws, i, 8, row.get("avg_fav_4h_%"))

        ch_data = df[(df["channel"] == row["channel"]) & df["pct_1h"].notna()]
        if len(ch_data):
            mx = ws.cell(i, 9, f"{ch_data['pct_1h'].max():+.2f}%")
            mx.alignment = center(); mx.fill = fill(C["l_green"]); mx.font = bold_font(C["dg"], 9)
            mn = ws.cell(i, 10, f"{ch_data['pct_1h'].min():+.2f}%")
            mn.alignment = center(); mn.fill = fill(C["l_red"]); mn.font = bold_font(C["dr"], 9)

        wr_v  = row.get("win_1h_%") or 0
        fav_v = row.get("avg_fav_1h_%") or 0
        score = round(wr_v * (1 + fav_v/100), 2)
        sc = ws.cell(i, 11, f"{score:.1f}")
        sc.alignment = center(); sc.font = bold_font(C["purple"], 10)


def write_feature_sheet(wb, df):
    ws = wb.create_sheet("Feature Lab")
    ws.sheet_view.showGridLines = False
    title_row(ws, "FEATURE LAB - Telegram Oncesi Indikator Ayrismasi", C["purple"], 9)

    feature_df = build_feature_diagnostics(df[df["market"] != "not_found"])
    ws.merge_cells("A2:L2")
    ws.cell(2, 1,
        "  LONG/SHORT mesajlarindan onceki 5m, 30m, 1h ve 4h indikatorleri. "
        "Amac kazanc degil, mesaj oncesi anlamli ayrisma olup olmadigini gormek."
    ).font = reg_font(C["dim"], 9)
    ws.cell(2, 1).fill = fill("F9FAFB")
    ws.row_dimensions[2].height = 20

    if feature_df.empty:
        ws.cell(
            4,
            1,
            "Feature ozeti icin yeterli veri bulunamadi. Gerekirse backtest_signals.py --force ile "
            "pre-signal kolonlarini doldurun.",
        ).font = reg_font(C["orange"])
        return

    header_row(
        ws, 4,
        ["Pencere", "Indikator", "Kolon", "Formula", "Veri",
         "LONG Ort", "SHORT Ort", "Delta", "Corr Yon", "Cohen d", "Perm p", "Durum"],
        C["purple"],
    )
    set_col_widths(ws, [10, 18, 22, 34, 9, 12, 12, 12, 11, 11, 10, 10])

    def write_effect(col_idx, row_idx, value):
        if value is None or pd.isna(value):
            ws.cell(row_idx, col_idx, "-").alignment = center()
            return
        cell = ws.cell(row_idx, col_idx, float(value))
        cell.number_format = "0.000"
        cell.alignment = center()
        if abs(value) >= 0.30:
            cell.fill = fill(C["l_green"] if value > 0 else C["l_red"])
            cell.font = bold_font(C["dg"] if value > 0 else C["dr"], 9)
        elif abs(value) >= 0.15:
            cell.fill = fill("ECFDF5" if value > 0 else "FEE2E2")
            cell.font = reg_font(C["green"] if value > 0 else C["red"], 9)

    def write_pvalue(col_idx, row_idx, value):
        if value is None or pd.isna(value):
            ws.cell(row_idx, col_idx, "-").alignment = center()
            return
        cell = ws.cell(row_idx, col_idx, float(value))
        cell.number_format = "0.0000"
        cell.alignment = center()
        if value < 0.05:
            cell.fill = fill(C["l_green"])
            cell.font = bold_font(C["dg"], 9)
        elif value < 0.10:
            cell.fill = fill(C["l_amber"])
            cell.font = bold_font(C["orange"], 9)

    for i, (_, row) in enumerate(feature_df.iterrows(), start=5):
        stripe_row(ws, i, 12, i % 2 == 0, height=16)
        ws.cell(i, 1, row["window"]).font = bold_font(C["purple"], 9)
        ws.cell(i, 1).alignment = center()
        ws.cell(i, 2, row["indicator"]).font = bold_font(C["dark"], 9)
        ws.cell(i, 3, row["column"]).font = reg_font(C["dim"], 8)
        ws.cell(i, 4, row["formula"]).font = reg_font(C["dim"], 8)
        ws.cell(i, 5, int(row["n"])).alignment = center()
        for col_idx, key in [(6, "long_mean"), (7, "short_mean"), (8, "delta_long_short")]:
            val = row.get(key)
            if val is not None and pd.notna(val):
                cell = ws.cell(i, col_idx, float(val))
                cell.number_format = "0.0000"
                cell.alignment = center()
            else:
                ws.cell(i, col_idx, "-").alignment = center()
        write_effect(9, i, row.get("corr_direction"))
        write_effect(10, i, row.get("cohen_d"))
        write_pvalue(11, i, row.get("perm_p"))

        status_cell = ws.cell(i, 12, row.get("status"))
        status_cell.alignment = center()
        if row.get("status") == "Anlamli":
            status_cell.fill = fill(C["l_green"])
            status_cell.font = bold_font(C["dg"], 9)
        elif row.get("status") == "Sinirda":
            status_cell.fill = fill(C["l_amber"])
            status_cell.font = bold_font(C["orange"], 9)
        else:
            status_cell.fill = fill(C["gray"])
            status_cell.font = reg_font(C["dim"], 8)


def write_hourly_sheet(wb, df):
    ws = wb.create_sheet("🕐 Saat & Gun Analizi")
    ws.sheet_view.showGridLines = False
    title_row(ws, "🕐  SAAT VE GUN BAZLI WIN RATE ANALİZİ", C["orange"], 9)

    sub = df[df["market"] != "not_found"]

    header_row(ws, 2, ["Saat","Sinyal Sayisi","+1sa WR%","+4sa WR%",
                       "Ort +1sa Fav%","Zaman Dilimi"], C["orange"])
    set_col_widths(ws, [8,13,11,11,14,22])

    hour_wr = win_rate_table(sub, "hour").sort_values("hour")
    for i, (_, row) in enumerate(hour_wr.iterrows(), start=3):
        stripe_row(ws, i, 6, i % 2 == 0, height=15)
        h = int(row["hour"])
        ws.cell(i, 1, f"{h:02d}:00").alignment = center()
        ws.cell(i, 2, int(row["n"])).alignment = center()
        wr_cell(ws, i, 3, row.get("win_1h_%"))
        wr_cell(ws, i, 4, row.get("win_4h_%"))
        pct_cell(ws, i, 5, row.get("avg_fav_1h_%"))
        tb_map = {**{hh: "Gece (0-6)" for hh in range(0,6)},
                  **{hh: "Sabah (6-12)" for hh in range(6,12)},
                  **{hh: "Ogleden Sonra (12-18)" for hh in range(12,18)},
                  **{hh: "Aksam (18-24)" for hh in range(18,24)}}
        ws.cell(i, 6, tb_map.get(h, "")).font = reg_font(C["dim"], 8)

    day_row_start = len(hour_wr) + 5
    ws.merge_cells(f"A{day_row_start}:F{day_row_start}")
    ws.cell(day_row_start, 1, "  HAFTANIN GUNU BAZLI ANALİZ").font = bold_font(C["orange"])
    ws.cell(day_row_start, 1).fill = fill(C["l_amber"])
    ws.row_dimensions[day_row_start].height = 20

    header_row(ws, day_row_start+1,
               ["Gun","Sinyal Sayisi","+1sa WR%","+4sa WR%","Ort +1sa Fav%",""], C["orange"])

    day_order = ["Monday","Tuesday","Wednesday","Thursday","Friday","Saturday","Sunday"]
    day_wr = win_rate_table(sub, "weekday")
    day_wr["_order"] = day_wr["weekday"].apply(lambda d: day_order.index(d) if d in day_order else 9)
    day_wr = day_wr.sort_values("_order")
    TR = {"Monday":"Pazartesi","Tuesday":"Sali","Wednesday":"Carsamba",
          "Thursday":"Persembe","Friday":"Cuma","Saturday":"Cumartesi","Sunday":"Pazar"}

    for i, (_, row) in enumerate(day_wr.iterrows(), start=day_row_start+2):
        stripe_row(ws, i, 5, i % 2 == 0, height=15)
        ws.cell(i, 1, TR.get(row["weekday"], row["weekday"])).font = reg_font(C["dark"])
        ws.cell(i, 2, int(row["n"])).alignment = center()
        wr_cell(ws, i, 3, row.get("win_1h_%"))
        wr_cell(ws, i, 4, row.get("win_4h_%"))
        pct_cell(ws, i, 5, row.get("avg_fav_1h_%"))


def write_correlation_sheet(wb, df):
    """
    Hedef odaklı korelasyon sayfası.
    Her feature'ın win_1h, win_4h, fav_1h, fav_4h ile korelasyonu,
    |corr| büyükten küçüğe sıralı.
    """
    ws = wb.create_sheet("📐 Korelasyon")
    ws.sheet_view.showGridLines = False
    title_row(ws, "📐  HEDEF ODAKLI KORELASYON — Feature vs Win Rate / Fav Hareket", C["blue"], 9)

    tcorr = target_correlation(df)
    if tcorr.empty:
        ws.cell(3, 1, "Korelasyon icin yeterli sayisal feature bulunamadi.").font = reg_font(C["orange"])
        return

    ws.merge_cells("A2:I2")
    ws.cell(2, 1,
        "  📌  Her feature'ın win rate ve fav hareket ile korelasyonu. "
        "|r| > 0.10 = zayıf sinyal, |r| > 0.15 = orta, |r| > 0.25 = güçlü. "
        "Sıralama: en güçlü etkiden en zayıfa."
    ).font = reg_font(C["dim"], 9)
    ws.cell(2, 1).fill = fill("F9FAFB")
    ws.row_dimensions[2].height = 20

    ws.merge_cells("A3:I3")
    ws.cell(3, 1,
        "  ⚠  Raw pre_pct_* ve btc_pct_* korelasyonda görünür ancak KARAR AĞACINA GİRMEZ. "
        "Sebep: +%5 hareket LONG için avantajlı, SHORT için dezavantajlıdır — yön bilgisi taşımaz. "
        "Ağaçta yalnızca _dir versiyonları (= raw × dir_signal) kullanılır. "
        "Ayrıca: ağaç class_weight='balanced' + balanced_accuracy ile eğitilir — sınıf dengesizliği düzeltilir."
    ).font = reg_font(C["orange"], 9)
    ws.cell(3, 1).fill = fill("FFF7ED")
    ws.row_dimensions[3].height = 20

    targets = ["win_1h", "win_4h", "fav_1h", "fav_4h"]
    HDR = ["Feature", "Açıklama", "Veri (n)",
           "win_1h\nCorr", "win_4h\nCorr", "fav_1h\nCorr", "fav_4h\nCorr",
           "Max |Corr|", "Yorum"]
    header_row(ws, 5, HDR, C["blue"], h=30)
    set_col_widths(ws, [24, 30, 9, 11, 11, 11, 11, 12, 36])

    def _corr_cell(ws, row, col, val):
        if val is None or (isinstance(val, float) and pd.isna(val)):
            ws.cell(row, col, "—").alignment = center()
            return
        cell = ws.cell(row, col, float(val))
        cell.number_format = "+0.000;-0.000"
        cell.alignment = center()
        cell.border = thin_border()
        av = abs(val)
        if av >= 0.25:
            cell.fill = fill(C["l_green"] if val > 0 else C["l_red"])
            cell.font = bold_font(C["dg"] if val > 0 else C["dr"], 10)
        elif av >= 0.15:
            cell.fill = fill("ECFDF5" if val > 0 else "FEE2E2")
            cell.font = bold_font(C["green"] if val > 0 else C["red"], 9)
        elif av >= 0.10:
            cell.fill = fill("F0FDF4" if val > 0 else "FFF5F5")
            cell.font = reg_font(C["green"] if val > 0 else C["red"], 9)
        else:
            cell.fill = fill(C["gray"])
            cell.font = reg_font(C["dim"], 9)

    def _interpret(row_data):
        best_t = None
        best_v = 0
        for t in targets:
            v = row_data.get(t)
            if v is not None and pd.notna(v) and abs(v) > best_v:
                best_v = abs(v)
                best_t = t
        if best_v < 0.05:
            return "Etkisiz"
        direction = "artırıyor" if row_data.get(best_t, 0) > 0 else "düşürüyor"
        strength = "güçlü" if best_v >= 0.25 else "orta" if best_v >= 0.15 else "zayıf"
        target_label = best_t.replace("win_", "WR ").replace("fav_", "Fav ")
        return f"{target_label} {direction} ({strength})"

    for i, (_, row) in enumerate(tcorr.iterrows(), start=6):
        stripe_row(ws, i, 9, i % 2 == 0, height=17)
        ws.cell(i, 1, row["feature"]).font = bold_font(C["dark"], 9)
        ws.cell(i, 2, row["label"]).font = reg_font(C["dim"], 8)
        ws.cell(i, 3, int(row["n"])).alignment = center()
        for j, t in enumerate(targets, start=4):
            _corr_cell(ws, i, j, row.get(t))

        # Max |corr|
        mac = row.get("max_abs_corr", 0)
        mc = ws.cell(i, 8, f"{mac:.3f}")
        mc.alignment = center()
        if mac >= 0.25:
            mc.fill = fill(C["l_green"]); mc.font = bold_font(C["dg"], 10)
        elif mac >= 0.15:
            mc.fill = fill("ECFDF5"); mc.font = bold_font(C["green"], 9)
        elif mac >= 0.10:
            mc.fill = fill("F0FDF4"); mc.font = reg_font(C["green"], 9)
        else:
            mc.fill = fill(C["gray"]); mc.font = reg_font(C["dim"], 9)

        ws.cell(i, 9, _interpret(row)).font = reg_font(C["dim"], 8)


def write_signal_quality_sheet(wb, df):
    """
    Sinyal kalite filtresi sayfası.

    Kaliteli sinyal kriterleri:
      taker_buy_ratio > 0.55  →  alış baskısı hakim (LONG için)
      taker_buy_ratio < 0.45  →  satış baskısı hakim (SHORT için)
      lag_score_1h  < 0.35    →  hareketin büyük kısmı henüz gelmemiş
      direction in LONG/LONG_IMPL/SHORT

    Amaç: "kaliteli" sinyal havuzunun win rate'ini tüm sinyallerle karşılaştırmak.
    """
    ws = wb.create_sheet("🎯 Sinyal Kalitesi")
    ws.sheet_view.showGridLines = False
    title_row(ws, "🎯  SİNYAL KALİTE FİLTRESİ — Taker Buy + Lag Skoru", C["orange"], 11)

    found = df[
        (df["market"] != "not_found") &
        df["pct_1h"].notna() &
        df["win_1h"].notna()
    ].copy()

    has_taker = "candle_taker_buy_ratio" in found.columns and found["candle_taker_buy_ratio"].notna().any()
    has_lag   = "lag_score_1h" in found.columns and found["lag_score_1h"].notna().any()

    # ── Kalite maskesi ────────────────────────────────────────────────
    _LONG_DIRS  = {"LONG", "LONG_IMPL"}
    _SHORT_DIRS = {"SHORT"}

    if has_taker and has_lag:
        quality_mask = (
            (
                (found["direction"].isin(_LONG_DIRS)  & (found["candle_taker_buy_ratio"] > 0.55)) |
                (found["direction"].isin(_SHORT_DIRS) & (found["candle_taker_buy_ratio"] < 0.45))
            ) &
            (found["lag_score_1h"] < 0.35)
        )
    elif has_taker:
        quality_mask = (
            (found["direction"].isin(_LONG_DIRS)  & (found["candle_taker_buy_ratio"] > 0.55)) |
            (found["direction"].isin(_SHORT_DIRS) & (found["candle_taker_buy_ratio"] < 0.45))
        )
    elif has_lag:
        quality_mask = found["lag_score_1h"] < 0.35
    else:
        ws.cell(3, 1, "⚠  Taker buy ratio ve lag_score henüz hesaplanmamış. "
                      "backtest_signals.py --force çalıştırın.").font = reg_font(C["orange"])
        return

    quality  = found[quality_mask]
    all_dir  = found[found["direction"].isin(_LONG_DIRS | _SHORT_DIRS)]
    qual_dir = quality[quality["direction"].isin(_LONG_DIRS | _SHORT_DIRS)]

    def _wr(sub): return round(sub["win_1h"].mean() * 100, 1) if len(sub) else None
    def _fav(sub): return round(sub["fav_1h"].mean(), 2) if "fav_1h" in sub and len(sub) else None

    # ── KPI Kutuları ─────────────────────────────────────────────────
    kpis = [
        ("A", "Tüm Sinyaller",    len(all_dir),                  C["blue"]),
        ("C", "Kaliteli Sinyal",  len(qual_dir),                 C["green"]),
        ("E", "Genel WR% (+1h)",  f"{_wr(all_dir) or 0:.1f}%",  C["purple"]),
        ("G", "Kalite WR% (+1h)", f"{_wr(qual_dir) or 0:.1f}%", C["orange"]),
    ]
    for col, lbl, val, clr in kpis:
        c2 = chr(ord(col) + 1)
        for r in [3, 4, 5]:
            ws.merge_cells(f"{col}{r}:{c2}{r}")
        ws.cell(3, ord(col) - 64, lbl).font = bold_font(); ws.cell(3, ord(col) - 64).fill = fill(clr)
        ws.cell(3, ord(col) - 64).alignment = center()
        vc = ws.cell(4, ord(col) - 64, val)
        vc.font = Font(name="Arial", bold=True, size=20, color=clr)
        vc.fill = fill(C["white"]); vc.alignment = center()
        ws.cell(5, ord(col) - 64).fill = fill(clr)
        ws.cell(5, ord(col) - 64 + 1).fill = fill(clr)
    for r in [3, 4, 5]:
        ws.row_dimensions[r].height = 28

    # ── Filtre açıklama notu ──────────────────────────────────────────
    ws.merge_cells("A6:H6")
    note_parts = []
    if has_taker: note_parts.append("taker_buy_ratio > 0.55 (LONG) / < 0.45 (SHORT)")
    if has_lag:   note_parts.append("lag_score_1h < 0.35 (hareket henüz gelmemiş)")
    ws.cell(6, 1, f"  📌  Filtre: {' + '.join(note_parts)}").font = reg_font(C["dim"], 9)
    ws.cell(6, 1).fill = fill("F9FAFB"); ws.cell(6, 1).alignment = left()
    ws.row_dimensions[6].height = 16

    # ── Periyot karşılaştırma tablosu ────────────────────────────────
    HDR = ["Periyot", "Tüm\nSinyal", "Tüm\nWR%", "Tüm Ort\nFav%",
           "Kaliteli\nSinyal", "Kaliteli\nWR%", "Kaliteli\nOrt Fav%", "Fark (WR%)"]
    header_row(ws, 8, HDR, C["orange"], h=28)
    set_col_widths(ws, [10, 11, 11, 13, 13, 13, 15, 13])

    for i, period in enumerate(["5m", "30m", "1h", "4h", "1d"], start=9):
        stripe_row(ws, i, 8, i % 2 == 0, height=17)
        wc = f"win_{period}"; fc = f"fav_{period}"
        all_p  = found[found[wc].notna()] if wc in found else pd.DataFrame()
        qual_p = quality[quality[wc].notna()] if wc in quality else pd.DataFrame()

        ws.cell(i, 1, f"+{period}").font = bold_font(C["dark"], 10)
        ws.cell(i, 2, len(all_p)).alignment = center()
        wr_cell(ws, i, 3, all_p[wc].mean() * 100 if len(all_p) else None)
        if fc in all_p: pct_cell(ws, i, 4, all_p[fc].mean() if len(all_p) else None)
        ws.cell(i, 5, len(qual_p)).alignment = center()
        wr_cell(ws, i, 6, qual_p[wc].mean() * 100 if len(qual_p) else None)
        if fc in qual_p: pct_cell(ws, i, 7, qual_p[fc].mean() if len(qual_p) else None)

        all_wr  = all_p[wc].mean()  * 100 if len(all_p)  else None
        qual_wr = qual_p[wc].mean() * 100 if len(qual_p) else None
        if all_wr is not None and qual_wr is not None:
            diff = round(qual_wr - all_wr, 1)
            dc = ws.cell(i, 8, f"{diff:+.1f}%")
            dc.alignment = center()
            if diff > 0:   dc.fill = fill(C["l_green"]); dc.font = bold_font(C["dg"], 10)
            elif diff < 0: dc.fill = fill(C["l_red"]);   dc.font = bold_font(C["dr"], 10)

    # ── Kanal bazlı kalite karşılaştırması ───────────────────────────
    ch_start = 11
    ws.merge_cells(f"A{ch_start}:H{ch_start}")
    ws.cell(ch_start, 1, "  KANAL BAZLI: TÜM SİNYAL vs KALİTELİ SİNYAL WR%").font = bold_font(C["orange"])
    ws.cell(ch_start, 1).fill = fill(C["l_amber"])
    ws.row_dimensions[ch_start].height = 20

    header_row(ws, ch_start + 1,
               ["Kanal", "Tüm\nSinyal", "Tüm\nWR%", "Kaliteli\nSinyal", "Kaliteli\nWR%",
                "Fark", "Ortalama\nTaker Buy", "Ortalama\nLag Score"],
               C["orange"], h=28)
    set_col_widths(ws, [26, 11, 11, 13, 13, 11, 16, 16])

    channels = sorted(all_dir["channel"].unique())
    for j, ch in enumerate(channels, start=ch_start + 2):
        stripe_row(ws, j, 8, j % 2 == 0, height=16)
        ch_all  = all_dir[all_dir["channel"] == ch]
        ch_qual = qual_dir[qual_dir["channel"] == ch]

        ws.cell(j, 1, f"@{ch}").font = bold_font(C["blue"], 9)
        ws.cell(j, 2, len(ch_all)).alignment = center()
        wr_cell(ws, j, 3, ch_all["win_1h"].mean() * 100 if len(ch_all) else None)
        ws.cell(j, 4, len(ch_qual)).alignment = center()
        wr_cell(ws, j, 5, ch_qual["win_1h"].mean() * 100 if len(ch_qual) else None)

        all_w  = ch_all["win_1h"].mean()  * 100 if len(ch_all)  else None
        qual_w = ch_qual["win_1h"].mean() * 100 if len(ch_qual) else None
        if all_w is not None and qual_w is not None:
            d = round(qual_w - all_w, 1)
            dc = ws.cell(j, 6, f"{d:+.1f}%")
            dc.alignment = center()
            if d > 0:   dc.fill = fill(C["l_green"]); dc.font = bold_font(C["dg"], 9)
            elif d < 0: dc.fill = fill(C["l_red"]);   dc.font = bold_font(C["dr"], 9)

        if has_taker and "candle_taker_buy_ratio" in ch_all:
            avg_tbr = ch_all["candle_taker_buy_ratio"].mean()
            if pd.notna(avg_tbr):
                tc = ws.cell(j, 7, f"{avg_tbr:.3f}")
                tc.alignment = center()
                if avg_tbr > 0.55:   tc.fill = fill(C["l_green"]); tc.font = bold_font(C["dg"], 9)
                elif avg_tbr < 0.45: tc.fill = fill(C["l_red"]);   tc.font = bold_font(C["dr"], 9)
                else:                tc.font = reg_font(size=9); tc.alignment = center()

        if has_lag and "lag_score_1h" in ch_all:
            avg_lag = ch_all["lag_score_1h"].mean()
            if pd.notna(avg_lag):
                lc = ws.cell(j, 8, f"{avg_lag:.3f}")
                lc.alignment = center()
                if avg_lag > 0.5:   lc.fill = fill(C["l_red"]);   lc.font = bold_font(C["dr"], 9)
                elif avg_lag < 0.2: lc.fill = fill(C["l_green"]); lc.font = bold_font(C["dg"], 9)
                else:               lc.font = reg_font(size=9); lc.alignment = center()


def write_tree_sheet(wb, df, tree_results):
    ws = wb.create_sheet("🌳 Karar Agaci")
    ws.sheet_view.showGridLines = False
    title_row(ws, "🌳  KARAR AGACI — +1 SAAT WİN RATE ONGORUSU", C["purple"], 8)

    if not tree_results:
        ws.cell(3, 1, "⚠ scikit-learn yuklu degil. pip install scikit-learn")
        return
    if "error" in tree_results:
        ws.cell(3, 1, f"⚠ {tree_results['error']}")
        return

    metrics = [
        ("A", "Ornek Sayisi", tree_results["n_samples"], C["blue"]),
        ("C", "Baseline WR", f"{tree_results['baseline_wr']}%", C["orange"]),
        ("E", "Train Accuracy", f"{tree_results['train_acc']}%", C["green"]),
        ("G", "CV Accuracy",
         f"{tree_results['cv_acc']}% ± {tree_results['cv_std']}%"
         if tree_results["cv_acc"] else "—", C["purple"]),
    ]
    for col, lbl, val, clr in metrics:
        c2 = chr(ord(col)+1)
        for r in [3,4,5]:
            ws.merge_cells(f"{col}{r}:{c2}{r}")
        ws.cell(3, ord(col)-64, lbl).font = bold_font()
        ws.cell(3, ord(col)-64).fill = fill(clr)
        ws.cell(3, ord(col)-64).alignment = center()
        vc = ws.cell(4, ord(col)-64, val)
        vc.font = Font(name="Arial", bold=True, size=16, color=clr)
        vc.fill = fill(C["white"]); vc.alignment = center()
        ws.cell(5, ord(col)-64).fill = fill(clr)
        ws.cell(5, ord(col)-64+1).fill = fill(clr)
    for r in [3,4,5]:
        ws.row_dimensions[r].height = 26

    best_d = tree_results.get("best_depth", 3)
    n_used = tree_results.get("n_features_used", "?")
    excluded = tree_results.get("excluded_by_filter", [])
    ws.merge_cells("A6:H6")
    ws.cell(6, 1,
        f"  ⚠  {tree_results['n_samples']} örnek  ·  Derinlik: {best_d}  ·  "
        f"min_leaf: {tree_results.get('best_min_leaf', 5)}  ·  "
        f"Kullanılan özellik: {n_used}  ·  "
        f"Korelasyon filtresinde dışlanan: {len(excluded)} özellik "
        f"({', '.join(excluded[:5])}{'…' if len(excluded) > 5 else ''})  ·  "
        "CV≈Train ise model sağlıklı, fark büyükse overfitting."
    ).font = reg_font(C["orange"], 9)
    ws.cell(6, 1).fill = fill(C["l_amber"])
    ws.row_dimensions[6].height = 24

    # ── Derinlik karşılaştırma tablosu ────────────────────────────
    depth_results = tree_results.get("depth_results", [])
    if depth_results:
        ws.merge_cells("A8:E8")
        ws.cell(8, 1, "  DERİNLİK KARŞILAŞTIRMASI — En iyi CV accuracy secildi").font = bold_font(C["blue"])
        ws.cell(8, 1).fill = fill(C["l_blue"])
        ws.row_dimensions[8].height = 20

        header_row(ws, 9, ["Derinlik", "Min Leaf", "Train Acc", "CV Acc", "Seçildi"], C["blue"], h=22)
        set_col_widths(ws, [12, 12, 14, 18, 12])

        for i, dr in enumerate(depth_results, start=10):
            stripe_row(ws, i, 5, i % 2 == 0, height=17)
            ws.cell(i, 1, dr["depth"]).alignment = center()
            ws.cell(i, 2, dr["min_leaf"]).alignment = center()
            ws.cell(i, 3, f"{dr['train_acc']}%").alignment = center()
            cv_txt = f"{dr['cv_acc']}% ± {dr['cv_std']}%" if dr["cv_acc"] else "—"
            ws.cell(i, 4, cv_txt).alignment = center()
            if dr["depth"] == best_d:
                ws.cell(i, 5, "✅").alignment = center()
                ws.cell(i, 5).font = bold_font(C["dg"], 12)
                for c in range(1, 6):
                    ws.cell(i, c).fill = fill(C["l_green"])
            else:
                ws.cell(i, 5, "").alignment = center()

        imp_row = 10 + len(depth_results) + 1
    else:
        imp_row = 8
    ws.merge_cells(f"A{imp_row}:D{imp_row}")
    ws.cell(imp_row, 1, "  OZELLİK ONEMİ (Feature Importance)").font = bold_font(C["purple"])
    ws.cell(imp_row, 1).fill = fill(C["l_purp"])
    ws.row_dimensions[imp_row].height = 20

    header_row(ws, imp_row+1, ["Ozellik","Onem Skoru","Gorsel","Yorum"], C["purple"])
    set_col_widths(ws, [26,14,30,40,40,40,40,40])

    interp = {
        "Kanal":              "Hangi kanaldan geldigi en belirleyici faktor",
        "Yon (Long=0/Short=1)": "Long mu Short mu oldugu da etkili",
        "Market Cap Kategorisi": "Buyuk/kucuk cap coinlerde sinyal guvenirligi farki",
        "Zaman Dilimi":       "Gece/sabah/aksam sinyalleri farkli basari orani",
        "Saat":               "Spesifik saat de performansi etkiliyor",
        "Haftanin Gunu":      "Haftanin gunu de anlamli bir faktor",
        "Giris Var":           "Mesajda net entry verilmesi sonucu etkiliyor olabilir",
        "Pre +1sa Momentum":   "Signal oncesi 1 saatlik ivme devam/ters donus sinyali olabilir",
        "Pre +1sa Range":      "Signal oncesi sikisma veya genisleme rejimini yansitir",
        "Candle Hacmi":        "Signal anindaki mutlak hacim seviyesi",
        "Taker Buy Oran":      ">0.55 = alis baskisi; <0.45 = satis baskisi — yon dogrulamasi icin kritik",
        "Trade Sayisi":        "Ayni candle icindeki islem sayisi",
        "Range":               "Signal candle volatilitesi",
        "Range/Open":          "Volatilitenin fiyat seviyesine gore normalize hali",
        "Range/Close":         "Volatilitenin kapanisa gore normalize hali",
        "Hacim Degisimi":      "Bir onceki candle'a gore hacim siklasmasi",
        "Goreceli Hacim (24)": "Kisa pencereye gore hacim sapmasi",
        "Goreceli Hacim (48)": "Uzun pencereye gore hacim sapmasi",
        "Trade Yogunlugu":     "Dar range icinde ne kadar fazla islem oldugu",
        "Likidite Baskisi":    "Range basina hacim; birikim/dagitim izi verebilir",
        "Price Efficiency":    "Candle trend mi yoksa chop mu onu gosterir",
        "Vol/Hacim":           "Volatilite-hacim dengesizligini yakalar",
        "Lag Skoru (1h)":      "+1=hareket onceden tamamlandi (gec giris), 0=denge, -1=kontra-momentum",
        "Google Trend Skoru":  "Sinyal anindaki Google arama yogunlugu (0-100); yuksek=kalabalik ilgi",
        "Trend Momentumu":     "trend_score / 30gun ort; >1.2=yukselen ilgi, <0.8=azalan ilgi",
        "Yönlü Trend":         "trend_momentum × yön (+1/-1); LONG+yüksek=iyi, SHORT+yüksek=riskli",
        "BTC 1sa Degisim":     "BTC son 1 saatteki yüzde degisim; piyasa momentum yönünü gösterir",
        "BTC 24sa Degisim":    "BTC son 24 saatteki yüzde degisim; makro trend bağlamı",
        "Korku/Acgozluluk":    "Fear & Greed Index 0-100; düşük=korku (dip?), yüksek=açgözlülük (tepe?)",
        "Kanal WR 30g":        "Kanalın son 30 gündeki kümülatif win rate'i; güvenilirlik göstergesi",
    }
    for i, (feat, score) in enumerate(tree_results["importances"], start=imp_row+2):
        stripe_row(ws, i, 4, i % 2 == 0, height=16)
        ws.cell(i, 1, feat).font = reg_font(C["dark"])
        sc = ws.cell(i, 2, f"{score:.4f}")
        sc.alignment = center(); sc.font = bold_font(C["purple"], 10)
        bar_len = int(score * 20)
        ws.cell(i, 3, "█" * bar_len + "░" * (20-bar_len)).font = Font(
            name="Courier New", size=9,
            color=C["purple"] if score > 0.2 else (C["blue"] if score > 0.1 else C["dim"]))
        ws.cell(i, 4, interp.get(feat, "")).font = reg_font(C["dim"], 8)

    rule_start = imp_row + len(tree_results["importances"]) + 4
    ws.merge_cells(f"A{rule_start}:H{rule_start}")
    ws.cell(rule_start, 1, "  KARAR AGACI KURALLARI (Win Rate'e Gore Sirali)").font = bold_font(C["purple"])
    ws.cell(rule_start, 1).fill = fill(C["l_purp"])
    ws.row_dimensions[rule_start].height = 20

    header_row(ws, rule_start+1,
               ["Kosullar","Ornek\nSayisi","Kazanan","Win Rate%","Tahmin"],
               C["purple"], h=28)
    ws.column_dimensions["A"].width = 65

    for i, rule in enumerate(tree_results["rules"], start=rule_start+2):
        stripe_row(ws, i, 5, i % 2 == 0, height=16)
        ws.cell(i, 1, rule["kosullar"]).font = reg_font(size=9)
        ws.cell(i, 2, rule["ornek_sayisi"]).alignment = center()
        ws.cell(i, 3, rule["kazanan"]).alignment = center()
        wr_cell(ws, i, 4, rule["win_rate_%"])
        tc = ws.cell(i, 5, rule["tahmin"]); tc.alignment = center()
        if "✅" in rule["tahmin"]:
            tc.fill = fill(C["l_green"]); tc.font = bold_font(C["dg"], 9)
        else:
            tc.fill = fill(C["l_red"]); tc.font = bold_font(C["dr"], 9)

    tree_text_row = rule_start + len(tree_results["rules"]) + 4
    ws.merge_cells(f"A{tree_text_row}:H{tree_text_row}")
    ws.cell(tree_text_row, 1, "  HAM KARAR AGACI (scikit-learn export_text)").font = bold_font(C["dark"])
    ws.cell(tree_text_row, 1).fill = fill(C["mid"])
    for j, line in enumerate(tree_results["tree_text"].split("\n"), start=tree_text_row+1):
        ws.merge_cells(f"A{j}:H{j}")
        ws.cell(j, 1, line).font = Font(name="Courier New", size=9, color=C["dark"])
        ws.row_dimensions[j].height = 13


def write_rf_sheet(wb, df, rf_results):
    """
    Random Forest analiz sayfası.
    KPI kutuları + config karşılaştırma tablosu + feature importance.
    """
    ws = wb.create_sheet("🌲 Random Forest")
    ws.sheet_view.showGridLines = False
    title_row(ws, "🌲  RANDOM FOREST — Topluluk Öğrenmesi ile Win Rate Tahmini", C["green"], 8)

    if not rf_results:
        ws.cell(3, 1, "⚠ scikit-learn yüklü değil. pip install scikit-learn").font = reg_font(C["orange"])
        return
    if "error" in rf_results:
        ws.cell(3, 1, f"⚠ {rf_results['error']}").font = reg_font(C["orange"])
        return

    # ── KPI Kutuları ─────────────────────────────────────────────
    oob_txt = f"{rf_results['oob_score']}%" if rf_results.get("oob_score") is not None else "—"
    cv_txt  = (
        f"{rf_results['cv_acc']}% ± {rf_results['cv_std']}%"
        if rf_results.get("cv_acc") else "—"
    )
    metrics = [
        ("A", "Örnek Sayısı",    rf_results["n_samples"],       C["blue"]),
        ("C", "Baseline WR",     f"{rf_results['baseline_wr']}%", C["orange"]),
        ("E", "CV Balanced Acc", cv_txt,                          C["green"]),
        ("G", "OOB Skoru",       oob_txt,                         C["purple"]),
    ]
    for col, lbl, val, clr in metrics:
        c2 = chr(ord(col) + 1)
        for r in [3, 4, 5]:
            ws.merge_cells(f"{col}{r}:{c2}{r}")
        ws.cell(3, ord(col)-64, lbl).font = bold_font()
        ws.cell(3, ord(col)-64).fill = fill(clr)
        ws.cell(3, ord(col)-64).alignment = center()
        vc = ws.cell(4, ord(col)-64, val)
        vc.font = Font(name="Arial", bold=True, size=16, color=clr)
        vc.fill = fill(C["white"]); vc.alignment = center()
        ws.cell(5, ord(col)-64).fill     = fill(clr)
        ws.cell(5, ord(col)-64+1).fill   = fill(clr)
    for r in [3, 4, 5]:
        ws.row_dimensions[r].height = 26

    # ── Açıklama notu ─────────────────────────────────────────────
    best_n   = rf_results.get("best_n_estimators", "?")
    best_ml  = rf_results.get("best_min_leaf", "?")
    n_used   = rf_results.get("n_features_used", "?")
    excluded = rf_results.get("excluded_by_filter", [])
    ws.merge_cells("A6:H6")
    ws.cell(6, 1,
        f"  ℹ  {rf_results['n_samples']} örnek  ·  Ağaç: {best_n}  ·  "
        f"min_leaf: {best_ml}  ·  Özellik: {n_used}  ·  "
        f"Korelasyon filtresinde dışlanan: {len(excluded)} özellik  ·  "
        "OOB = her ağacın kendi bootstrap dışı örnekleri üzerindeki balanced accuracy tahmini."
    ).font = reg_font(C["dim"], 9)
    ws.cell(6, 1).fill = fill("F9FAFB")
    ws.row_dimensions[6].height = 22

    # ── Config karşılaştırma tablosu ─────────────────────────────
    cfg_results = rf_results.get("config_results", [])
    if cfg_results:
        ws.merge_cells("A8:F8")
        ws.cell(8, 1, "  KONFİGÜRASYON KARŞILAŞTIRMASI — Gap Penaltılı Skor ile En İyi Config Seçildi").font = bold_font(C["green"])
        ws.cell(8, 1).fill = fill(C["l_green"])
        ws.row_dimensions[8].height = 20

        header_row(ws, 9,
                   ["Ağaç Sayısı", "Min Leaf", "Train Acc", "CV Balanced Acc", "OOB Skoru", "Seçildi"],
                   C["green"], h=22)
        set_col_widths(ws, [14, 12, 14, 24, 14, 12])

        for i, cr in enumerate(cfg_results, start=10):
            stripe_row(ws, i, 6, i % 2 == 0, height=17)
            ws.cell(i, 1, cr["n_estimators"]).alignment = center()
            ws.cell(i, 2, cr["min_leaf"]).alignment     = center()
            ws.cell(i, 3, f"{cr['train_acc']}%").alignment = center()
            cv_t = f"{cr['cv_acc']}% ± {cr['cv_std']}%" if cr.get("cv_acc") else "—"
            ws.cell(i, 4, cv_t).alignment = center()
            oob_t = f"{cr['oob_acc']}%" if cr.get("oob_acc") is not None else "—"
            ws.cell(i, 5, oob_t).alignment = center()
            is_best = (cr["n_estimators"] == best_n and cr["min_leaf"] == best_ml)
            if is_best:
                ws.cell(i, 6, "✅").alignment = center()
                ws.cell(i, 6).font = bold_font(C["dg"], 12)
                for c in range(1, 7):
                    ws.cell(i, c).fill = fill(C["l_green"])
            else:
                ws.cell(i, 6, "").alignment = center()

        imp_row = 10 + len(cfg_results) + 1
    else:
        imp_row = 8

    # ── Feature importance ────────────────────────────────────────
    ws.merge_cells(f"A{imp_row}:D{imp_row}")
    ws.cell(imp_row, 1, "  ÖZELLİK ÖNEMİ — Gini Safsızlık Azalması (tüm ağaçların ortalaması)").font = bold_font(C["green"])
    ws.cell(imp_row, 1).fill = fill(C["l_green"])
    ws.row_dimensions[imp_row].height = 20

    ws.merge_cells(f"A{imp_row+1}:H{imp_row+1}")
    ws.cell(imp_row+1, 1,
        "  ℹ  RF feature importance = her ağaçtaki Gini azalmasının ağırlıklı ortalaması. "
        "Tek ağaca göre çok daha stabil; küçük veri setlerinde bile güvenilir sıralama verir."
    ).font = reg_font(C["dim"], 8)
    ws.cell(imp_row+1, 1).fill = fill("F9FAFB")
    ws.row_dimensions[imp_row+1].height = 18

    header_row(ws, imp_row+2, ["Özellik", "Önem Skoru", "Görsel", "Yorum"], C["green"], h=22)
    set_col_widths(ws, [26, 14, 30, 42])

    interp = {
        "Kanal":               "Hangi kanaldan geldigi en belirleyici faktor",
        "Yon (Long=0/Short=1)": "Long mu Short mu oldugu da etkili",
        "Market Cap Kategorisi": "Buyuk/kucuk cap coinlerde sinyal guvenirligi farki",
        "Zaman Dilimi":        "Gece/sabah/aksam sinyalleri farkli basari orani",
        "Saat":                "Spesifik saat de performansi etkiliyor",
        "Haftanin Gunu":       "Haftanin gunu de anlamli bir faktor",
        "Giris Var":           "Mesajda net entry verilmesi sonucu etkiliyor olabilir",
        "Candle Hacmi":        "Signal anindaki mutlak hacim seviyesi",
        "Taker Buy Oran":      ">0.55 = alis baskisi; <0.45 = satis baskisi — yon dogrulamasi icin kritik",
        "Trade Sayisi":        "Ayni candle icindeki islem sayisi",
        "Range":               "Signal candle volatilitesi",
        "Range/Open":          "Volatilitenin fiyat seviyesine gore normalize hali",
        "Range/Close":         "Volatilitenin kapanisa gore normalize hali",
        "Hacim Degisimi":      "Bir onceki candle'a gore hacim siklasmasi",
        "Goreceli Hacim (24)": "Kisa pencereye gore hacim sapmasi",
        "Goreceli Hacim (48)": "Uzun pencereye gore hacim sapmasi",
        "Trade Yogunlugu":     "Dar range icinde ne kadar fazla islem oldugu",
        "Likidite Baskisi":    "Range basina hacim; birikim/dagitim izi verebilir",
        "Price Efficiency":    "Candle trend mi yoksa chop mu onu gosterir",
        "Vol/Hacim":           "Volatilite-hacim dengesizligini yakalar",
        "Lag Skoru (1h)":      "+1=hareket onceden tamamlandi (gec giris), 0=denge, -1=kontra-momentum",
        "Google Trend Skoru":  "Sinyal anindaki Google arama yogunlugu (0-100); yuksek=kalabalik ilgi",
        "Trend Momentumu":     "trend_score / 30gun ort; >1.2=yukselen ilgi, <0.8=azalan ilgi",
        "Yönlü Trend":         "trend_momentum × yön (+1/-1); LONG+yüksek=iyi, SHORT+yüksek=riskli",
        "BTC 1sa Yönlü":       "BTC 1sa hareketi × yön; makro koşulların işlem yönüyle uyumu",
        "BTC 24sa Yönlü":      "BTC 24sa hareketi × yön; daha uzun vadeli makro bağlam",
        "Korku/Acgozluluk":    "Fear & Greed Index 0-100; düşük=korku (dip?), yüksek=açgözlülük (tepe?)",
        "Kanal WR 30g":        "Kanalın son 30 gündeki kümülatif win rate'i; güvenilirlik göstergesi",
    }

    visible_importances = [(f, s) for f, s in rf_results["importances"] if s >= 0.001]
    for i, (feat, score) in enumerate(visible_importances, start=imp_row+3):
        stripe_row(ws, i, 4, i % 2 == 0, height=16)
        ws.cell(i, 1, feat).font = reg_font(C["dark"])
        sc = ws.cell(i, 2, f"{score:.4f}")
        sc.alignment = center(); sc.font = bold_font(C["green"], 10)
        bar_len   = int(score * 20)
        bar_color = C["green"] if score > 0.2 else (C["blue"] if score > 0.1 else C["dim"])
        ws.cell(i, 3, "█" * bar_len + "░" * (20 - bar_len)).font = Font(
            name="Courier New", size=9, color=bar_color)
        ws.cell(i, 4, interp.get(feat, "")).font = reg_font(C["dim"], 8)

    # ── Surrogate Tree Kuralları ──────────────────────────────────
    surr_rules    = rf_results.get("surrogate_rules", [])
    surr_fidelity = rf_results.get("surr_fidelity", None)
    surr_min_leaf = rf_results.get("surr_min_leaf", "?")
    surr_start    = imp_row + 3 + len(visible_importances) + 2

    ws.merge_cells(f"A{surr_start}:H{surr_start}")
    ws.cell(surr_start, 1,
        "  🌲  SURROGATE TREE KURALLARI — RF Kararlarının Yorumlanabilir Özeti"
    ).font = bold_font(C["green"])
    ws.cell(surr_start, 1).fill = fill(C["l_green"])
    ws.row_dimensions[surr_start].height = 20

    fid_txt = f"{surr_fidelity}%" if surr_fidelity is not None else "—"
    ws.merge_cells(f"A{surr_start+1}:H{surr_start+1}")
    ws.cell(surr_start+1, 1,
        f"  ℹ  Surrogate Tree: RF'in tahminleri (predict_proba ≥ 0.5) üzerinde derinlik-4 DT eğitildi. "
        f"Fidelity={fid_txt} (RF kararlarının ne kadarını doğru taklit ediyor).  "
        f"min_leaf={surr_min_leaf}.  "
        "Yaprak win rate'leri gerçek y etiketlerinden hesaplanır, RF tahmininden değil."
    ).font = reg_font(C["dim"], 9)
    ws.cell(surr_start+1, 1).fill = fill("F0FDF4")
    ws.row_dimensions[surr_start+1].height = 22

    if not surr_rules:
        ws.cell(surr_start+2, 1, "Kural çıkarılamadı.").font = reg_font(C["orange"])
    else:
        header_row(ws, surr_start+2,
                   ["Koşullar", "Örnek\nSayısı", "Kazanan", "Win Rate%", "Tahmin"],
                   C["green"], h=28)
        ws.column_dimensions["A"].width = 65

        for i, rule in enumerate(surr_rules, start=surr_start+3):
            stripe_row(ws, i, 5, i % 2 == 0, height=16)
            ws.cell(i, 1, rule["kosullar"]).font = reg_font(size=9)
            ws.cell(i, 2, rule["ornek_sayisi"]).alignment = center()
            ws.cell(i, 3, rule["kazanan"]).alignment = center()
            wr_cell(ws, i, 4, rule["win_rate_%"])
            tc = ws.cell(i, 5, rule["tahmin"]); tc.alignment = center()
            if "✅" in rule["tahmin"]:
                tc.fill = fill(C["l_green"]); tc.font = bold_font(C["dg"], 9)
            else:
                tc.fill = fill(C["l_red"]);   tc.font = bold_font(C["dr"], 9)


def write_marketcap_sheet(wb, df, mcap_data):
    ws = wb.create_sheet("💰 Market Cap Etkisi")
    ws.sheet_view.showGridLines = False
    title_row(ws, "💰  PİYASA DEGERİ (MARKET CAP) ETKİSİ — Buyukluk ve Sinyal Performansi",
              C["blue"], 11)

    sub = df[df["market"] != "not_found"]

    header_row(ws, 2,
               ["Buyukluk Kategorisi","Sinyal Sayisi","Ort Market Cap ($)",
                "+1sa WR%","+4sa WR%","Ort +1sa Fav%","Ort +4sa Fav%","Yorum"], C["blue"])
    set_col_widths(ws, [22,13,20,11,11,14,14,40])

    tier_order  = ["HIGH","MID","LOW","UNKNOWN"]
    tier_labels = {"HIGH":"🔵 LARGE CAP","MID":"🟡 MID CAP",
                   "LOW":"🔴 SMALL CAP","UNKNOWN":"⚫ Bilinmiyor"}
    tier_notes  = {
        "HIGH": "Buyuk piyasa degeri (BTC,ETH,SOL vb.) — Daha likit ve tahmin edilebilir",
        "MID":  "Orta piyasa degeri — Dengeli risk/firsat profili",
        "LOW":  "Kucuk piyasa degeri — Yuksek volatilite, dusuk likidite",
        "UNKNOWN": "CoinGecko'dan market cap verisi alinamadi",
    }

    if "cap_tier" in sub.columns:
        tier_wr = win_rate_table(sub, "cap_tier")
        for i, tier in enumerate(tier_order, start=3):
            row = tier_wr[tier_wr["cap_tier"] == tier]
            stripe_row(ws, i, 8, i % 2 == 0, height=16)
            ws.cell(i, 1, tier_labels[tier]).font = bold_font(C["blue"], 9)
            if len(row):
                r = row.iloc[0]
                ws.cell(i, 2, int(r["n"])).alignment = center()
                tier_caps = [mcap_data.get(t) for t in
                             sub[sub["cap_tier"]==tier]["ticker"].unique()
                             if mcap_data.get(t)]
                if tier_caps:
                    ws.cell(i, 3, f"${np.mean(tier_caps):,.0f}").alignment = center()
                wr_cell(ws, i, 4, r.get("win_1h_%"))
                wr_cell(ws, i, 5, r.get("win_4h_%"))
                pct_cell(ws, i, 6, r.get("avg_fav_1h_%"))
                pct_cell(ws, i, 7, r.get("avg_fav_4h_%"))
            ws.cell(i, 8, tier_notes[tier]).font = reg_font(C["dim"], 8)
    else:
        ws.cell(3, 1, "Market cap verisi yuklenmedi.").font = reg_font(C["orange"])

    # Ticker bazli tablo
    tbl_start = 7
    ws.merge_cells(f"A{tbl_start}:K{tbl_start}")
    ws.cell(tbl_start, 1, "  TİCKER BAZLI PİYASA DEGERİ VE PERFORMANS").font = bold_font(C["blue"])
    ws.cell(tbl_start, 1).fill = fill(C["l_blue"])
    ws.row_dimensions[tbl_start].height = 20

    header_row(ws, tbl_start+1,
               ["Ticker","Buyukluk","Market Cap ($)",
                "Sinyal\nSayisi","+1sa WR%","+4sa WR%",
                "Ort +1sa\nFav%","Ort +4sa\nFav%","Max +1sa%","Min +1sa%","Kanal(lar)"],
               C["blue"], h=28)

    ticker_wr = win_rate_table(sub, "ticker").sort_values("win_1h_%",
                                                          ascending=False, na_position="last")
    for i, (_, row) in enumerate(ticker_wr.iterrows(), start=tbl_start+2):
        stripe_row(ws, i, 11, i % 2 == 0, height=15)
        tkr = row["ticker"]
        ws.cell(i, 1, tkr).font = bold_font(C["blue"], 9)
        ws.cell(i, 1).alignment = center()

        cap = mcap_data.get(tkr)
        tier = "UNKNOWN"
        tkr_rows = sub[sub["ticker"] == tkr]
        if "cap_tier" in sub.columns and len(tkr_rows):
            tier = tkr_rows["cap_tier"].iloc[0]
        tier_c = {"HIGH": C["blue"], "MID": C["orange"], "LOW": C["red"], "UNKNOWN": C["dim"]}
        tl = ws.cell(i, 2, tier_labels.get(tier, tier))
        tl.alignment = center(); tl.font = bold_font(tier_c.get(tier, C["dim"]), 8)

        if cap:
            ws.cell(i, 3, f"${cap:,.0f}").alignment = center()
        ws.cell(i, 4, int(row["n"])).alignment = center()
        wr_cell(ws, i, 5, row.get("win_1h_%"))
        wr_cell(ws, i, 6, row.get("win_4h_%"))
        pct_cell(ws, i, 7, row.get("avg_fav_1h_%"))
        pct_cell(ws, i, 8, row.get("avg_fav_4h_%"))

        t_data = sub[(sub["ticker"]==tkr) & sub["pct_1h"].notna()]
        if len(t_data):
            mx = ws.cell(i, 9, f"{t_data['pct_1h'].max():+.2f}%")
            mx.alignment = center(); mx.fill = fill(C["l_green"]); mx.font = bold_font(C["dg"], 9)
            mn = ws.cell(i, 10, f"{t_data['pct_1h'].min():+.2f}%")
            mn.alignment = center(); mn.fill = fill(C["l_red"]); mn.font = bold_font(C["dr"], 9)

        chs = sorted(sub[sub["ticker"]==tkr]["channel"].unique())
        ws.cell(i, 11, ", ".join(chs)).font = reg_font(C["dim"], 8)


def write_trend_subset_sheet(wb, df, target: str = "win_1h"):
    """
    Trend Zengin Analiz sayfası.
    Sadece trend_score dolu olan sinyaller için ayrı korelasyon + karar ağacı.
    """
    ws = wb.create_sheet("📈 Trend Zengin Analiz")
    ws.sheet_view.showGridLines = False
    title_row(ws, "📈  TREND ZENGİN ANALİZ — Sadece Trend Verisi Olan Sinyaller", C["green"], 9)

    # trend_score dolu olan sinyaller
    trend_sub = df[
        (df["market"] != "not_found") &
        df["trend_score"].notna() &
        df["win_1h"].notna()
    ].copy()

    total_signals = len(df[df["market"] != "not_found"])
    trend_count = len(trend_sub)

    ws.merge_cells("A2:I2")
    ws.cell(2, 1,
        f"  📌  Trend verisi olan: {trend_count} / {total_signals} sinyal ({trend_count/max(total_signals,1)*100:.0f}%). "
        "Bu alt küme için ayrı korelasyon ve karar ağacı. Trend feature'ları gerçek değerleriyle kullanılıyor (median dolgu yok)."
    ).font = reg_font(C["dim"], 9)
    ws.cell(2, 1).fill = fill("F9FAFB")
    ws.row_dimensions[2].height = 22

    ws.merge_cells("A3:I3")
    ws.cell(3, 1,
        "  ⚠  Raw pre_pct_* (momentum) korelasyonda görünür ancak KARAR AĞACINA GİRMEZ. "
        "Sebep: +%5 hareket LONG için avantajlı, SHORT için dezavantajlıdır — yön bilgisi taşımaz. "
        "Karar ağacında yalnızca pre_pct_*_dir (yönlü versiyon = pre_pct × dir_signal) kullanılır."
    ).font = reg_font(C["orange"], 9)
    ws.cell(3, 1).fill = fill("FFF7ED")
    ws.row_dimensions[3].height = 20

    if trend_count < 30:
        ws.cell(5, 1, f"⚠  Trend verisi olan sinyal sayısı çok az ({trend_count}). "
                      "En az 30 sinyal gerekli.").font = reg_font(C["orange"])
        return

    # ── BÖLÜM 1: Hedef odaklı korelasyon ────────────────────────
    cur_row = 5
    ws.merge_cells(f"A{cur_row}:I{cur_row}")
    ws.cell(cur_row, 1, "  HEDEF ODAKLI KORELASYON — Trend Alt Kümesi").font = bold_font(C["green"])
    ws.cell(cur_row, 1).fill = fill(C["l_green"])
    ws.row_dimensions[cur_row].height = 20
    cur_row += 1

    tcorr = target_correlation(trend_sub)
    if tcorr.empty:
        ws.cell(cur_row, 1, "Yeterli veri yok.").font = reg_font(C["orange"])
        cur_row += 2
    else:
        targets_list = ["win_1h", "win_4h", "fav_1h", "fav_4h"]
        HDR = ["Feature", "Açıklama", "n",
               "win_1h", "win_4h", "fav_1h", "fav_4h",
               "Max |r|", "Yorum"]
        header_row(ws, cur_row, HDR, C["green"], h=26)
        set_col_widths(ws, [24, 28, 8, 10, 10, 10, 10, 10, 32])
        cur_row += 1

        for _, row in tcorr.iterrows():
            stripe_row(ws, cur_row, 9, cur_row % 2 == 0, height=16)
            ws.cell(cur_row, 1, row["feature"]).font = bold_font(C["dark"], 9)
            ws.cell(cur_row, 2, row["label"]).font = reg_font(C["dim"], 8)
            ws.cell(cur_row, 3, int(row["n"])).alignment = center()

            for j, t in enumerate(targets_list, start=4):
                val = row.get(t)
                if val is not None and pd.notna(val):
                    cc = ws.cell(cur_row, j, float(val))
                    cc.number_format = "+0.000;-0.000"
                    cc.alignment = center()
                    av = abs(val)
                    if av >= 0.20:
                        cc.fill = fill(C["l_green"] if val > 0 else C["l_red"])
                        cc.font = bold_font(C["dg"] if val > 0 else C["dr"], 9)
                    elif av >= 0.10:
                        cc.fill = fill("F0FDF4" if val > 0 else "FFF5F5")
                        cc.font = reg_font(C["green"] if val > 0 else C["red"], 9)
                    else:
                        cc.font = reg_font(C["dim"], 9)
                else:
                    ws.cell(cur_row, j, "—").alignment = center()

            mac = row.get("max_abs_corr", 0)
            mc = ws.cell(cur_row, 8, f"{mac:.3f}")
            mc.alignment = center()
            if mac >= 0.20:
                mc.fill = fill(C["l_green"]); mc.font = bold_font(C["dg"], 9)
            elif mac >= 0.10:
                mc.font = reg_font(C["green"], 9)
            else:
                mc.font = reg_font(C["dim"], 9)

            # Kısa yorum
            best_v = 0
            best_t = ""
            for t in targets_list:
                v = row.get(t)
                if v is not None and pd.notna(v) and abs(v) > best_v:
                    best_v = abs(v)
                    best_t = t
            if best_v >= 0.10:
                d = "↑" if row.get(best_t, 0) > 0 else "↓"
                ws.cell(cur_row, 9, f"{best_t} {d} ({best_v:.3f})").font = reg_font(C["dim"], 8)
            else:
                ws.cell(cur_row, 9, "Zayıf").font = reg_font(C["dim"], 8)
            cur_row += 1

        cur_row += 1

    # ── BÖLÜM 2: Karar ağacı (trend alt kümesi) ─────────────────
    ws.merge_cells(f"A{cur_row}:I{cur_row}")
    ws.cell(cur_row, 1, "  KARAR AĞACI — Trend Alt Kümesi").font = bold_font(C["purple"])
    ws.cell(cur_row, 1).fill = fill(C["l_purp"])
    ws.row_dimensions[cur_row].height = 20
    cur_row += 1

    tree_res = decision_tree_analysis(trend_sub, target=target)
    if not tree_res:
        ws.cell(cur_row, 1, "⚠ scikit-learn yüklü değil.").font = reg_font(C["orange"])
        return
    if "error" in tree_res:
        ws.cell(cur_row, 1, f"⚠ {tree_res['error']}").font = reg_font(C["orange"])
        return

    # KPI satırı
    ws.cell(cur_row, 1, f"Örnek: {tree_res['n_samples']}").font = bold_font(C["blue"], 9)
    ws.cell(cur_row, 2, f"Baseline: {tree_res['baseline_wr']}%").font = bold_font(C["orange"], 9)
    ws.cell(cur_row, 3, f"Train: {tree_res['train_acc']}%").font = bold_font(C["green"], 9)
    cv_txt = f"CV: {tree_res['cv_acc']}% ± {tree_res['cv_std']}%" if tree_res["cv_acc"] else "CV: —"
    ws.cell(cur_row, 4, cv_txt).font = bold_font(C["purple"], 9)
    best_d = tree_res.get("best_depth", 3)
    ws.cell(cur_row, 5, f"Derinlik: {best_d}").font = bold_font(C["dark"], 9)
    cur_row += 1

    # Derinlik karşılaştırma
    dr_list = tree_res.get("depth_results", [])
    if dr_list:
        header_row(ws, cur_row, ["Derinlik", "Train", "CV Acc", "", ""], C["purple"], h=20)
        cur_row += 1
        for dr in dr_list:
            stripe_row(ws, cur_row, 5, cur_row % 2 == 0, height=15)
            ws.cell(cur_row, 1, dr["depth"]).alignment = center()
            ws.cell(cur_row, 2, f"{dr['train_acc']}%").alignment = center()
            cv_t = f"{dr['cv_acc']}% ± {dr['cv_std']}%" if dr["cv_acc"] else "—"
            ws.cell(cur_row, 3, cv_t).alignment = center()
            if dr["depth"] == best_d:
                ws.cell(cur_row, 4, "✅").font = bold_font(C["dg"])
            cur_row += 1
        cur_row += 1

    # Feature importance
    ws.merge_cells(f"A{cur_row}:E{cur_row}")
    ws.cell(cur_row, 1, "  ÖZELLİK ÖNEMİ (Trend Alt Kümesi)").font = bold_font(C["purple"])
    ws.cell(cur_row, 1).fill = fill(C["l_purp"])
    cur_row += 1

    header_row(ws, cur_row, ["Özellik", "Önem", "Görsel", "", ""], C["purple"], h=20)
    cur_row += 1
    for feat, score in tree_res["importances"]:
        if score < 0.001:
            continue
        stripe_row(ws, cur_row, 5, cur_row % 2 == 0, height=15)
        ws.cell(cur_row, 1, feat).font = reg_font(C["dark"], 9)
        ws.cell(cur_row, 2, f"{score:.4f}").alignment = center()
        ws.cell(cur_row, 2).font = bold_font(C["purple"], 9)
        bar = int(score * 20)
        ws.cell(cur_row, 3, "█" * bar + "░" * (20 - bar)).font = Font(
            name="Courier New", size=9,
            color=C["purple"] if score > 0.15 else C["dim"])
        cur_row += 1
    cur_row += 1

    # Kurallar
    rules = tree_res.get("rules", [])
    if rules:
        ws.merge_cells(f"A{cur_row}:E{cur_row}")
        ws.cell(cur_row, 1, "  KARAR KURALLARI (Trend Alt Kümesi)").font = bold_font(C["purple"])
        ws.cell(cur_row, 1).fill = fill(C["l_purp"])
        cur_row += 1

        header_row(ws, cur_row, ["Koşullar", "n", "Kazanan", "WR%", "Tahmin"], C["purple"], h=22)
        ws.column_dimensions["A"].width = 55
        cur_row += 1
        for rule in rules:
            stripe_row(ws, cur_row, 5, cur_row % 2 == 0, height=16)
            ws.cell(cur_row, 1, rule["kosullar"]).font = reg_font(size=9)
            ws.cell(cur_row, 2, rule["ornek_sayisi"]).alignment = center()
            ws.cell(cur_row, 3, rule["kazanan"]).alignment = center()
            wr_cell(ws, cur_row, 4, rule["win_rate_%"])
            tc = ws.cell(cur_row, 5, rule["tahmin"]); tc.alignment = center()
            if "✅" in rule["tahmin"]:
                tc.fill = fill(C["l_green"]); tc.font = bold_font(C["dg"], 9)
            else:
                tc.fill = fill(C["l_red"]); tc.font = bold_font(C["dr"], 9)
            cur_row += 1


def write_pump_warning_sheet(wb, df):
    """
    Katman 1 — Pump Uyarı Sayfası (ML'e girmiyor, sadece tanımlayıcı).
    pump_before_4h / pump_before_24h değerleriyle sinyal öncesi pump'ları gösterir.
    """
    ws = wb.create_sheet("⚠ Pump Uyarisi")
    ws.sheet_view.showGridLines = False
    title_row(ws, "⚠  SİNYAL ÖNCESİ PUMP UYARISI — Katman 1 (ML'e Girmiyor)", C["orange"], 11)

    found = df[
        (df["market"] != "not_found") &
        df["pct_1h"].notna() &
        df["win_1h"].notna()
    ].copy()

    has_pump4  = "pump_before_4h"  in found.columns and found["pump_before_4h"].notna().any()
    has_pump24 = "pump_before_24h" in found.columns and found["pump_before_24h"].notna().any()

    if not has_pump4 and not has_pump24:
        ws.cell(3, 1, "⚠  pump_before verileri henüz hesaplanmamış. "
                      "backtest_signals.py --force çalıştırın.").font = reg_font(C["orange"])
        return

    ws.merge_cells("A2:K2")
    ws.cell(2, 1,
        "  📌  pump_before_4h/24h: Sinyal öncesi 4sa/24sa'deki fiyat hareketi. "
        "Yüksek değer = sinyal zaten pump olmuş bir coin için geliyor (geç giriş riski)."
    ).font = reg_font(C["dim"], 9)
    ws.cell(2, 1).fill = fill("F9FAFB")
    ws.row_dimensions[2].height = 20

    # ── KPI'lar ──────────────────────────────────────────────────
    pump4_med  = found["pump_before_4h"].median()  if has_pump4  else None
    pump24_med = found["pump_before_24h"].median() if has_pump24 else None

    # pump_before > %5 olanlar "pump sonrası sinyal"
    if has_pump4:
        high_pump_4h = found[found["pump_before_4h"] > 5]
        normal_4h    = found[found["pump_before_4h"] <= 5]
    if has_pump24:
        high_pump_24h = found[found["pump_before_24h"] > 10]
        normal_24h    = found[found["pump_before_24h"] <= 10]

    kpi_row = 3
    kpis = []
    if has_pump4:
        kpis.append(("A", "Medyan Pump 4sa", f"{pump4_med:+.1f}%", C["blue"]))
        kpis.append(("C", "Pump>5% (4sa)",
                     f"{len(high_pump_4h)}/{len(found)}", C["orange"]))
    if has_pump24:
        kpis.append(("E", "Medyan Pump 24sa", f"{pump24_med:+.1f}%", C["purple"]))
        kpis.append(("G", "Pump>10% (24sa)",
                     f"{len(high_pump_24h)}/{len(found)}", C["red"]))

    for col_l, lbl, val, clr in kpis:
        c2 = chr(ord(col_l) + 1)
        for r in [kpi_row, kpi_row+1, kpi_row+2]:
            ws.merge_cells(f"{col_l}{r}:{c2}{r}")
        ws.cell(kpi_row, ord(col_l)-64, lbl).font = bold_font()
        ws.cell(kpi_row, ord(col_l)-64).fill = fill(clr)
        ws.cell(kpi_row, ord(col_l)-64).alignment = center()
        vc = ws.cell(kpi_row+1, ord(col_l)-64, val)
        vc.font = Font(name="Arial", bold=True, size=18, color=clr)
        vc.fill = fill(C["white"]); vc.alignment = center()
        ws.cell(kpi_row+2, ord(col_l)-64).fill = fill(clr)
        ws.cell(kpi_row+2, ord(col_l)-64+1).fill = fill(clr)
    for r in [kpi_row, kpi_row+1, kpi_row+2]:
        ws.row_dimensions[r].height = 26

    # ── Pump vs Normal WR karşılaştırma ──────────────────────────
    tbl_row = kpi_row + 4
    ws.merge_cells(f"A{tbl_row}:K{tbl_row}")
    ws.cell(tbl_row, 1, "  PUMP ÖNCESİ SİNYAL vs NORMAL SİNYAL — Win Rate Karşılaştırması").font = bold_font(C["orange"])
    ws.cell(tbl_row, 1).fill = fill(C["l_amber"])
    ws.row_dimensions[tbl_row].height = 20

    HDR = ["Kategori", "Sinyal\nSayısı", "+1sa WR%", "+4sa WR%", "+1gün WR%",
           "Ort +1sa\nFav%", "Ort +4sa\nFav%", "Medyan\nPump4h%", "Medyan\nPump24h%"]
    header_row(ws, tbl_row+1, HDR, C["orange"], h=28)
    set_col_widths(ws, [26, 12, 11, 11, 11, 14, 14, 14, 14])

    def _wr(sub, col):
        v = sub[col].dropna()
        return round(v.mean() * 100, 1) if len(v) else None

    def _fav(sub, col):
        v = sub[col].dropna()
        return round(v.mean(), 2) if len(v) else None

    categories = []
    if has_pump4:
        categories.append(("🟢 Normal (pump4h ≤ 5%)", normal_4h))
        categories.append(("🔴 Pump (pump4h > 5%)",   high_pump_4h))
    if has_pump24:
        categories.append(("🟢 Normal (pump24h ≤ 10%)", normal_24h))
        categories.append(("🔴 Pump (pump24h > 10%)",   high_pump_24h))
    categories.append(("📊 Tüm Sinyaller", found))

    for i, (cat_label, sub) in enumerate(categories, start=tbl_row+2):
        stripe_row(ws, i, 9, i % 2 == 0, height=17)
        ws.cell(i, 1, cat_label).font = bold_font(C["dark"], 9)
        ws.cell(i, 2, len(sub)).alignment = center()
        wr_cell(ws, i, 3, _wr(sub, "win_1h"))
        wr_cell(ws, i, 4, _wr(sub, "win_4h"))
        wr_cell(ws, i, 5, _wr(sub, "win_1d"))
        pct_cell(ws, i, 6, _fav(sub, "fav_1h"))
        pct_cell(ws, i, 7, _fav(sub, "fav_4h"))
        if has_pump4:
            med4 = sub["pump_before_4h"].median()
            if pd.notna(med4):
                pct_cell(ws, i, 8, med4)
        if has_pump24:
            med24 = sub["pump_before_24h"].median()
            if pd.notna(med24):
                pct_cell(ws, i, 9, med24)

    # ── Kanal bazlı pump analizi ─────────────────────────────────
    ch_start = tbl_row + len(categories) + 3
    ws.merge_cells(f"A{ch_start}:K{ch_start}")
    ws.cell(ch_start, 1, "  KANAL BAZLI — Ortalama Pump Before Değerleri").font = bold_font(C["orange"])
    ws.cell(ch_start, 1).fill = fill(C["l_amber"])
    ws.row_dimensions[ch_start].height = 20

    header_row(ws, ch_start+1,
               ["Kanal", "Sinyal", "Ort Pump4h%", "Ort Pump24h%",
                "WR%(normal)", "WR%(pump)", "Fark"], C["orange"], h=24)

    channels = sorted(found["channel"].unique())
    for j, ch in enumerate(channels, start=ch_start+2):
        stripe_row(ws, j, 7, j % 2 == 0, height=16)
        ch_data = found[found["channel"] == ch]
        ws.cell(j, 1, f"@{ch}").font = bold_font(C["blue"], 9)
        ws.cell(j, 2, len(ch_data)).alignment = center()

        if has_pump4:
            avg4 = ch_data["pump_before_4h"].mean()
            if pd.notna(avg4):
                pct_cell(ws, j, 3, avg4)
        if has_pump24:
            avg24 = ch_data["pump_before_24h"].mean()
            if pd.notna(avg24):
                pct_cell(ws, j, 4, avg24)

        if has_pump4:
            ch_norm = ch_data[ch_data["pump_before_4h"] <= 5]
            ch_pump = ch_data[ch_data["pump_before_4h"] > 5]
            wr_n = _wr(ch_norm, "win_1h")
            wr_p = _wr(ch_pump, "win_1h")
            wr_cell(ws, j, 5, wr_n)
            wr_cell(ws, j, 6, wr_p)
            if wr_n is not None and wr_p is not None:
                d = round(wr_p - wr_n, 1)
                dc = ws.cell(j, 7, f"{d:+.1f}%")
                dc.alignment = center()
                if d > 0:   dc.fill = fill(C["l_green"]); dc.font = bold_font(C["dg"], 9)
                elif d < 0: dc.fill = fill(C["l_red"]);   dc.font = bold_font(C["dr"], 9)


def write_market_context_sheet(wb, df):
    """
    Piyasa Bağlamı Sayfası — BTC hareketi ve Fear & Greed Index ile win rate ilişkisi.
    """
    ws = wb.create_sheet("🌐 Piyasa Baglami")
    ws.sheet_view.showGridLines = False
    title_row(ws, "🌐  PİYASA BAĞLAMI — BTC + Fear & Greed Index ile Win Rate", C["blue"], 11)

    found = df[
        (df["market"] != "not_found") &
        df["pct_1h"].notna() &
        df["win_1h"].notna()
    ].copy()

    has_btc = "btc_pct_24h" in found.columns and found["btc_pct_24h"].notna().any()
    has_fng = "fear_greed"  in found.columns and found["fear_greed"].notna().any()

    if not has_btc and not has_fng:
        ws.cell(3, 1, "⚠  BTC ve FnG verileri henüz doldurulmamış. "
                      "backtest_signals.py ve backtest_context.py çalıştırın.").font = reg_font(C["orange"])
        return

    ws.merge_cells("A2:K2")
    ws.cell(2, 1,
        "  📌  BTC yönü ve piyasa duygusu (Fear & Greed) ile sinyal performansı ilişkisi."
    ).font = reg_font(C["dim"], 9)
    ws.cell(2, 1).fill = fill("F9FAFB")
    ws.row_dimensions[2].height = 18

    cur_row = 3

    # ── BTC Bağlamı ─────────────────────────────────────────────
    if has_btc:
        ws.merge_cells(f"A{cur_row}:K{cur_row}")
        ws.cell(cur_row, 1, "  BTC 24 SAATLİK DEĞİŞİM — Piyasa Yönü ve Win Rate").font = bold_font(C["blue"])
        ws.cell(cur_row, 1).fill = fill(C["l_blue"])
        ws.row_dimensions[cur_row].height = 20
        cur_row += 1

        HDR = ["BTC Durumu", "Sinyal\nSayısı", "+1sa WR%", "+4sa WR%",
               "Ort +1sa\nFav%", "Medyan\nBTC 24h%"]
        header_row(ws, cur_row, HDR, C["blue"], h=28)
        set_col_widths(ws, [26, 12, 11, 11, 14, 14])
        cur_row += 1

        def _wr_s(sub, col):
            v = sub[col].dropna()
            return round(v.mean() * 100, 1) if len(v) else None

        btc24 = pd.to_numeric(found["btc_pct_24h"], errors="coerce")
        btc_up   = found[btc24 > 1]
        btc_flat = found[(btc24 >= -1) & (btc24 <= 1)]
        btc_down = found[btc24 < -1]

        btc_cats = [
            ("📈 BTC Yükseliş (>+1%)", btc_up),
            ("➡️ BTC Yatay (-1% ~ +1%)", btc_flat),
            ("📉 BTC Düşüş (<-1%)", btc_down),
            ("📊 Tümü", found[btc24.notna()]),
        ]
        for cat_label, sub in btc_cats:
            stripe_row(ws, cur_row, 6, cur_row % 2 == 0, height=17)
            ws.cell(cur_row, 1, cat_label).font = bold_font(C["dark"], 9)
            ws.cell(cur_row, 2, len(sub)).alignment = center()
            wr_cell(ws, cur_row, 3, _wr_s(sub, "win_1h"))
            wr_cell(ws, cur_row, 4, _wr_s(sub, "win_4h"))
            pct_cell(ws, cur_row, 5, sub["fav_1h"].mean() if len(sub) and "fav_1h" in sub else None)
            if len(sub) and "btc_pct_24h" in sub:
                med = sub["btc_pct_24h"].median()
                if pd.notna(med):
                    pct_cell(ws, cur_row, 6, med)
            cur_row += 1
        cur_row += 1

    # ── Fear & Greed Index ──────────────────────────────────────
    if has_fng:
        ws.merge_cells(f"A{cur_row}:K{cur_row}")
        ws.cell(cur_row, 1, "  FEAR & GREED INDEX — Piyasa Duygusu ve Win Rate").font = bold_font(C["green"])
        ws.cell(cur_row, 1).fill = fill(C["l_green"])
        ws.row_dimensions[cur_row].height = 20
        cur_row += 1

        HDR = ["Duygu Durumu", "Sinyal\nSayısı", "+1sa WR%", "+4sa WR%",
               "Ort +1sa\nFav%", "Medyan FnG"]
        header_row(ws, cur_row, HDR, C["green"], h=28)
        cur_row += 1

        fng = pd.to_numeric(found["fear_greed"], errors="coerce")
        ext_fear = found[fng <= 25]
        fear     = found[(fng > 25) & (fng <= 45)]
        neutral  = found[(fng > 45) & (fng <= 55)]
        greed    = found[(fng > 55) & (fng <= 75)]
        ext_grd  = found[fng > 75]

        fng_cats = [
            ("😱 Extreme Fear (0-25)",  ext_fear),
            ("😟 Fear (26-45)",          fear),
            ("😐 Neutral (46-55)",       neutral),
            ("😏 Greed (56-75)",         greed),
            ("🤑 Extreme Greed (76-100)", ext_grd),
            ("📊 Tümü",                  found[fng.notna()]),
        ]
        for cat_label, sub in fng_cats:
            stripe_row(ws, cur_row, 6, cur_row % 2 == 0, height=17)
            ws.cell(cur_row, 1, cat_label).font = bold_font(C["dark"], 9)
            ws.cell(cur_row, 2, len(sub)).alignment = center()
            wr_cell(ws, cur_row, 3, _wr_s(sub, "win_1h") if len(sub) else None)
            wr_cell(ws, cur_row, 4, _wr_s(sub, "win_4h") if len(sub) else None)
            pct_cell(ws, cur_row, 5, sub["fav_1h"].mean() if len(sub) and "fav_1h" in sub else None)
            if len(sub):
                med_fng = sub["fear_greed"].median()
                if pd.notna(med_fng):
                    ws.cell(cur_row, 6, int(med_fng)).alignment = center()
            cur_row += 1
        cur_row += 1

    # ── Channel WR 30d (varsa) ──────────────────────────────────
    if "channel_win_rate_30d" in found.columns and found["channel_win_rate_30d"].notna().any():
        ws.merge_cells(f"A{cur_row}:K{cur_row}")
        ws.cell(cur_row, 1, "  KANAL GÜVENİLİRLİK — Son 30 Gün Kümülatif Win Rate").font = bold_font(C["purple"])
        ws.cell(cur_row, 1).fill = fill(C["l_purp"])
        ws.row_dimensions[cur_row].height = 20
        cur_row += 1

        HDR = ["Kanal Güvenilirlik", "Sinyal\nSayısı", "+1sa WR%", "+4sa WR%",
               "Ort +1sa\nFav%", "Medyan\nKanal WR30d"]
        header_row(ws, cur_row, HDR, C["purple"], h=28)
        cur_row += 1

        cwr = pd.to_numeric(found["channel_win_rate_30d"], errors="coerce")
        high_rel  = found[cwr > 55]
        mid_rel   = found[(cwr >= 45) & (cwr <= 55)]
        low_rel   = found[cwr < 45]

        rel_cats = [
            ("🟢 Güvenilir Kanal (WR>55%)", high_rel),
            ("🟡 Orta Kanal (WR 45-55%)",   mid_rel),
            ("🔴 Düşük Kanal (WR<45%)",     low_rel),
            ("📊 Tümü",                     found[cwr.notna()]),
        ]
        for cat_label, sub in rel_cats:
            stripe_row(ws, cur_row, 6, cur_row % 2 == 0, height=17)
            ws.cell(cur_row, 1, cat_label).font = bold_font(C["dark"], 9)
            ws.cell(cur_row, 2, len(sub)).alignment = center()
            wr_cell(ws, cur_row, 3, _wr_s(sub, "win_1h") if len(sub) else None)
            wr_cell(ws, cur_row, 4, _wr_s(sub, "win_4h") if len(sub) else None)
            pct_cell(ws, cur_row, 5, sub["fav_1h"].mean() if len(sub) and "fav_1h" in sub else None)
            if len(sub):
                med = sub["channel_win_rate_30d"].median()
                if pd.notna(med):
                    ws.cell(cur_row, 6, f"{med:.1f}%").alignment = center()
            cur_row += 1


# ─────────────────────────────────────────────────────────────────
# RUN (programatik cagri icin)
# ─────────────────────────────────────────────────────────────────

def run(btdb: str = BT_DB_PATH, out: str = OUT_PATH,
        no_volume: bool = False, target: str = "win_1h"):
    """ML analizini calistir ve Excel uret."""

    print("=" * 55)
    print("  SİNYAL ML ANALİZ MOTORU")
    print("=" * 55)

    print("\n[1/5] Backtest verisi yukleniyor...")
    df = load_backtest(btdb)
    print(f"      → {len(df)} satir")
    if len(df) == 0:
        print("⚠  Veri yok. Once backtest_signals.py calistir.")
        return

    df = enrich_features(df)
    found = df[df["market"] != "not_found"]
    print(f"      → Fiyat verisi olan: {len(found)} sinyal")

    mcap_data = {}
    if not no_volume:
        print("\n[2/5] Market cap verisi cekiliyor (CoinGecko)...")
        unique_tickers = found["ticker"].unique().tolist()
        mcap_data = fetch_marketcap_data(unique_tickers)
        df["market_cap_usd"] = df["ticker"].map(mcap_data)
        df = assign_cap_tier(df)
        found = df[df["market"] != "not_found"]
    else:
        print("\n[2/5] Market cap verisi atlandi (--no-volume)")
        df["market_cap_usd"] = np.nan
        df["cap_tier"] = "UNKNOWN"

    print("\n[3/5] Korelasyon matrisi hesaplaniyor...")
    corr = correlation_matrix(df)
    print(f"      → {corr.shape[0]}×{corr.shape[1]} matris")

    print(f"\n[4/6] Karar agaci egitiliyor (hedef: {target})...")
    tree_res = decision_tree_analysis(found, target=target)
    if tree_res and "error" not in tree_res:
        print(f"      → Train: {tree_res['train_acc']}%  "
              f"CV: {tree_res['cv_acc']}%  "
              f"Baseline: {tree_res['baseline_wr']}%  "
              f"Derinlik: {tree_res['best_depth']}")
        print(f"      → En onemli ozellik: {tree_res['importances'][0][0]} "
              f"({tree_res['importances'][0][1]:.4f})")
    elif tree_res and tree_res.get("error"):
        print(f"      ⚠ {tree_res['error']}")

    print(f"\n[5/6] Random Forest egitiliyor (hedef: {target})...")
    rf_res = random_forest_analysis(found, target=target)
    if rf_res and "error" not in rf_res:
        oob_txt = f"{rf_res['oob_score']}%" if rf_res.get("oob_score") is not None else "—"
        print(f"      → Train: {rf_res['train_acc']}%  "
              f"CV: {rf_res['cv_acc']}%  "
              f"OOB: {oob_txt}  "
              f"Agac: {rf_res['best_n_estimators']}  "
              f"min_leaf: {rf_res['best_min_leaf']}")
        print(f"      → En onemli ozellik: {rf_res['importances'][0][0]} "
              f"({rf_res['importances'][0][1]:.4f})")
    elif rf_res and rf_res.get("error"):
        print(f"      ⚠ {rf_res['error']}")

    print(f"\n[6/6] Excel yaziliyor: {out}")
    wb = Workbook()
    write_overview_sheet(wb, df, mcap_data)
    write_channel_sheet(wb, df)
    write_feature_sheet(wb, df)
    write_signal_quality_sheet(wb, df)
    write_hourly_sheet(wb, df)
    write_correlation_sheet(wb, df)
    write_tree_sheet(wb, df, tree_res)
    write_rf_sheet(wb, df, rf_res)
    write_marketcap_sheet(wb, df, mcap_data)
    write_trend_subset_sheet(wb, df, target=target)
    write_pump_warning_sheet(wb, df)
    write_market_context_sheet(wb, df)
    wb.save(out)
    print(f"      ✅ Kaydedildi: {out}")
    print("\n  Sayfalar:")
    print("  📊 Genel Bakis        — KPI ozeti ve periyot win rate'leri")
    print("  🏆 Kanal Performansi  — Kanal siralamasi ve skor")
    print("  🔬 Feature Lab        — Telegram oncesi indikator ayrismasi")
    print("  🎯 Sinyal Kalitesi    — Taker buy + lag filtresi vs tum sinyaller")
    print("  🕐 Saat & Gun         — Zamana gore win rate analizi")
    print("  📐 Korelasyon         — Degiskenler arasi iliskiler")
    print("  🌳 Karar Agaci        — Yorumlanabilir kurallar (tek agac, derinlik 3-5)")
    print("  🌲 Random Forest      — Topluluk modeli (overfitting azaltilmis, OOB dogrulama)")
    print("  💰 Market Cap Etkisi  — Piyasa degeri ve performans iliskisi")
    print("  📈 Trend Zengin Analiz — Trend verisi olan alt kume icin ayri korelasyon+agac")
    print("  ⚠  Pump Uyarisi       — Sinyal oncesi pump bilgilendirmesi (Katman 1)")
    print("  🌐 Piyasa Baglami     — BTC + Fear & Greed + Kanal Guvenilirligi")


# ─────────────────────────────────────────────────────────────────
# CLI
# ─────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Sinyal ML analiz motoru")
    parser.add_argument("--btdb",      default=BT_DB_PATH)
    parser.add_argument("--out",       default=OUT_PATH)
    parser.add_argument("--no-volume", action="store_true")
    parser.add_argument("--target",    default="win_1h",
                        choices=["win_1h", "win_4h", "win_1d"])
    args = parser.parse_args()
    run(btdb=str(Path(args.btdb).resolve()),
        out=str(Path(args.out).resolve()),
        no_volume=args.no_volume,
        target=args.target)


if __name__ == "__main__":
    main()
