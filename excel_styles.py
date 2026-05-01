"""
Ortak Excel Stil Modulu
========================
Tum rapor/backtest/ml scriptleri tarafindan paylasilan
renk paleti, font, fill, alignment ve yardimci fonksiyonlar.
"""

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────
# RENK PALETİ
# ─────────────────────────────────────────────────────────────────

COLORS = {
    "dark":    "1F2937",
    "green":   "059669",
    "blue":    "1D4ED8",
    "orange":  "B45309",
    "purple":  "7C3AED",
    "red":     "991B1B",
    "white":   "FFFFFF",
    "gray":    "F3F4F6",
    "mid":     "E5E7EB",
    "l_green": "D1FAE5",
    "l_red":   "FEE2E2",
    "l_blue":  "DBEAFE",
    "l_amber": "FEF3C7",
    "l_purp":  "EDE9FE",
    "dg":      "065F46",   # deep green
    "dr":      "7F1D1D",   # deep red
    "dim":     "6B7280",
    "text":    "374151",
}

C = COLORS   # kisa alias — mevcut kodla uyumlu


# ─────────────────────────────────────────────────────────────────
# TEMEL STİL FONKSİYONLARI
# ─────────────────────────────────────────────────────────────────

def fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def bold_font(color: str = "FFFFFF", size: int = 10) -> Font:
    return Font(name="Arial", bold=True, color=color, size=size)


def reg_font(color: str = "111827", size: int = 9) -> Font:
    return Font(name="Arial", bold=False, color=color, size=size)


def center() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def left() -> Alignment:
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


def thin_border() -> Border:
    s = Side(style="thin", color="D1D5DB")
    return Border(left=s, right=s, top=s, bottom=s)


# ─────────────────────────────────────────────────────────────────
# KISA ALİASLAR  (mevcut kod uyumluluğu icin)
# ─────────────────────────────────────────────────────────────────

_fill = fill
_bf   = bold_font
_rf   = reg_font
_cen  = center
_lft  = left
_bdr  = thin_border


# ─────────────────────────────────────────────────────────────────
# YÜKSEK SEVİYE YARDIMCILAR
# ─────────────────────────────────────────────────────────────────

def title_row(ws, text: str, bg: str, col_count: int,
              row: int = 1, font_size: int = 13, height: int = 34):
    """Sayfa basligini birlestirilmis hucreye yazar."""
    ws.merge_cells(f"A{row}:{get_column_letter(col_count)}{row}")
    c = ws.cell(row, 1, text)
    c.font = Font(name="Arial", bold=True, size=font_size, color="FFFFFF")
    c.fill = fill(bg)
    c.alignment = center()
    ws.row_dimensions[row].height = height


def header_row(ws, row: int, labels: list, bg: str, h: int = 24):
    """Baslik satirini renkli olarak yazar."""
    for col, lbl in enumerate(labels, 1):
        c = ws.cell(row, col, lbl)
        c.font = bold_font()
        c.fill = fill(bg)
        c.alignment = center()
        c.border = thin_border()
    ws.row_dimensions[row].height = h


def stripe_row(ws, row: int, col_count: int,
               alt: bool = False, height: int = 14):
    """Zebra satiri uygular (alternatif gri/beyaz)."""
    bg = C["gray"] if alt else C["white"]
    for col in range(1, col_count + 1):
        cell = ws.cell(row, col)
        cell.fill = fill(bg)
        cell.border = thin_border()
        cell.font = reg_font()
        cell.alignment = left()
    ws.row_dimensions[row].height = height


def set_col_widths(ws, widths: list):
    """Sutun genisliklerini sirayla ayarlar."""
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def wr_cell(ws, row: int, col: int, wr_value):
    """Win rate hucresini yuzdeyle renklendirir."""
    if wr_value is None:
        ws.cell(row, col, "—").alignment = center()
        return
    c = ws.cell(row, col, f"{wr_value:.1f}%")
    c.alignment = center()
    if wr_value >= 65:
        c.fill = fill(C["l_green"]); c.font = bold_font(C["dg"], 9)
    elif wr_value >= 50:
        c.fill = fill("ECFDF5"); c.font = reg_font(C["green"], 9)
    elif wr_value >= 40:
        c.fill = fill(C["l_amber"]); c.font = reg_font(C["orange"], 9)
    else:
        c.fill = fill(C["l_red"]); c.font = bold_font(C["dr"], 9)


def pct_cell(ws, row: int, col: int, value, direction: str = None):
    """
    Yuzde degisim hucresini renklendirir.
    direction verilirse yon bazli (LONG/SHORT),
    verilmezse mutlak deger bazli renklendirir.
    """
    import numpy as np
    if value is None or (isinstance(value, float) and np.isnan(value)):
        ws.cell(row, col, "—").alignment = center()
        return

    c = ws.cell(row, col, f"{value:+.2f}%")
    c.alignment = center()
    c.font = reg_font(size=9)

    if direction:
        # Yon bazli: LONG -> pozitif=yesil, SHORT -> negatif=yesil
        favorable = (value > 0 and direction == "LONG") or \
                    (value < 0 and direction == "SHORT")
        unfavorable = (value < 0 and direction == "LONG") or \
                      (value > 0 and direction == "SHORT")
        if favorable:
            c.fill = fill(C["l_green"])
            c.font = bold_font(C["dg"], 9)
        elif unfavorable:
            c.fill = fill(C["l_red"])
            c.font = bold_font(C["dr"], 9)
        else:
            c.fill = fill(C["l_amber"])
    else:
        # Mutlak deger bazli
        if value > 1:
            c.fill = fill(C["l_green"]); c.font = bold_font(C["dg"], 9)
        elif value > 0:
            c.fill = fill("ECFDF5"); c.font = reg_font(C["green"], 9)
        elif value > -1:
            c.fill = fill(C["l_amber"]); c.font = reg_font(C["orange"], 9)
        else:
            c.fill = fill(C["l_red"]); c.font = bold_font(C["dr"], 9)
